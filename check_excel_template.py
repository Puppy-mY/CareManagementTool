"""
Excelテンプレートの内容を確認するスクリプト
"""
import openpyxl
import json

# Excelファイルを開く
file_path = r"E:\CareManagementTool\templates\forms\LTC_Certification_Change_R7.11-.xlsx"
wb = openpyxl.load_workbook(file_path)

print("=" * 80)
print("介護保険 要介護・要支援 [認定区分変更]申請書 - テンプレート分析")
print("=" * 80)

# シート一覧
print("\n【シート一覧】")
for sheet_name in wb.sheetnames:
    print(f"  - {sheet_name}")

# メインシート（最初のシート）を取得
sheet = wb.active
print(f"\n【アクティブシート】: {sheet.title}")
print(f"最大行: {sheet.max_row}, 最大列: {sheet.max_column}")

# セルの内容を確認（最初の30行、30列）
print("\n【セル内容サンプル（A1:AD30）】")
print("-" * 80)

for row in range(1, min(31, sheet.max_row + 1)):
    row_data = []
    for col in range(1, min(31, sheet.max_column + 1)):
        cell = sheet.cell(row=row, column=col)
        if cell.value:
            col_letter = openpyxl.utils.get_column_letter(col)
            row_data.append(f"{col_letter}{row}: {cell.value}")

    if row_data:
        print(f"行{row:2d}: {' | '.join(row_data)}")

# 特定のキーワードを含むセルを検索
print("\n【キーワード検索結果】")
print("-" * 80)

keywords = [
    "氏名", "フリガナ", "ふりがな", "生年月日", "性別",
    "住所", "郵便番号", "電話", "被保険者",
    "事業所", "申請", "認定", "主治医", "医療機関",
    "調査", "理由", "有効期間"
]

for keyword in keywords:
    print(f"\n[{keyword}]を含むセル:")
    found = False
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str) and keyword in cell.value:
                col_letter = openpyxl.utils.get_column_letter(cell.column)
                print(f"  {col_letter}{cell.row}: {cell.value}")
                found = True
    if not found:
        print(f"  (見つかりませんでした)")

print("\n" + "=" * 80)
print("分析完了")
print("=" * 80)

wb.close()
