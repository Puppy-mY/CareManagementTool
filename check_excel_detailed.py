"""
Excelテンプレートの詳細分析
"""
import openpyxl

# Excelファイルを開く
file_path = r"E:\CareManagementTool\templates\forms\LTC_Certification_Change_R7.11-.xlsx"
wb = openpyxl.load_workbook(file_path)
sheet = wb.active

print("=" * 100)
print("介護保険 要介護・要支援 [認定区分変更]申請書 - 詳細分析")
print("=" * 100)

# セクションごとに分析
sections = [
    ("申請者情報エリア", 6, 11, 1, 30),
    ("被保険者情報エリア", 12, 20, 1, 30),
    ("前回認定・申請理由エリア", 21, 23, 1, 30),
    ("医療保険情報エリア", 25, 27, 1, 30),
    ("認定調査情報エリア", 29, 35, 1, 30),
    ("主治医情報エリア", 37, 42, 1, 30),
    ("同意・署名エリア", 44, 52, 1, 30),
]

for section_name, start_row, end_row, start_col, end_col in sections:
    print(f"\n【{section_name}】（行{start_row}～{end_row}）")
    print("-" * 100)

    for row in range(start_row, end_row + 1):
        cells_with_value = []
        for col in range(start_col, end_col + 1):
            cell = sheet.cell(row=row, column=col)
            if cell.value:
                col_letter = openpyxl.utils.get_column_letter(col)
                # セルの値を最大50文字に制限
                value_str = str(cell.value)[:50]
                cells_with_value.append(f"{col_letter}{row}: {value_str}")

        if cells_with_value:
            print(f"  行{row:2d}: {' | '.join(cells_with_value)}")

# 入力フィールドの可能性があるセルを探す（値が空白のセル）
print("\n" + "=" * 100)
print("【入力フィールド候補（空白セルの分析）】")
print("=" * 100)

# 特定の範囲で結合セルを探す
print("\n結合セル一覧:")
for merged_range in sheet.merged_cells.ranges:
    print(f"  {merged_range}")

wb.close()
