from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.contrib import messages
from django.core.paginator import Paginator
from django.db.models import Q
from django.http import HttpResponse, JsonResponse, HttpResponseRedirect
from django.urls import reverse
from django.conf import settings
from datetime import date
import json
import os
import urllib.parse
from .models import Assessment
from clients.models import Client
from .forms import DetailedAssessmentForm

# openpyxlの利用可能性チェック
try:
    from openpyxl import load_workbook
    from openpyxl.cell.cell import MergedCell

    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


@login_required
def assessment_list(request):
    # 全アセスメントを取得（クライアントサイドでフィルタリング）
    assessments = Assessment.objects.select_related("client", "assessor").order_by(
        "-assessment_date", "-id"
    )

    # 作成者一覧を取得（アセスメントを作成したことがあるユーザーのみ）
    all_assessors = User.objects.filter(
        id__in=Assessment.objects.values_list("assessor_id", flat=True).distinct()
    ).order_by("username")

    context = {
        "assessments": assessments,
        "all_assessors": all_assessors,
    }

    return render(request, "assessments/assessment_list.html", context)


@login_required
def my_assessment_list(request):
    # 自分が担当する利用者のアセスメントのみ取得
    assessments = Assessment.objects.select_related("client", "assessor").filter(
        client__created_by=request.user
    ).order_by("-assessment_date", "-id")

    # 作成者一覧を取得（絞り込み後のアセスメントに関わるユーザーのみ）
    all_assessors = User.objects.filter(
        id__in=assessments.values_list("assessor_id", flat=True).distinct()
    ).order_by("username")

    context = {
        "assessments": assessments,
        "all_assessors": all_assessors,
    }

    return render(request, "assessments/my_assessment_list.html", context)


@login_required
def assessment_detail(request, pk):
    assessment = get_object_or_404(Assessment, pk=pk)

    # 同じ利用者の他のアセスメント履歴を取得
    assessment_history = (
        Assessment.objects.filter(client=assessment.client)
        .exclude(pk=pk)
        .order_by("-assessment_date")
    )

    context = {
        "assessment": assessment,
        "assessment_history": assessment_history,
    }

    return render(request, "assessments/assessment_detail.html", context)


@login_required
def assessment_edit(request, pk):
    assessment = get_object_or_404(Assessment, pk=pk)

    if request.method == "POST":
        form = DetailedAssessmentForm(request.POST)

        if form.is_valid():
            # actionパラメータを取得
            action = request.POST.get("action", "save")

            # 詳細データをJSONフィールドに保存
            detailed_data = {}

            # 各セクションのデータを整理
            detailed_data["basic_info"] = {
                "life_history": form.cleaned_data.get("life_history", ""),
                "personal_hopes": form.cleaned_data.get("personal_hopes", ""),
                "family_hopes": form.cleaned_data.get("family_hopes", ""),
                "special_situation_status": request.POST.get(
                    "special_situation_status", ""
                ),
                "special_situation": request.POST.getlist("special_situation"),
                "special_situation_other": request.POST.get(
                    "special_situation_other", ""
                ),
            }

            detailed_data["insurance_info"] = {
                "care_level": request.POST.get("care_level", ""),
                "dementia_level": request.POST.get("dementia_level", ""),
                "burden_ratio": form.cleaned_data.get("burden_ratio", ""),
                "medical_insurance": form.cleaned_data.get("medical_insurance", ""),
                "disability_handbook": form.cleaned_data.get(
                    "disability_handbook", False
                ),
                "difficulty_certificate": form.cleaned_data.get(
                    "difficulty_certificate", False
                ),
                "life_protection": form.cleaned_data.get("life_protection", False),
            }

            detailed_data["family_situation"] = {
                "family_member1": form.cleaned_data.get("family_member1", ""),
                "family_member2": form.cleaned_data.get("family_member2", ""),
            }

            detailed_data["living_situation"] = {
                "home_environment": request.POST.get("home_environment", ""),
                "other_home_environment_detail_text": request.POST.get(
                    "other_home_environment_detail_text", ""
                ),
                "housing_type": request.POST.get("housing_type", ""),
                "other_housing_type_detail_text": request.POST.get(
                    "other_housing_type_detail_text", ""
                ),
                "has_elevator": request.POST.get("has_elevator", ""),
                "housing_ownership": request.POST.get("housing_ownership", ""),
                "other_ownership_detail_text": request.POST.get(
                    "other_ownership_detail_text", ""
                ),
                "private_room": request.POST.get("private_room", ""),
                "air_conditioning": request.POST.get("air_conditioning", ""),
                "toilet_type": request.POST.get("toilet_type", ""),
                "toilet_handrail": request.POST.get("toilet_handrail", ""),
                "toilet_step": request.POST.get("toilet_step", ""),
                "bathroom": request.POST.get("bathroom", ""),
                "bathroom_handrail": request.POST.get("bathroom_handrail", ""),
                "bathroom_step": request.POST.get("bathroom_step", ""),
                "sleeping_arrangement": request.POST.get("sleeping_arrangement", ""),
                "other_sleeping_detail_text": request.POST.get(
                    "other_sleeping_detail_text", ""
                ),
                "room_level_difference": request.POST.get("room_level_difference", ""),
                "housing_modification": request.POST.get("housing_modification", ""),
                "housing_modification_need": request.POST.get(
                    "housing_modification_need", ""
                ),
                "housing_notes": request.POST.get("housing_notes", ""),
            }

            detailed_data["services"] = {
                "care_services": form.cleaned_data.get("care_services", []),
                "welfare_equipment": form.cleaned_data.get("welfare_equipment", []),
                "other_services": form.cleaned_data.get("other_services", ""),
                "informal_services": form.cleaned_data.get("informal_services", []),
                "other_services_detail_text": form.cleaned_data.get(
                    "other_services_detail_text", ""
                ),
                "social_relationships": form.cleaned_data.get(
                    "social_relationships", ""
                ),
            }

            detailed_data["health_status"] = {
                "disease_1": form.cleaned_data.get("disease_1", ""),
                "disease_2": form.cleaned_data.get("disease_2", ""),
                "disease_3": form.cleaned_data.get("disease_3", ""),
                "medical_history": form.cleaned_data.get("medical_history", ""),
                "health_basic_info_details": form.cleaned_data.get(
                    "health_basic_info_details", ""
                ),
                "main_doctor_hospital": form.cleaned_data.get(
                    "main_doctor_hospital", ""
                ),
                "main_doctor_name": form.cleaned_data.get("main_doctor_name", ""),
                "visiting_doctor_hospital": form.cleaned_data.get(
                    "visiting_doctor_hospital", ""
                ),
                "visiting_doctor_name": form.cleaned_data.get(
                    "visiting_doctor_name", ""
                ),
                "medication_status": form.cleaned_data.get("medication_status", ""),
                "medication_content": form.cleaned_data.get("medication_content", ""),
                "allergy_details": form.cleaned_data.get("allergy_details", ""),
                "main_doctor": form.cleaned_data.get("main_doctor", ""),
                "outpatient_clinic": form.cleaned_data.get("outpatient_clinic", ""),
                "medication": form.cleaned_data.get("medication", ""),
                "hospital": form.cleaned_data.get("hospital", ""),
                "medication_details": form.cleaned_data.get("medication_details", ""),
                "allergies": form.cleaned_data.get("allergies", False),
            }

            detailed_data["physical_status"] = {
                "skin_condition": form.cleaned_data.get("skin_condition", []),
                "infection_status": form.cleaned_data.get("infection_status", []),
                "special_treatment": form.cleaned_data.get("special_treatment", []),
                "pain_location": form.cleaned_data.get("pain_location", ""),
                "pain_existence": form.cleaned_data.get("pain_existence", ""),
                "height": (
                    str(form.cleaned_data.get("height", ""))
                    if form.cleaned_data.get("height")
                    else ""
                ),
                "weight": (
                    str(form.cleaned_data.get("weight", ""))
                    if form.cleaned_data.get("weight")
                    else ""
                ),
                "vision": form.cleaned_data.get("vision", ""),
                "hearing": form.cleaned_data.get("hearing", ""),
                "physical_notes": form.cleaned_data.get("physical_notes", ""),
            }

            detailed_data["cognitive_function"] = {
                "dementia_presence": form.cleaned_data.get("dementia_presence", ""),
                "dementia_severity": form.cleaned_data.get("dementia_severity", ""),
                "dementia_details": form.cleaned_data.get("dementia_details", ""),
                "bpsd_presence": form.cleaned_data.get("bpsd_presence", ""),
                "bpsd_symptoms": form.cleaned_data.get("bpsd_symptoms", []),
                "bpsd_details": form.cleaned_data.get("bpsd_details", ""),
                "conversation": form.cleaned_data.get("conversation", ""),
                "communication": form.cleaned_data.get("communication", ""),
                "cognitive_notes": form.cleaned_data.get("cognitive_notes", ""),
            }

            detailed_data["basic_activities"] = {
                "turning_over": form.cleaned_data.get("turning_over", ""),
                "getting_up": form.cleaned_data.get("getting_up", ""),
                "sitting": form.cleaned_data.get("sitting", ""),
                "standing_up": form.cleaned_data.get("standing_up", ""),
                "standing": form.cleaned_data.get("standing", ""),
                "transfer": form.cleaned_data.get("transfer", ""),
                "indoor_mobility": form.cleaned_data.get("indoor_mobility", ""),
                "outdoor_mobility": form.cleaned_data.get("outdoor_mobility", ""),
                "indoor_mobility_equipment": form.cleaned_data.get(
                    "indoor_mobility_equipment", ""
                ),
                "outdoor_mobility_equipment": form.cleaned_data.get(
                    "outdoor_mobility_equipment", ""
                ),
                "basic_activity_notes": form.cleaned_data.get(
                    "basic_activity_notes", ""
                ),
            }

            detailed_data["adl"] = {
                "eating_method": form.cleaned_data.get("eating_method", ""),
                "eating_assistance": form.cleaned_data.get("eating_assistance", ""),
                "swallowing": form.cleaned_data.get("swallowing", ""),
                "meal_form_main": form.cleaned_data.get("meal_form_main", ""),
                "meal_form_side": form.cleaned_data.get("meal_form_side", ""),
                "water_thickening": form.cleaned_data.get("water_thickening", False),
                "eating_restriction": form.cleaned_data.get("eating_restriction", ""),
                "eating_utensils": form.cleaned_data.get("eating_utensils", []),
                "eating_notes": form.cleaned_data.get("eating_notes", ""),
                "oral_hygiene_assistance": form.cleaned_data.get(
                    "oral_hygiene_assistance", ""
                ),
                "natural_teeth_presence": form.cleaned_data.get(
                    "natural_teeth_presence", ""
                ),
                "denture_presence": form.cleaned_data.get("denture_presence", ""),
                "denture_type": form.cleaned_data.get("denture_type", ""),
                "denture_location": form.cleaned_data.get("denture_location", ""),
                "tooth_decay": form.cleaned_data.get("tooth_decay", False),
                "oral_notes": form.cleaned_data.get("oral_notes", ""),
                "bathing_assistance": form.cleaned_data.get("bathing_assistance", ""),
                "bathing_form": form.cleaned_data.get("bathing_form", ""),
                "bathing_restriction": form.cleaned_data.get(
                    "bathing_restriction", False
                ),
                "dressing_upper": form.cleaned_data.get("dressing_upper", ""),
                "dressing_lower": form.cleaned_data.get("dressing_lower", ""),
                "bathing_notes": form.cleaned_data.get("bathing_notes", ""),
                "excretion_assistance": form.cleaned_data.get(
                    "excretion_assistance", ""
                ),
                "urination": form.cleaned_data.get("urination", False),
                "urinary_incontinence": form.cleaned_data.get(
                    "urinary_incontinence", False
                ),
                "defecation": form.cleaned_data.get("defecation", False),
                "fecal_incontinence": form.cleaned_data.get(
                    "fecal_incontinence", False
                ),
                "daytime_location": form.cleaned_data.get("daytime_location", []),
                "daytime_method": form.cleaned_data.get("daytime_method", []),
                "nighttime_location": form.cleaned_data.get("nighttime_location", []),
                "nighttime_method": form.cleaned_data.get("nighttime_method", []),
                "excretion_supplies": form.cleaned_data.get("excretion_supplies", []),
                "excretion_notes": form.cleaned_data.get("excretion_notes", ""),
            }

            detailed_data["iadl"] = {
                "cooking": form.cleaned_data.get("cooking", ""),
                "cleaning": form.cleaned_data.get("cleaning", ""),
                "washing": form.cleaned_data.get("washing", ""),
                "shopping": form.cleaned_data.get("shopping", ""),
                "money_management": form.cleaned_data.get("money_management", ""),
                "iadl_notes": form.cleaned_data.get("iadl_notes", ""),
            }

            # 基本情報を更新
            assessment.assessment_type = request.POST.get(
                "assessment_type", assessment.assessment_type
            )
            assessment.interview_location = request.POST.get(
                "interview_location", assessment.interview_location
            )

            # 作成日を更新
            created_date = request.POST.get("created_date")
            if created_date:
                from datetime import datetime

                try:
                    assessment.assessment_date = datetime.strptime(
                        created_date, "%Y-%m-%d"
                    ).date()
                except ValueError:
                    pass

            # その他フィールドの処理
            assessment_type_other = request.POST.get("assessment_type_other", "")
            interview_location_other = request.POST.get("interview_location_other", "")

            # その他フィールドをモデルフィールドに保存
            assessment.assessment_type_other = assessment_type_other
            assessment.interview_location_other = interview_location_other

            # その他フィールドをbasic_infoに追加
            detailed_data["basic_info"]["assessment_type_other"] = assessment_type_other
            detailed_data["basic_info"][
                "interview_location_other"
            ] = interview_location_other

            # 健康状態データに疾患名・発症日を追加
            for i in range(1, 7):
                detailed_data["health_status"][f"disease_name_{i}"] = request.POST.get(
                    f"disease_name_{i}", ""
                )
                detailed_data["health_status"][f"onset_date_{i}"] = request.POST.get(
                    f"onset_date_{i}", ""
                )

            # 特別な医療処置を追加
            detailed_data["health_status"]["special_medical_treatment_status"] = (
                request.POST.get("special_medical_treatment_status", "")
            )
            detailed_data["health_status"]["special_medical_treatment"] = (
                request.POST.getlist("special_medical_treatment")
            )

            # 基本情報の詳細を追加
            detailed_data["health_status"]["health_basic_info_details"] = (
                request.POST.get("health_basic_info_details", "")
            )

            # フォームで定義されていないが、テンプレートで使用されているフィールドを直接POSTから取得
            additional_fields = [
                "main_doctor_hospital",
                "main_doctor_name",
                "visiting_doctor_hospital",
                "visiting_doctor_name",
                "family_doctor_hospital_1",
                "family_doctor_name_1",
                "family_doctor_hospital_2",
                "family_doctor_name_2",
                "family_doctor_hospital_3",
                "family_doctor_name_3",
                "family_doctor_hospital_4",
                "family_doctor_name_4",
                "hospital_visit_status",
                "hospital_visit_method",
                "hospital_visit_method_other",
                "medical_institution_notes",
                "medication_status",
                "medication_content",
                "allergy_details",
                "medication_allergy_notes",
                "has_sleeping_medication",
                "has_laxative",
                "has_allergy",
                "has_skin_disease",
                "skin_disease_other",
                "infection_presence",
                "infection_other_text",
                "skin_infection_details",
                "has_paralysis",
                "has_pain",
                "uses_hearing_aid",
                "uses_glasses",
                "sensory_function_details",
            ]

            for field in additional_fields:
                if field in request.POST:
                    detailed_data["health_status"][field] = request.POST.get(field, "")

            # physical_status関連フィールドをphysical_statusに移動
            if "physical_status_notes" in request.POST:
                detailed_data["physical_status"]["notes"] = request.POST.get(
                    "physical_status_notes", ""
                )
            if "paralysis_pain_details" in request.POST:
                detailed_data["physical_status"]["paralysis_pain_details"] = (
                    request.POST.get("paralysis_pain_details", "")
                )
            if "skin_infection_details" in request.POST:
                detailed_data["physical_status"]["skin_infection_details"] = (
                    request.POST.get("skin_infection_details", "")
                )
            if "paralysis_location" in request.POST:
                detailed_data["physical_status"]["paralysis_location"] = (
                    request.POST.get("paralysis_location", "")
                )
            if "pain_location" in request.POST:
                detailed_data["physical_status"]["pain_location"] = request.POST.get(
                    "pain_location", ""
                )

            # MultipleChoiceFieldの特別処理
            detailed_data["health_status"]["skin_disease"] = request.POST.getlist(
                "skin_disease"
            )
            detailed_data["health_status"]["infection"] = request.POST.getlist(
                "infection"
            )

            # その他のセクションのフィールドも追加
            # 基本動作関連
            basic_activity_additional = [
                "indoor_mobility_equipment_other",
                "outdoor_mobility_equipment_other",
            ]
            for field in basic_activity_additional:
                if field in request.POST:
                    detailed_data["basic_activities"][field] = request.POST.get(
                        field, ""
                    )

            # ADL関連
            adl_additional = [
                "water_thickening_level",
                "eating_restriction_detail",
                "eating_utensils_other_detail",
                "bathing_restriction_detail",
                "urinary_incontinence_frequency",
                "fecal_incontinence_frequency",
            ]
            for field in adl_additional:
                if field in request.POST:
                    detailed_data["adl"][field] = request.POST.get(field, "")

            # 認知機能関連
            cognitive_additional = ["dementia_details", "bpsd_details"]
            for field in cognitive_additional:
                if field in request.POST:
                    detailed_data["cognitive_function"][field] = request.POST.get(
                        field, ""
                    )

            # BPSD症状（複数選択）の処理
            if "bpsd_symptoms" in request.POST:
                detailed_data["cognitive_function"]["bpsd_symptoms"] = (
                    request.POST.getlist("bpsd_symptoms")
                )

            # サービス関連
            services_additional = ["other_services_detail_text", "services_notes"]
            for field in services_additional:
                if field in request.POST:
                    detailed_data["services"][field] = request.POST.get(field, "")

            # 居住状況関連
            living_additional = [
                "other_home_environment_detail_text",
                "other_housing_type_detail_text",
                "other_ownership_detail_text",
                "other_sleeping_detail_text",
                "housing_notes",
            ]
            for field in living_additional:
                if field in request.POST:
                    detailed_data["living_situation"][field] = request.POST.get(
                        field, ""
                    )

            # JSONフィールドにデータを設定
            assessment.basic_info = detailed_data["basic_info"]
            assessment.insurance_info = detailed_data["insurance_info"]
            assessment.family_situation = detailed_data["family_situation"]
            assessment.living_situation = detailed_data["living_situation"]
            assessment.services = detailed_data["services"]
            assessment.health_status = detailed_data["health_status"]
            assessment.physical_status = detailed_data["physical_status"]
            assessment.cognitive_function = detailed_data["cognitive_function"]
            assessment.basic_activities = detailed_data["basic_activities"]
            assessment.adl = detailed_data["adl"]
            assessment.iadl = detailed_data["iadl"]

            assessment.save()

            # actionによる分岐処理
            if action == "excel":
                # Excel出力（自動保存済み）
                return generate_assessment_excel(assessment, request)
            else:
                # 通常の保存
                messages.success(request, "アセスメントを更新しました")
                url = (
                    reverse("client_detail", kwargs={"pk": assessment.client.pk})
                    + "#assessment-history"
                )
                return HttpResponseRedirect(url)
        else:
            # フォームバリデーションエラーが発生した場合の処理
            print("=== FORM VALIDATION ERROR (EDIT) ===")
            print(f"Form errors: {form.errors}")
            print(f"Form non-field errors: {form.non_field_errors()}")

            # 各フィールドのエラーを詳細出力
            for field, errors in form.errors.items():
                print(f"Field '{field}' errors: {errors}")
            print("=== END FORM VALIDATION ERROR ===")

            messages.error(
                request, "フォームにエラーがあります。入力内容を確認してください。"
            )

            # フォームの選択肢を設定（バリデーションエラー後も選択肢を保持）
            form.fields["assessment_type"].widget.choices = (
                Assessment.ASSESSMENT_TYPE_CHOICES
            )
            form.fields["interview_location"].widget.choices = (
                Assessment.INTERVIEW_LOCATION_CHOICES
            )
    else:
        # 既存のデータでフォームを初期化
        initial_data = {
            "assessment_type": assessment.assessment_type,
            "interview_location": assessment.interview_location,
        }

        # その他フィールドの初期値を設定
        if assessment.basic_info:
            initial_data["assessment_type_other"] = assessment.basic_info.get(
                "assessment_type_other", ""
            )
            initial_data["interview_location_other"] = assessment.basic_info.get(
                "interview_location_other", ""
            )

        # 作成日の初期値を設定
        initial_data["created_date"] = assessment.assessment_date

        # プロフィールが存在しない場合のエラーを回避
        creator_name = ""
        if assessment.assessor:
            try:
                creator_name = assessment.assessor.profile.get_full_name()
            except Exception:
                creator_name = assessment.assessor.username
        initial_data["creator"] = creator_name

        # JSONフィールドから既存データを読み込み
        if assessment.basic_info:
            for key, value in assessment.basic_info.items():
                initial_data[key] = value
        if assessment.insurance_info:
            for key, value in assessment.insurance_info.items():
                initial_data[key] = value
        if assessment.family_situation:
            for key, value in assessment.family_situation.items():
                initial_data[key] = value
        if assessment.living_situation:
            for key, value in assessment.living_situation.items():
                initial_data[key] = value
        if assessment.services:
            for key, value in assessment.services.items():
                initial_data[key] = value
        if assessment.health_status:
            for key, value in assessment.health_status.items():
                # 数値フィールドの処理
                if key in ["height", "weight"] and value:
                    try:
                        initial_data[key] = float(value) if value != "" else None
                    except (ValueError, TypeError):
                        initial_data[key] = None
                else:
                    initial_data[key] = value
        if assessment.physical_status:
            for key, value in assessment.physical_status.items():
                # 数値フィールドの処理
                if key in ["height", "weight"] and value:
                    try:
                        initial_data[key] = float(value) if value != "" else None
                    except (ValueError, TypeError):
                        initial_data[key] = None
                elif key == "notes":
                    # physical_status.notesをphysical_status_notesフィールドにマッピング
                    initial_data["physical_status_notes"] = value
                elif key == "paralysis_pain_details":
                    initial_data["paralysis_pain_details"] = value
                elif key == "skin_infection_details":
                    initial_data["skin_infection_details"] = value
                elif key == "paralysis_location":
                    initial_data["paralysis_location"] = value
                elif key == "pain_location":
                    initial_data["pain_location"] = value
                else:
                    initial_data[key] = value

        # 既存データの互換性のため、health_statusにphysical_status関連フィールドがある場合の処理
        if assessment.health_status:
            # physical_status_notes
            if "physical_status_notes" in assessment.health_status:
                if (
                    "physical_status_notes" not in initial_data
                    or not initial_data["physical_status_notes"]
                ):
                    initial_data["physical_status_notes"] = assessment.health_status[
                        "physical_status_notes"
                    ]
            # paralysis_pain_details
            if "paralysis_pain_details" in assessment.health_status:
                if (
                    "paralysis_pain_details" not in initial_data
                    or not initial_data["paralysis_pain_details"]
                ):
                    initial_data["paralysis_pain_details"] = assessment.health_status[
                        "paralysis_pain_details"
                    ]
            # skin_infection_details
            if "skin_infection_details" in assessment.health_status:
                if (
                    "skin_infection_details" not in initial_data
                    or not initial_data["skin_infection_details"]
                ):
                    initial_data["skin_infection_details"] = assessment.health_status[
                        "skin_infection_details"
                    ]
            # paralysis_location
            if "paralysis_location" in assessment.health_status:
                if (
                    "paralysis_location" not in initial_data
                    or not initial_data["paralysis_location"]
                ):
                    initial_data["paralysis_location"] = assessment.health_status[
                        "paralysis_location"
                    ]
            # pain_location
            if "pain_location" in assessment.health_status:
                if (
                    "pain_location" not in initial_data
                    or not initial_data["pain_location"]
                ):
                    initial_data["pain_location"] = assessment.health_status[
                        "pain_location"
                    ]
        if assessment.cognitive_function:
            for key, value in assessment.cognitive_function.items():
                initial_data[key] = value
        if assessment.basic_activities:
            for key, value in assessment.basic_activities.items():
                initial_data[key] = value
        if assessment.adl:
            for key, value in assessment.adl.items():
                initial_data[key] = value
        if assessment.iadl:
            for key, value in assessment.iadl.items():
                initial_data[key] = value

        form = DetailedAssessmentForm(initial=initial_data)

        # フォームの選択肢を設定
        form.fields["assessment_type"].widget.choices = (
            Assessment.ASSESSMENT_TYPE_CHOICES
        )
        form.fields["interview_location"].widget.choices = (
            Assessment.INTERVIEW_LOCATION_CHOICES
        )

    context = {
        "form": form,
        "assessment": assessment,
        "assessment_types": Assessment.ASSESSMENT_TYPE_CHOICES,
        "interview_locations": Assessment.INTERVIEW_LOCATION_CHOICES,
        "is_edit": True,
    }

    return render(request, "assessments/detailed_assessment_form.html", context)


@login_required
def detailed_assessment_create(request):
    """アセスメント作成ビュー"""
    client_id = request.GET.get("client")
    selected_client = None

    if client_id:
        selected_client = get_object_or_404(Client, pk=client_id)

    if request.method == "POST":
        form = DetailedAssessmentForm(request.POST)

        if form.is_valid():
            # actionパラメータを取得
            action = request.POST.get("action", "save")

            # 基本アセスメント作成
            client_pk = request.POST.get("selected_client")
            client = get_object_or_404(Client, pk=client_pk)

            assessment = Assessment(
                client=client,
                assessor=request.user,
                assessment_type=request.POST.get("assessment_type", "new"),
                interview_location=request.POST.get("interview_location", "home"),
            )

            # 作成日を設定
            created_date = request.POST.get("created_date")
            if created_date:
                from datetime import datetime

                try:
                    assessment.assessment_date = datetime.strptime(
                        created_date, "%Y-%m-%d"
                    ).date()
                except ValueError:
                    pass

            # 詳細データをJSONフィールドに保存
            detailed_data = {}

            # 各セクションのデータを整理
            detailed_data["basic_info"] = {
                "life_history": form.cleaned_data.get("life_history", ""),
                "personal_hopes": form.cleaned_data.get("personal_hopes", ""),
                "family_hopes": form.cleaned_data.get("family_hopes", ""),
                "assessment_type_other": request.POST.get("assessment_type_other", ""),
                "interview_location_other": request.POST.get(
                    "interview_location_other", ""
                ),
                "special_situation_status": request.POST.get(
                    "special_situation_status", ""
                ),
                "special_situation": request.POST.getlist("special_situation"),
                "special_situation_other": request.POST.get(
                    "special_situation_other", ""
                ),
            }

            detailed_data["insurance_info"] = {
                "care_level": request.POST.get("care_level", ""),
                "dementia_level": request.POST.get("dementia_level", ""),
                "burden_ratio": form.cleaned_data.get("burden_ratio", ""),
                "medical_insurance": form.cleaned_data.get("medical_insurance", ""),
                "disability_handbook": form.cleaned_data.get(
                    "disability_handbook", False
                ),
                "difficulty_certificate": form.cleaned_data.get(
                    "difficulty_certificate", False
                ),
                "life_protection": form.cleaned_data.get("life_protection", False),
            }

            detailed_data["family_situation"] = {
                "family_member1": form.cleaned_data.get("family_member1", ""),
                "family_member2": form.cleaned_data.get("family_member2", ""),
            }

            detailed_data["living_situation"] = {
                "home_environment": request.POST.get("home_environment", ""),
                "other_home_environment_detail_text": request.POST.get(
                    "other_home_environment_detail_text", ""
                ),
                "housing_type": request.POST.get("housing_type", ""),
                "other_housing_type_detail_text": request.POST.get(
                    "other_housing_type_detail_text", ""
                ),
                "has_elevator": request.POST.get("has_elevator", ""),
                "housing_ownership": request.POST.get("housing_ownership", ""),
                "other_ownership_detail_text": request.POST.get(
                    "other_ownership_detail_text", ""
                ),
                "private_room": request.POST.get("private_room", ""),
                "air_conditioning": request.POST.get("air_conditioning", ""),
                "toilet_type": request.POST.get("toilet_type", ""),
                "toilet_handrail": request.POST.get("toilet_handrail", ""),
                "toilet_step": request.POST.get("toilet_step", ""),
                "bathroom": request.POST.get("bathroom", ""),
                "bathroom_handrail": request.POST.get("bathroom_handrail", ""),
                "bathroom_step": request.POST.get("bathroom_step", ""),
                "sleeping_arrangement": request.POST.get("sleeping_arrangement", ""),
                "other_sleeping_detail_text": request.POST.get(
                    "other_sleeping_detail_text", ""
                ),
                "room_level_difference": request.POST.get("room_level_difference", ""),
                "housing_modification": request.POST.get("housing_modification", ""),
                "housing_modification_need": request.POST.get(
                    "housing_modification_need", ""
                ),
                "housing_notes": request.POST.get("housing_notes", ""),
            }

            detailed_data["services"] = {
                "care_services": form.cleaned_data.get("care_services", []),
                "welfare_equipment": form.cleaned_data.get("welfare_equipment", []),
                "other_services": form.cleaned_data.get("other_services", ""),
                "informal_services": form.cleaned_data.get("informal_services", []),
                "other_services_detail_text": form.cleaned_data.get(
                    "other_services_detail_text", ""
                ),
                "social_relationships": form.cleaned_data.get(
                    "social_relationships", ""
                ),
            }

            detailed_data["health_status"] = {
                # 疾患名・発症日フィールド
                "disease_name_1": request.POST.get("disease_name_1", ""),
                "disease_name_2": request.POST.get("disease_name_2", ""),
                "disease_name_3": request.POST.get("disease_name_3", ""),
                "disease_name_4": request.POST.get("disease_name_4", ""),
                "disease_name_5": request.POST.get("disease_name_5", ""),
                "disease_name_6": request.POST.get("disease_name_6", ""),
                "onset_date_1": request.POST.get("onset_date_1", ""),
                "onset_date_2": request.POST.get("onset_date_2", ""),
                "onset_date_3": request.POST.get("onset_date_3", ""),
                "onset_date_4": request.POST.get("onset_date_4", ""),
                "onset_date_5": request.POST.get("onset_date_5", ""),
                "onset_date_6": request.POST.get("onset_date_6", ""),
                # 既存フィールド
                "disease_1": form.cleaned_data.get("disease_1", ""),
                "disease_2": form.cleaned_data.get("disease_2", ""),
                "disease_3": form.cleaned_data.get("disease_3", ""),
                "medical_history": form.cleaned_data.get("medical_history", ""),
                "health_basic_info_details": form.cleaned_data.get(
                    "health_basic_info_details", ""
                ),
                "main_doctor_hospital": form.cleaned_data.get(
                    "main_doctor_hospital", ""
                ),
                "main_doctor_name": form.cleaned_data.get("main_doctor_name", ""),
                "visiting_doctor_hospital": form.cleaned_data.get(
                    "visiting_doctor_hospital", ""
                ),
                "visiting_doctor_name": form.cleaned_data.get(
                    "visiting_doctor_name", ""
                ),
                "medication_status": form.cleaned_data.get("medication_status", ""),
                "medication_content": form.cleaned_data.get("medication_content", ""),
                "allergy_details": form.cleaned_data.get("allergy_details", ""),
                "main_doctor": form.cleaned_data.get("main_doctor", ""),
                "outpatient_clinic": form.cleaned_data.get("outpatient_clinic", ""),
                "medication": form.cleaned_data.get("medication", ""),
                "hospital": form.cleaned_data.get("hospital", ""),
                "medication_details": form.cleaned_data.get("medication_details", ""),
                "allergies": form.cleaned_data.get("allergies", False),
                # 特別な医療処置
                "special_medical_treatment_status": request.POST.get(
                    "special_medical_treatment_status", ""
                ),
                "special_medical_treatment": request.POST.getlist(
                    "special_medical_treatment"
                ),
            }

            # フォームで定義されていないが、テンプレートで使用されているフィールドを直接POSTから取得
            additional_fields = [
                "main_doctor_hospital",
                "main_doctor_name",
                "visiting_doctor_hospital",
                "visiting_doctor_name",
                "family_doctor_hospital_1",
                "family_doctor_name_1",
                "family_doctor_hospital_2",
                "family_doctor_name_2",
                "family_doctor_hospital_3",
                "family_doctor_name_3",
                "family_doctor_hospital_4",
                "family_doctor_name_4",
                "hospital_visit_status",
                "hospital_visit_method",
                "hospital_visit_method_other",
                "medical_institution_notes",
                "medication_status",
                "medication_content",
                "allergy_details",
                "medication_allergy_notes",
                "has_sleeping_medication",
                "has_laxative",
                "has_allergy",
                "has_skin_disease",
                "skin_disease_other",
                "infection_presence",
                "infection_other_text",
                "skin_infection_details",
                "has_paralysis",
                "has_pain",
                "uses_hearing_aid",
                "uses_glasses",
                "sensory_function_details",
            ]

            for field in additional_fields:
                if field in request.POST:
                    detailed_data["health_status"][field] = request.POST.get(field, "")

            # physical_statusの初期化
            detailed_data["physical_status"] = {
                "skin_condition": form.cleaned_data.get("skin_condition", []),
                "infection_status": form.cleaned_data.get("infection_status", []),
                "special_treatment": form.cleaned_data.get("special_treatment", []),
                "pain_location": form.cleaned_data.get("pain_location", ""),
                "pain_existence": form.cleaned_data.get("pain_existence", ""),
                "height": (
                    str(form.cleaned_data.get("height", ""))
                    if form.cleaned_data.get("height")
                    else ""
                ),
                "weight": (
                    str(form.cleaned_data.get("weight", ""))
                    if form.cleaned_data.get("weight")
                    else ""
                ),
            }

            detailed_data["cognitive_function"] = {
                "dementia_presence": form.cleaned_data.get("dementia_presence", ""),
                "dementia_severity": form.cleaned_data.get("dementia_severity", ""),
                "dementia_details": form.cleaned_data.get("dementia_details", ""),
                "bpsd_presence": form.cleaned_data.get("bpsd_presence", ""),
                "bpsd_symptoms": form.cleaned_data.get("bpsd_symptoms", []),
                "bpsd_details": form.cleaned_data.get("bpsd_details", ""),
                "conversation": form.cleaned_data.get("conversation", ""),
                "communication": form.cleaned_data.get("communication", ""),
                "cognitive_notes": form.cleaned_data.get("cognitive_notes", ""),
            }

            detailed_data["basic_activities"] = {
                "turning_over": form.cleaned_data.get("turning_over", ""),
                "getting_up": form.cleaned_data.get("getting_up", ""),
                "sitting": form.cleaned_data.get("sitting", ""),
                "standing_up": form.cleaned_data.get("standing_up", ""),
                "standing": form.cleaned_data.get("standing", ""),
                "transfer": form.cleaned_data.get("transfer", ""),
                "indoor_mobility": form.cleaned_data.get("indoor_mobility", ""),
                "outdoor_mobility": form.cleaned_data.get("outdoor_mobility", ""),
                "indoor_mobility_equipment": form.cleaned_data.get(
                    "indoor_mobility_equipment", ""
                ),
                "outdoor_mobility_equipment": form.cleaned_data.get(
                    "outdoor_mobility_equipment", ""
                ),
                "basic_activity_notes": form.cleaned_data.get(
                    "basic_activity_notes", ""
                ),
            }

            detailed_data["adl"] = {
                "eating_method": form.cleaned_data.get("eating_method", ""),
                "eating_assistance": form.cleaned_data.get("eating_assistance", ""),
                "swallowing": form.cleaned_data.get("swallowing", ""),
                "meal_form_main": form.cleaned_data.get("meal_form_main", ""),
                "meal_form_side": form.cleaned_data.get("meal_form_side", ""),
                "water_thickening": form.cleaned_data.get("water_thickening", False),
                "eating_restriction": form.cleaned_data.get("eating_restriction", ""),
                "eating_utensils": form.cleaned_data.get("eating_utensils", []),
                "eating_notes": form.cleaned_data.get("eating_notes", ""),
                "oral_hygiene_assistance": form.cleaned_data.get(
                    "oral_hygiene_assistance", ""
                ),
                "natural_teeth_presence": form.cleaned_data.get(
                    "natural_teeth_presence", ""
                ),
                "denture_presence": form.cleaned_data.get("denture_presence", ""),
                "denture_type": form.cleaned_data.get("denture_type", ""),
                "denture_location": form.cleaned_data.get("denture_location", ""),
                "tooth_decay": form.cleaned_data.get("tooth_decay", False),
                "oral_notes": form.cleaned_data.get("oral_notes", ""),
                "bathing_assistance": form.cleaned_data.get("bathing_assistance", ""),
                "bathing_form": form.cleaned_data.get("bathing_form", ""),
                "bathing_restriction": form.cleaned_data.get(
                    "bathing_restriction", False
                ),
                "dressing_upper": form.cleaned_data.get("dressing_upper", ""),
                "dressing_lower": form.cleaned_data.get("dressing_lower", ""),
                "bathing_notes": form.cleaned_data.get("bathing_notes", ""),
                "excretion_assistance": form.cleaned_data.get(
                    "excretion_assistance", ""
                ),
                "urination": form.cleaned_data.get("urination", False),
                "urinary_incontinence": form.cleaned_data.get(
                    "urinary_incontinence", False
                ),
                "defecation": form.cleaned_data.get("defecation", False),
                "fecal_incontinence": form.cleaned_data.get(
                    "fecal_incontinence", False
                ),
                "daytime_location": form.cleaned_data.get("daytime_location", []),
                "daytime_method": form.cleaned_data.get("daytime_method", []),
                "nighttime_location": form.cleaned_data.get("nighttime_location", []),
                "nighttime_method": form.cleaned_data.get("nighttime_method", []),
                "excretion_supplies": form.cleaned_data.get("excretion_supplies", []),
                "excretion_notes": form.cleaned_data.get("excretion_notes", ""),
            }

            detailed_data["iadl"] = {
                "cooking": form.cleaned_data.get("cooking", ""),
                "cleaning": form.cleaned_data.get("cleaning", ""),
                "washing": form.cleaned_data.get("washing", ""),
                "shopping": form.cleaned_data.get("shopping", ""),
                "money_management": form.cleaned_data.get("money_management", ""),
                "iadl_notes": form.cleaned_data.get("iadl_notes", ""),
            }

            # physical_status関連フィールドをphysical_statusに移動
            if "physical_status_notes" in request.POST:
                detailed_data["physical_status"]["notes"] = request.POST.get(
                    "physical_status_notes", ""
                )
            if "paralysis_pain_details" in request.POST:
                detailed_data["physical_status"]["paralysis_pain_details"] = (
                    request.POST.get("paralysis_pain_details", "")
                )
            if "skin_infection_details" in request.POST:
                detailed_data["physical_status"]["skin_infection_details"] = (
                    request.POST.get("skin_infection_details", "")
                )
            if "paralysis_location" in request.POST:
                detailed_data["physical_status"]["paralysis_location"] = (
                    request.POST.get("paralysis_location", "")
                )
            if "pain_location" in request.POST:
                detailed_data["physical_status"]["pain_location"] = request.POST.get(
                    "pain_location", ""
                )

            # MultipleChoiceFieldの特別処理
            detailed_data["health_status"]["skin_disease"] = request.POST.getlist(
                "skin_disease"
            )
            detailed_data["health_status"]["infection"] = request.POST.getlist(
                "infection"
            )

            # その他のセクションのフィールドも追加
            # 基本動作関連
            basic_activity_additional = [
                "indoor_mobility_equipment_other",
                "outdoor_mobility_equipment_other",
            ]
            for field in basic_activity_additional:
                if field in request.POST:
                    detailed_data["basic_activities"][field] = request.POST.get(
                        field, ""
                    )

            # ADL関連
            adl_additional = [
                "water_thickening_level",
                "eating_restriction_detail",
                "eating_utensils_other_detail",
                "bathing_restriction_detail",
                "urinary_incontinence_frequency",
                "fecal_incontinence_frequency",
            ]
            for field in adl_additional:
                if field in request.POST:
                    detailed_data["adl"][field] = request.POST.get(field, "")

            # 認知機能関連
            cognitive_additional = ["dementia_details", "bpsd_details"]
            for field in cognitive_additional:
                if field in request.POST:
                    detailed_data["cognitive_function"][field] = request.POST.get(
                        field, ""
                    )

            # BPSD症状（複数選択）の処理
            if "bpsd_symptoms" in request.POST:
                detailed_data["cognitive_function"]["bpsd_symptoms"] = (
                    request.POST.getlist("bpsd_symptoms")
                )

            # サービス関連
            services_additional = ["other_services_detail_text", "services_notes"]
            for field in services_additional:
                if field in request.POST:
                    detailed_data["services"][field] = request.POST.get(field, "")

            # 居住状況関連
            living_additional = [
                "other_home_environment_detail_text",
                "other_housing_type_detail_text",
                "other_ownership_detail_text",
                "other_sleeping_detail_text",
                "housing_notes",
            ]
            for field in living_additional:
                if field in request.POST:
                    detailed_data["living_situation"][field] = request.POST.get(
                        field, ""
                    )

            # physical_statusの基本データを設定（既存の追加フィールドを保持）
            detailed_data["physical_status"].update(
                {
                    "skin_condition": form.cleaned_data.get("skin_condition", []),
                    "infection_status": form.cleaned_data.get("infection_status", []),
                    "special_treatment": form.cleaned_data.get("special_treatment", []),
                    "pain_location": form.cleaned_data.get("pain_location", ""),
                    "pain_existence": form.cleaned_data.get("pain_existence", ""),
                    "height": (
                        str(form.cleaned_data.get("height", ""))
                        if form.cleaned_data.get("height")
                        else ""
                    ),
                    "weight": (
                        str(form.cleaned_data.get("weight", ""))
                        if form.cleaned_data.get("weight")
                        else ""
                    ),
                    "vision": form.cleaned_data.get("vision", ""),
                    "hearing": form.cleaned_data.get("hearing", ""),
                    "physical_notes": form.cleaned_data.get("physical_notes", ""),
                }
            )



            # JSONフィールドにデータを設定
            assessment.basic_info = detailed_data["basic_info"]
            assessment.insurance_info = detailed_data["insurance_info"]
            assessment.family_situation = detailed_data["family_situation"]
            assessment.living_situation = detailed_data["living_situation"]
            assessment.services = detailed_data["services"]
            assessment.health_status = detailed_data["health_status"]
            assessment.physical_status = detailed_data["physical_status"]
            assessment.cognitive_function = detailed_data["cognitive_function"]
            assessment.basic_activities = detailed_data["basic_activities"]
            assessment.adl = detailed_data["adl"]
            assessment.iadl = detailed_data["iadl"]

            assessment.save()

            # actionによる分岐処理
            if action == "excel":
                # Excel出力（自動保存済み）
                return generate_assessment_excel(assessment, request)
            else:
                # 通常の保存
                messages.success(request, "アセスメントを保存しました")
                url = (
                    reverse("client_detail", kwargs={"pk": assessment.client.pk})
                    + "#assessment-history"
                )
                return HttpResponseRedirect(url)
        else:
            # フォームバリデーションエラーの詳細をログ出力
            print("=== FORM VALIDATION ERROR (CREATE) ===")
            print(f"Form errors: {form.errors}")
            print(f"Form non-field errors: {form.non_field_errors()}")

            # 各フィールドのエラーを詳細出力
            for field, errors in form.errors.items():
                print(f"Field '{field}' errors: {errors}")
            print("=== END FORM VALIDATION ERROR ===")

            messages.error(
                request, "フォームにエラーがあります。入力内容を確認してください。"
            )

            # フォームの選択肢を設定（バリデーションエラー後も選択肢を保持）
            form.fields["assessment_type"].widget.choices = (
                Assessment.ASSESSMENT_TYPE_CHOICES
            )
            form.fields["interview_location"].widget.choices = (
                Assessment.INTERVIEW_LOCATION_CHOICES
            )
    else:
        form = DetailedAssessmentForm()

        # フォームの選択肢を設定
        form.fields["assessment_type"].widget.choices = (
            Assessment.ASSESSMENT_TYPE_CHOICES
        )
        form.fields["interview_location"].widget.choices = (
            Assessment.INTERVIEW_LOCATION_CHOICES
        )

    clients = Client.objects.all()

    context = {
        "form": form,
        "clients": clients,
        "selected_client": selected_client,
        "assessment_types": Assessment.ASSESSMENT_TYPE_CHOICES,
        "interview_locations": Assessment.INTERVIEW_LOCATION_CHOICES,
        "today": date.today().strftime("%Y-%m-%d"),
        "is_edit": False,
    }

    # 前回アセスメントデータの取得
    if selected_client:
        latest_assessment = (
            Assessment.objects.filter(client=selected_client)
            .order_by("-assessment_date", "-created_at")
            .first()
        )
        if latest_assessment:
            context["has_previous_assessment"] = True
            context["previous_assessment_data"] = json.dumps(
                {
                    "basic_info": latest_assessment.basic_info or {},
                    "insurance_info": latest_assessment.insurance_info or {},
                    "family_situation": latest_assessment.family_situation or {},
                    "living_situation": latest_assessment.living_situation or {},
                    "services": latest_assessment.services or {},
                    "health_status": latest_assessment.health_status or {},
                    "physical_status": latest_assessment.physical_status or {},
                    "cognitive_function": latest_assessment.cognitive_function or {},
                    "basic_activities": latest_assessment.basic_activities or {},
                    "adl": latest_assessment.adl or {},
                    "iadl": latest_assessment.iadl or {},
                },
                ensure_ascii=False,
            )

    return render(request, "assessments/detailed_assessment_form.html", context)


def generate_assessment_excel(assessment, request=None):
    """アセスメントのExcel生成（共通関数）"""
    if not OPENPYXL_AVAILABLE:
        raise Exception(
            "Excel出力機能は現在利用できません。openpyxlライブラリのインストールが必要です。"
        )

    try:
        # --- 保存先パスの事前特定（ファイル名とサブフォルダ） ---
        network_base_path = r"N:\居宅関係★★\利用者フォルダ"
        
        # 苗字のふりがな（苗字）と氏名を組み合わせてファイル名を生成
        furigana = assessment.client.furigana or ""
        last_name_kana = furigana.split()[0] if furigana.split() else ""
        full_name = (assessment.client.name or "").replace(" ", "").replace("　", "")
        save_filename = f"{last_name_kana}_{full_name}様.xlsm"

        # 担当者（アセッサー）の苗字を取得してサブフォルダを特定
        staff_last_name = ""
        if assessment.assessor and hasattr(assessment.assessor, "profile"):
            staff_last_name = assessment.assessor.profile.last_name
        
        target_dir = os.path.join(network_base_path, staff_last_name) if staff_last_name else network_base_path
        existing_file_path = os.path.join(target_dir, save_filename)
        # ----------------------------------------------------

        # 座標設定ファイルを読み込み
        coordinates_path = os.path.join(
            settings.BASE_DIR, "static", "config", "excel_coordinates_assessment.json"
        )
        with open(coordinates_path, "r", encoding="utf-8") as f:
            coordinates = json.load(f)

        coords = coordinates["assessment_sheet"]

        # テンプレート（ベースファイル）の決定
        # 既存ファイルがあればそれをベースにし、なければ原本を使用する
        default_template_path = os.path.join(
            settings.BASE_DIR, "templates", "forms", "assessment_sheet_genpon.xlsm"
        )
        
        import tempfile
        import shutil

        # 一時ファイルを作成
        with tempfile.NamedTemporaryFile(suffix=".xlsm", delete=False) as temp_file:
            temp_filepath = temp_file.name

        # ベースファイルのコピー
        if os.path.exists(existing_file_path):
            shutil.copy2(existing_file_path, temp_filepath)
            print(f"INFO: Using existing file as template: {existing_file_path}")
        else:
            if not os.path.exists(default_template_path):
                raise Exception(f"Excelテンプレートが見つかりません: {default_template_path}")
            shutil.copy2(default_template_path, temp_filepath)
            print(f"INFO: Using default template: {default_template_path}")

        # openpyxl でワークブックを読み込み（マクロ付きファイルに対応）
        workbook = load_workbook(temp_filepath, keep_vba=True)

        # シート名 → ワークシートオブジェクト取得ヘルパー
        def get_sheet(name):
            if name in workbook.sheetnames:
                return workbook[name]
            return None

        # セルへの書き込みヘルパー（openpyxl の MergedCell には直接書き込めないため上位セルに書く）
        def write_cell(ws, cell_ref, value):
            """openpyxl 用セル書き込み。結合セルの場合は左上セルに書き込む。"""
            if not cell_ref or ws is None:
                return
            if value is None:
                value = ""
            if isinstance(value, list):
                value = "、".join([str(v) for v in value])
            try:
                cell = ws[cell_ref]
                if isinstance(cell, MergedCell):
                    # 結合セルの左上セルを特定して書き込む
                    for merged_range in ws.merged_cells.ranges:
                        if cell_ref in merged_range:
                            top_left = ws.cell(merged_range.min_row, merged_range.min_col)
                            top_left.value = value
                            return
                else:
                    cell.value = value
            except Exception as e:
                print(f"DEBUG: セル {cell_ref} への書き込みエラー (シート: {ws.title}, 値: {value}): {str(e)}")



            # --- 和暦変換関数 ---
            def to_wareki(date):
                if not date: return ""
                from datetime import date as date_class, datetime as datetime_class
                # datetime.datetime → datetime.date に変換
                if isinstance(date, datetime_class):
                    date = date.date()
                year, month, day = date.year, date.month, date.day
                if date >= date_class(2019, 5, 1):
                    wareki_year, era = year - 2018, "令和"
                elif date >= date_class(1989, 1, 8):
                    wareki_year, era = year - 1988, "平成"
                elif date >= date_class(1926, 12, 25):
                    wareki_year, era = year - 1925, "昭和"
                elif date >= date_class(1912, 7, 30):
                    wareki_year, era = year - 1911, "大正"
                else:
                    wareki_year, era = year - 1867, "明治"

                if wareki_year == 1:
                    return f"{era}元年{month}月{day}日"
                return f"{era}{wareki_year}年{month}月{day}日"

            # 選択肢の値→表示名変換用辞書
            CHOICE_LABELS = {
                "basic_no_assistance": "つかまらないでできる",
                "basic_with_assistance": "何かにつかまればできる",
                "basic_cannot": "できない",
                "sitting_can_do": "できる",
                "sitting_self_support": "自分の手で支えればできる",
                "sitting_support_needed": "支えてもらえればできる",
                "sitting_cannot": "できない",
                "standing_no_support": "支えなしでできる",
                "standing_with_something": "何か支えがあればできる",
                "standing_cannot": "できない",
                "mobility_independent": "自立",
                "mobility_supervision": "見守り",
                "mobility_partial_assistance": "一部介助",
                "mobility_full_assistance": "全介助",
                "equipment_none": "なし",
                "equipment_wheelchair": "車椅子",
                "equipment_walker": "歩行器",
                "equipment_cane": "杖",
                "equipment_crutches": "松葉杖",
                "equipment_other": "その他",
                "eating_method_oral": "経口摂取",
                "eating_method_tube_oral": "経管栄養+経口摂取",
                "eating_method_tube_only": "経管栄養（胃ろう）",
                "eating_method_tube_nasal": "経管栄養（経鼻）",
                "eating_method_tube_gastronomy": "経管栄養（腸ろう）",
                "eating_assistance_independent": "自立",
                "eating_assistance_supervision": "見守り",
                "eating_assistance_partial_assistance": "一部介助",
                "eating_assistance_full_assistance": "全介助",
                "eating_restriction_yes": "あり",
                "eating_restriction_no": "なし",
                "eating_restriction_unknown": "不明",
                "swallowing_can_do": "できる",
                "swallowing_supervision_needed": "見守り等が必要",
                "swallowing_cannot": "できない",
                "meal_main_normal": "普通",
                "meal_main_soft": "軟飯",
                "meal_main_porridge": "全粥",
                "meal_main_paste": "ペースト",
                "water_thickening_yes": "あり",
                "water_thickening_no": "なし",
                "teeth_yes": "あり",
                "teeth_no": "なし",
                "denture_complete": "総義歯",
                "denture_partial": "部分義歯",
                "denture_upper": "上顎",
                "denture_lower": "下顎",
                "denture_both": "上下",
                "bathing_form_regular_bath": "一般浴",
                "bathing_form_sitting_bath": "寝台浴",
                "bathing_form_shower_bath": "シャワー浴",
                "bathing_form_chair_bath": "チェアー浴",
                "dressing_independent": "自立",
                "dressing_supervision": "見守り",
                "dressing_partial_assistance": "一部介助",
                "dressing_full_assistance": "全介助",
                "excretion_yes": "あり",
                "excretion_no": "なし",
                "incontinence_yes": "あり",
                "incontinence_no": "なし",
                "location_toilet": "トイレ",
                "location_portable_toilet": "Pトイレ",
                "location_bed": "ベッド",
                "iadl_independent": "自立",
                "iadl_supervision": "見守り",
                "iadl_partial_assistance": "一部介助",
                "iadl_full_assistance": "全介助",
                "dementia_yes": "あり",
                "dementia_no": "なし",
                "dementia_mild": "軽度",
                "dementia_moderate": "中等度",
                "dementia_severe": "重度",
                "bpsd_none": "なし",
                "bpsd_persecution_delusion": "被害妄想",
                "bpsd_confabulation": "作話",
                "bpsd_mood_instability": "感情の不安定",
                "bpsd_day_night_reversal": "昼夜逆転",
                "bpsd_home_return_desire": "帰宅願望",
                "bpsd_loud_voice": "大声・奇声",
                "bpsd_violence": "暴力・暴言",
                "bpsd_collection_habit": "収集癖",
                "bpsd_care_resistance": "介護抵抗",
                "bpsd_restlessness": "落ち着きがない",
                "bpsd_wandering": "徘徊",
                "bpsd_destructive_behavior": "破壊行為",
                "bpsd_severe_forgetfulness": "ひどい物忘れ",
                "bpsd_selfish_behavior": "自分勝手な行動",
                "bpsd_agitation": "不穏",
                "bpsd_depression_tendency": "うつ傾向",
                "skin_bedsore": "床ずれ",
                "skin_eczema": "湿疹",
                "skin_itching": "かゆみ",
                "skin_athletes_foot": "水虫",
                "skin_shingles": "帯状疱疹",
                "infection_tuberculosis": "結核",
                "infection_pneumonia": "肺炎",
                "infection_hepatitis": "肝炎",
                "infection_scabies": "疥癬",
                "infection_mrsa": "MRSA",
                "conversation_possible": "可能",
                "conversation_unclear": "不明瞭",
                "conversation_somewhat_difficult": "やや不自由",
                "conversation_completely_impossible": "全くできない",
                "conversation_impossible": "全くできない",
                "communication_possible": "可能",
                "communication_only_sometimes": "その場のみ可能",
                "home_family_cohabitation": "家族と同居",
                "home_living_alone": "一人暮らし",
                "home_elderly_household": "高齢世帯",
                "home_daytime_alone": "日中独居",
                "home_two_generation_house": "二世帯住宅",
                "home_other": "その他",
                "housing_detached_house": "一戸建て",
                "housing_apartment_complex": "集合住宅",
                "housing_public_housing": "公営住宅",
                "housing_condominium": "マンション",
                "housing_other": "その他",
                "ownership_owned": "所有",
                "ownership_rental": "賃貸",
                "ownership_lodging": "間借り",
                "ownership_other": "その他",
                "room_available": "あり",
                "room_not_available": "なし",
                "aircon_available": "あり",
                "aircon_not_available": "なし",
                "toilet_western": "洋式",
                "toilet_japanese": "和式",
                "bathroom_available": "あり",
                "bathroom_not_available": "なし",
                "sleeping_tatami_floor": "畳・床",
                "sleeping_regular_bed": "ベッド",
                "sleeping_care_bed": "介護用ベッド",
                "sleeping_other": "その他",
                "level_diff_available": "あり",
                "level_diff_not_available": "なし",
                "modification_completed": "あり",
                "modification_not_completed": "なし",
                "modification_need_needed": "あり",
                "modification_need_not_needed": "なし",
                "vision_normal": "正常",
                "vision_large_letters_ok": "大きい字は可",
                "vision_barely_visible": "ほぼ見えない",
                "vision_blind": "失明",
                "hearing_normal": "正常",
                "hearing_loud_voice_ok": "大きい声は可",
                "hearing_barely_audible": "ほぼ聞こえない",
                "hearing_deaf": "聞こえない",
                "glasses_yes": "使用",
                "glasses_no": "未使用",
                "hearing_aid_yes": "使用",
                "hearing_aid_no": "未使用",
                "service_home_help": "訪問介護",
                "service_visit_bath": "訪問入浴",
                "service_visit_nursing": "訪問看護",
                "service_visit_rehab": "訪問リハビリ",
                "service_day_service": "通所介護",
                "service_day_rehab": "通所リハビリ",
                "service_short_stay": "ショートステイ",
                "service_small_scale_multi": "小規模多機能",
                "welfare_wheelchair": "車いす",
                "welfare_walker": "車いす付属品",
                "welfare_special_bed": "特殊寝台",
                "welfare_special_bed_accessories": "特殊寝台付属品",
                "welfare_fall_prevention": "床ずれ防止用具",
                "welfare_position_changer": "体位変換器",
                "welfare_walking_aid": "手すり",
                "welfare_slope": "スロープ",
                "welfare_walking_frame": "歩行器",
                "welfare_walking_support": "歩行補助つえ",
                "welfare_detect_sensor": "排個感知機器",
                "welfare_mobility_lift": "移動用リフト",
                "welfare_automatic_drainage": "自動排泄処理装置",
                "informal_family_support": "家族による支援",
                "informal_neighbor_support": "近隣による支援",
                "informal_volunteer": "ボランティア",
                "informal_community_group": "地域活動グループ",
                "informal_friend_support": "友人による支援",
                "informal_npo_support": "NPO団体支援",
                "informal_religious_support": "宗教団体支援",
                "informal_meal_support": "食事支援",
                "informal_excretion_support": "排泄支援",
                "informal_linen_lease": "リネンリース",
                "informal_laundry": "洗濯",
                "informal_room_cleaning": "居室清掃",
                "informal_sputum_suction": "喀痰吸引",
                "informal_diaper_supply": "オムツ支給",
                "informal_taxi_voucher": "タクシー券",
                "informal_massage_voucher": "マッサージ券",
                "informal_fire_alarm": "火災報知器",
                "informal_auto_fire_extinguisher": "自動消火器",
                "informal_elderly_phone": "老人用電話",
                "informal_bedding_disinfection": "寝具乾燥消毒",
                "informal_induction_cooker": "電磁調理器",
                "informal_emergency_alert": "緊急通報装置",
                "informal_meal_delivery": "配食サービス",
                "informal_other_services": "その他",
                "utensil_chopsticks": "箸",
                "utensil_spoon": "スプーン",
                "utensil_apron": "エプロン",
                "utensil_assistive": "補助具",
                "utensil_other": "その他",
                # 排泄用品（excretion_supplies）
                "supply_rehabilitation_pants": "リハビリパンツ",
                "supply_paper_diaper": "紙おむつ",
                "supply_small_pad": "小パット",
                "supply_large_pad": "大パット",
                # 服薬状況（medication_status）
                "medication_self": "自己管理",
                "medication_family": "家族管理",
                "medication_support": "支援あり",
                "medication_none": "服薬なし",
                # 通院状況（hospital_visit_status）
                "visit_status_regular": "定期通院",
                "visit_status_occasional": "時々通院",
                "visit_status_home_visit": "往診",
                "visit_status_none": "通院なし",
                # 通院方法（hospital_visit_method）
                "visit_method_self": "本人",
                "visit_method_family": "家族",
                "visit_method_care_assistance": "通院乗降介助",
                "visit_method_care_taxi": "介護タクシー",
                # 口腔ケア介助（oral_hygiene_assistance）- MOBILITY_CHOICES
                "oral_assistance_independent": "自立",
                "oral_assistance_supervision": "見守り",
                "oral_assistance_partial_assistance": "一部介助",
                "oral_assistance_full_assistance": "全介助",
                # 入浴介助（bathing_assistance）- MOBILITY_CHOICES
                "bathing_assistance_independent": "自立",
                "bathing_assistance_supervision": "見守り",
                "bathing_assistance_partial_assistance": "一部介助",
                "bathing_assistance_full_assistance": "全介助",
                # 入浴制限（bathing_restriction）
                "bathing_restriction_yes": "あり",
                "bathing_restriction_no": "なし",
                # 排泄介助（excretion_assistance）- MOBILITY_CHOICES
                "excretion_assistance_independent": "自立",
                "excretion_assistance_supervision": "見守り",
                "excretion_assistance_partial_assistance": "一部介助",
                "excretion_assistance_full_assistance": "全介助",
            }

            def get_choice_label(value):
                """選択肢の値を表示名に変換"""
                if not value:
                    return ""
                # リストや辞書の場合はそのまま返す（変換しない）
                if isinstance(value, (list, dict)):
                    return value
                # f"prefix_{value}" パターンで value が空だった場合（例: "vision_"）は空文字を返す
                if isinstance(value, str) and value.endswith('_'):
                    return ""
                # 文字列の場合のみ変換
                return CHOICE_LABELS.get(value, value)

            def convert_list_to_text(value_list, prefix=""):
                """リスト形式の値を連結テキストに変換する（型安全）"""
                if not value_list:
                    return ""
                if not isinstance(value_list, list):
                    return str(value_list)
                labels = []
                for item in value_list:
                    raw_label = get_choice_label(f"{prefix}{item}") if prefix else get_choice_label(item)
                    label = str(raw_label) if raw_label is not None else ""
                    if label:
                        labels.append(label)
                return "、".join(labels)

            # --- アセスメントシートのワークシート設定と書き込み関数 ---
            assessment_sheet_name = coords.get("sheet_name", "（最新）アセスメントシート")
            worksheet = get_sheet(assessment_sheet_name)

            # 既存データをクリア（空白データが前回の値を残さないよう）
            for _key, _cell_ref in coords.items():
                if _key != "sheet_name" and _cell_ref:
                    try:
                        write_cell(worksheet, _cell_ref, "")
                    except Exception:
                        pass

            def safe_write_cell(cell_ref, value):
                try:
                    if cell_ref:
                        if value is None: value = ""
                        if isinstance(value, list):
                            value = "、".join([str(v) for v in value])
                        write_cell(worksheet, cell_ref, value)
                except Exception as e:
                    print(f"DEBUG: セル {cell_ref} への書き込みエラー (シート: {assessment_sheet_name}, 値: {value}): {str(e)}")

            def safe_write_wareki_cell(cell_ref, value):
                """和暦文字列を書き込む（openpyxlはテキストとして書き込む）"""
                try:
                    if cell_ref:
                        if value is None: value = ""
                        write_cell(worksheet, cell_ref, str(value))
                except Exception as e:
                    print(f"DEBUG: セル {cell_ref} への和暦書き込みエラー (値: {value}): {str(e)}")

            
            # クライアント基本情報
            safe_write_cell(coords.get("client_name"), assessment.client.name)
            safe_write_cell(coords.get("client_furigana"), assessment.client.furigana)
            safe_write_cell(
                coords.get("client_gender"), assessment.client.get_gender_display()
            )
            # アセスメント時点の年齢を計算して数値として書き込み
            client_age = ""
            if assessment.client.birth_date:
                from datetime import date as date_class
                ref_date = assessment.assessment_date or date_class.today()
                birth_date = assessment.client.birth_date
                client_age = ref_date.year - birth_date.year - ((ref_date.month, ref_date.day) < (birth_date.month, birth_date.day))
            safe_write_cell(coords.get("client_age"), client_age)
            safe_write_cell(coords.get("client_address"), assessment.client.address)
            safe_write_cell(coords.get("client_contact"), assessment.client.phone)
            safe_write_wareki_cell(
                coords.get("birth_date"),
                to_wareki(assessment.client.birth_date) if assessment.client.birth_date else ""
            )

            # アセスメント基本情報
            safe_write_wareki_cell(
                coords.get("assessment_date"),
                (
                    to_wareki(assessment.assessment_date)
                    if assessment.assessment_date
                    else ""
                ),
            )

            assessor_name = ""
            assessor_kana_raw = ""
            if assessment.assessor:
                try:
                    assessor_name = assessment.assessor.profile.get_full_name()
                    assessor_kana_raw = assessment.assessor.profile.get_full_name_kana()
                except Exception:
                    # Profileが存在しない場合はusernameを使用
                    assessor_name = assessment.assessor.username
            safe_write_cell(coords.get("assessor"), assessor_name)

            # アセスメントの理由（その他の場合は詳細を書き込む）
            if (
                assessment.assessment_type == "other"
                and assessment.assessment_type_other
            ):
                safe_write_cell(
                    coords.get("assessment_type"), assessment.assessment_type_other
                )
            else:
                safe_write_cell(
                    coords.get("assessment_type"),
                    assessment.get_assessment_type_display(),
                )

            # 面談場所（その他の場合は詳細を書き込む）
            if (
                assessment.interview_location == "other"
                and assessment.interview_location_other
            ):
                safe_write_cell(
                    coords.get("interview_location"),
                    assessment.interview_location_other,
                )
            else:
                safe_write_cell(
                    coords.get("interview_location"),
                    assessment.get_interview_location_display(),
                )

            # 被保険者番号は数値として書き込み
            insurance_val = assessment.client.insurance_number
            if insurance_val:
                try:
                    insurance_val = int(insurance_val)
                except (ValueError, TypeError):
                    pass
            safe_write_cell(coords.get("insurance_number"), insurance_val)
            safe_write_cell(
                coords.get("care_level_official"),
                (
                    assessment.client.get_care_level_display()
                    if assessment.client.care_level
                    else ""
                ),
            )
            safe_write_wareki_cell(
                coords.get("certification_date"),
                to_wareki(assessment.client.certification_date) if assessment.client.certification_date else ""
            )
            safe_write_wareki_cell(
                coords.get("certification_period_start"),
                to_wareki(assessment.client.certification_period_start) if assessment.client.certification_period_start else ""
            )
            safe_write_wareki_cell(
                coords.get("certification_period_end"),
                to_wareki(assessment.client.certification_period_end) if assessment.client.certification_period_end else ""
            )
            # 負担割合（「1割負担」→「1割」に変換）
            if assessment.client.care_burden:
                burden_text = assessment.client.care_burden.replace("負担", "")
                safe_write_cell(coords.get("burden_ratio"), burden_text)
            else:
                safe_write_cell(coords.get("burden_ratio"), "")

            # 保険関連情報（利用者モデルから取得）
            client = assessment.client

            # 医療保険
            safe_write_cell(
                coords.get("medical_insurance"), client.get_medical_insurance_type_display() if client.medical_insurance_type else ""
            )

            # 身体障がい者手帳
            if client.disability_handbook:
                handbook_text = f"あり（{client.get_disability_handbook_type_display()}）" if client.disability_handbook_type else "あり"
                safe_write_cell(coords.get("disability_handbook"), handbook_text)
            else:
                safe_write_cell(coords.get("disability_handbook"), "なし")

            # 難病申請
            if client.difficult_disease:
                safe_write_cell(coords.get("difficult_disease"), "あり")
            else:
                safe_write_cell(coords.get("difficult_disease"), "なし")

            # 生活保護
            if client.life_protection:
                safe_write_cell(coords.get("life_protection"), "あり")
            else:
                safe_write_cell(coords.get("life_protection"), "なし")

            # 障がい者日常生活自立度・認知症日常生活自立度
            if client.disability_level:
                safe_write_cell(
                    coords.get("disability_level"),
                    client.get_disability_level_display(),
                )
            if client.dementia_level:
                safe_write_cell(
                    coords.get("dementia_level"), client.get_dementia_level_display()
                )

            # 家族状況（利用者登録から取得）
            safe_write_cell(
                coords.get("family_member1_name"), assessment.client.family_name1
            )
            safe_write_cell(
                coords.get("family_member1_relation"),
                assessment.client.family_relationship_detail1,
            )
            safe_write_cell(
                coords.get("family_member1_address"), assessment.client.family_address1
            )
            safe_write_cell(
                coords.get("family_member1_contact"), assessment.client.family_contact1
            )
            safe_write_cell(
                coords.get("family_member1_care"),
                assessment.client.get_family_care_status1_display(),
            )
            safe_write_cell(
                coords.get("family_member1_living_together"),
                assessment.client.get_family_living_status1_display(),
            )
            safe_write_cell(
                coords.get("family_member1_job"),
                assessment.client.get_family_employment1_display(),
            )
            safe_write_cell(
                coords.get("family_member1_notes"), assessment.client.family_notes1
            )

            safe_write_cell(
                coords.get("family_member2_name"), assessment.client.family_name2
            )
            safe_write_cell(
                coords.get("family_member2_relation"),
                assessment.client.family_relationship_detail2,
            )
            safe_write_cell(
                coords.get("family_member2_address"), assessment.client.family_address2
            )
            safe_write_cell(
                coords.get("family_member2_contact"), assessment.client.family_contact2
            )
            safe_write_cell(
                coords.get("family_member2_care"),
                assessment.client.get_family_care_status2_display(),
            )
            safe_write_cell(
                coords.get("family_member2_living_together"),
                assessment.client.get_family_living_status2_display(),
            )
            safe_write_cell(
                coords.get("family_member2_job"),
                assessment.client.get_family_employment2_display(),
            )
            safe_write_cell(
                coords.get("family_member2_notes"), assessment.client.family_notes2
            )

            # 住居状況（居住状況）
            if assessment.living_situation:
                housing = assessment.living_situation
                # 家庭環境
                home_env = housing.get("home_environment", "")
                if home_env:
                    if home_env == "other":
                        other_detail = housing.get(
                            "other_home_environment_detail_text", ""
                        )
                        if other_detail:
                            safe_write_cell(
                                coords.get("home_environment"), other_detail
                            )
                        else:
                            safe_write_cell(
                                coords.get("home_environment"),
                                get_choice_label(f"home_{home_env}"),
                            )
                    else:
                        safe_write_cell(
                            coords.get("home_environment"),
                            get_choice_label(f"home_{home_env}"),
                        )
                # 住宅形態（エレベーター情報を含む）
                housing_type = housing.get("housing_type", "")
                if housing_type:
                    housing_text = ""
                    if housing_type == "other":
                        other_detail = housing.get("other_housing_type_detail_text", "")
                        housing_text = (
                            other_detail
                            if other_detail
                            else get_choice_label(f"housing_{housing_type}")
                        )
                    else:
                        housing_text = get_choice_label(f"housing_{housing_type}")

                    # エレベーター情報を追加（一戸建て以外の場合）
                    if housing_type != "detached_house":
                        has_elevator = housing.get("has_elevator", "")
                        if has_elevator == "yes":
                            housing_text = f"{housing_text}、エレベーター有"
                        elif has_elevator == "no":
                            housing_text = f"{housing_text}、エレベーター無"

                    safe_write_cell(coords.get("housing_type"), housing_text)

                # 住宅権利
                ownership = housing.get("housing_ownership", "")
                if ownership:
                    if ownership == "other":
                        other_detail = housing.get("other_ownership_detail_text", "")
                        if other_detail:
                            safe_write_cell(
                                coords.get("housing_ownership"), other_detail
                            )
                        else:
                            safe_write_cell(
                                coords.get("housing_ownership"),
                                get_choice_label(f"ownership_{ownership}"),
                            )
                    else:
                        safe_write_cell(
                            coords.get("housing_ownership"),
                            get_choice_label(f"ownership_{ownership}"),
                        )
                # 個室
                private_room = housing.get("private_room", "")
                if private_room:
                    safe_write_cell(
                        coords.get("private_room"),
                        get_choice_label(f"room_{private_room}"),
                    )
                # エアコン
                air_cond = housing.get("air_conditioning", "")
                if air_cond:
                    safe_write_cell(
                        coords.get("air_conditioning"),
                        get_choice_label(f"aircon_{air_cond}"),
                    )

                # トイレ（手すり、段差情報を含む）
                toilet = housing.get("toilet_type", "")
                if toilet:
                    toilet_text = get_choice_label(f"toilet_{toilet}")

                    # 手すり情報を追加
                    toilet_handrail = housing.get("toilet_handrail", "")
                    if toilet_handrail == "yes":
                        toilet_text = f"{toilet_text}、手すり有"
                    elif toilet_handrail == "no":
                        toilet_text = f"{toilet_text}、手すり無"

                    # 段差情報を追加
                    toilet_step = housing.get("toilet_step", "")
                    if toilet_step == "yes":
                        toilet_text = f"{toilet_text}、段差有"
                    elif toilet_step == "no":
                        toilet_text = f"{toilet_text}、段差無"

                    safe_write_cell(coords.get("toilet_type"), toilet_text)

                # 浴室（手すり、段差情報を含む）
                bathroom = housing.get("bathroom", "")
                if bathroom:
                    bathroom_text = get_choice_label(f"bathroom_{bathroom}")

                    # 「あり」の場合のみ、手すりと段差情報を追加
                    if bathroom == "available":
                        # 手すり情報を追加
                        bathroom_handrail = housing.get("bathroom_handrail", "")
                        if bathroom_handrail == "yes":
                            bathroom_text = f"{bathroom_text}、手すり有"
                        elif bathroom_handrail == "no":
                            bathroom_text = f"{bathroom_text}、手すり無"

                        # 段差情報を追加
                        bathroom_step = housing.get("bathroom_step", "")
                        if bathroom_step == "yes":
                            bathroom_text = f"{bathroom_text}、段差有"
                        elif bathroom_step == "no":
                            bathroom_text = f"{bathroom_text}、段差無"

                    safe_write_cell(coords.get("bathroom"), bathroom_text)
                # 就寝環境
                sleeping = housing.get("sleeping_arrangement", "")
                if sleeping:
                    if sleeping == "other":
                        other_detail = housing.get("other_sleeping_detail_text", "")
                        if other_detail:
                            safe_write_cell(
                                coords.get("sleeping_arrangement"), other_detail
                            )
                        else:
                            safe_write_cell(
                                coords.get("sleeping_arrangement"),
                                get_choice_label(f"sleeping_{sleeping}"),
                            )
                    else:
                        safe_write_cell(
                            coords.get("sleeping_arrangement"),
                            get_choice_label(f"sleeping_{sleeping}"),
                        )
                # 段差
                level_diff = housing.get("room_level_difference", "")
                if level_diff:
                    safe_write_cell(
                        coords.get("room_level_difference"),
                        get_choice_label(f"level_diff_{level_diff}"),
                    )
                # 住宅改修
                modification = housing.get("housing_modification", "")
                if modification:
                    safe_write_cell(
                        coords.get("housing_modification"),
                        get_choice_label(f"modification_{modification}"),
                    )
                # 住宅改修の必要性
                mod_need = housing.get("housing_modification_need", "")
                if mod_need:
                    safe_write_cell(
                        coords.get("housing_modification_need"),
                        get_choice_label(f"modification_need_{mod_need}"),
                    )
                # 特別な状況（basic_infoに保存されている）
                special_situation_labels = {
                    "abuse": "虐待", "terminal": "終末期", "sudden_caregiver_absence": "介護者の急変",
                    "financial_difficulty": "経済的困窮", "social_isolation": "社会的孤立",
                    "housing_problem": "住宅問題", "family_conflict": "家族間の対立",
                    "self_neglect": "セルフネグレクト", "disaster_victim": "被災者",
                    "multiple_care_burden": "多重介護負担", "mental_health_issue": "精神的問題",
                    "other": "その他",
                }
                if assessment.basic_info:
                    situation_list = assessment.basic_info.get("special_situation", [])
                    situation_other = assessment.basic_info.get("special_situation_other", "")
                    if isinstance(situation_list, list) and situation_list:
                        situation_texts = [special_situation_labels.get(s, s) for s in situation_list]
                        if "other" in situation_list and situation_other:
                            situation_texts = [t if t != "その他" else situation_other for t in situation_texts]
                        safe_write_cell(coords.get("special_circumstances"), "、".join(situation_texts))
                    else:
                        safe_write_cell(coords.get("special_circumstances"), "")

            # 健康状態
            if assessment.health_status:
                health = assessment.health_status

                # 疾患名と発症日（6件）
                for i in range(1, 7):
                    # 疾患名
                    disease_name = health.get(f"disease_name_{i}", "")
                    safe_write_cell(coords.get(f"disease_name_{i}"), disease_name)

                    # 発症日（和暦変換）
                    onset_date_str = health.get(f"onset_date_{i}", "")
                    onset_val = ""
                    if onset_date_str:
                        try:
                            from datetime import datetime
                            onset_date = datetime.strptime(onset_date_str, "%Y-%m-%d").date()
                            onset_val = to_wareki(onset_date)
                        except (ValueError, TypeError):
                            onset_val = onset_date_str
                    safe_write_wareki_cell(coords.get(f"onset_date_{i}"), onset_val)

                # 既往歴
                safe_write_cell(
                    coords.get("medical_history"), health.get("medical_history", "")
                )

                # 特別な医療処置（リストから文字列に変換し、日本語ラベルに変換）
                special_medical_treatment_status = health.get(
                    "special_medical_treatment_status", ""
                )
                special_medical_treatment = health.get("special_medical_treatment", [])

                # V28用の詳細テキスト
                if isinstance(special_medical_treatment, list):
                    # 特別な医療処置のラベルマッピング（テンプレートのvalueに対応）
                    treatment_labels = {
                        "none": "なし",
                        "iv_management": "点滴の管理",
                        "central_venous_nutrition": "中心静脈栄養",
                        "dialysis": "透析",
                        "stoma_care": "ストーマの処置",
                        "oxygen_therapy": "酸素療法",
                        "ventilator": "人工呼吸器",
                        "tracheostomy_care": "気管切開の処置",
                        "pain_nursing": "疼痛の看護",
                        "tube_feeding": "経管栄養",
                        "monitor_measurement": "モニター測定（血圧、心拍、酸素飽和度等）",
                        "pressure_ulcer_care": "褥瘡の処置",
                        "urinary_catheter": "導尿カテーテル",
                    }
                    treatment_texts = [
                        treatment_labels.get(t, t)
                        for t in special_medical_treatment
                        if t
                    ]
                    special_medical_treatment_text = (
                        "、".join([str(t) for t in treatment_texts if t]) if treatment_texts else ""
                    )
                else:
                    special_medical_treatment_text = special_medical_treatment

                # V28に詳細を書き込む（なしの場合も「なし」と記載）
                if special_medical_treatment_status == "none":
                    safe_write_cell("V28", "なし")
                else:
                    safe_write_cell("V28", special_medical_treatment_text)

                # I60には「あり」「なし」のみを書き込む
                if special_medical_treatment_status == "none":
                    safe_write_cell("I60", "なし")
                elif special_medical_treatment_status == "available":
                    safe_write_cell("I60", "あり（詳細は1ページ参照）")
                else:
                    safe_write_cell("I60", "")

                # 主治医情報（病院名と氏名を分けて記載）
                safe_write_cell(
                    coords.get("main_doctor_hospital"),
                    health.get("main_doctor_hospital", ""),
                )
                safe_write_cell(
                    coords.get("main_doctor_name"), health.get("main_doctor_name", "")
                )

                # 往診医情報（病院名と氏名を分けて記載）
                safe_write_cell(
                    coords.get("visiting_doctor_hospital"),
                    health.get("visiting_doctor_hospital", ""),
                )
                safe_write_cell(
                    coords.get("visiting_doctor_name"),
                    health.get("visiting_doctor_name", ""),
                )

                # かかりつけ医（1-4）全て病院名
                safe_write_cell(
                    coords.get("family_doctor_1"),
                    health.get("family_doctor_hospital_1", ""),
                )
                safe_write_cell(
                    coords.get("family_doctor_2"),
                    health.get("family_doctor_hospital_2", ""),
                )
                safe_write_cell(
                    coords.get("family_doctor_3"),
                    health.get("family_doctor_hospital_3", ""),
                )
                safe_write_cell(
                    coords.get("family_doctor_4"),
                    health.get("family_doctor_hospital_4", ""),
                )

                # 通院情報（通院状況、通院方法、通院情報の詳細をまとめて記載）
                hospital_visit_parts = []

                # 通院状況
                visit_status = health.get("hospital_visit_status", "")
                if visit_status:
                    # 接尾辞付きキーで変換
                    label = get_choice_label(f"visit_status_{visit_status}")
                    hospital_visit_parts.append(f"通院状況：{label}")

                # 通院方法
                visit_method = health.get("hospital_visit_method", "")
                if visit_method:
                    # その他の場合は詳細を含める
                    if visit_method == "other":
                        method_other = health.get("hospital_visit_method_other", "")
                        if method_other:
                            hospital_visit_parts.append(f"通院方法：{method_other}")
                        else:
                            hospital_visit_parts.append(f"通院方法：その他")
                    else:
                        # 接尾辞付きキーで変換
                        label = get_choice_label(f"visit_method_{visit_method}")
                        hospital_visit_parts.append(f"通院方法：{label}")

                # 1行目を作成
                first_line = (
                    "　".join(hospital_visit_parts) if hospital_visit_parts else ""
                )

                # 通院情報の詳細
                hospital_notes = health.get("medical_institution_notes", "")

                # 結合して出力
                if first_line and hospital_notes:
                    full_hospital_visit_text = f"{first_line}\n{hospital_notes}"
                elif first_line:
                    full_hospital_visit_text = first_line
                elif hospital_notes:
                    full_hospital_visit_text = hospital_notes
                else:
                    full_hospital_visit_text = ""

                safe_write_cell(
                    coords.get("hospital_visit_info"), full_hospital_visit_text
                )

                # アレルギー
                has_allergy = health.get("has_allergy", "")
                allergy_details = health.get("allergy_details", "")

                if has_allergy == "yes":
                    # 「あり」の場合、F37に「あり」、I37に詳細を記載
                    safe_write_cell(coords.get("has_allergy"), "あり")
                    safe_write_cell(coords.get("allergy_details"), allergy_details)
                elif has_allergy == "no":
                    # 「なし」の場合
                    safe_write_cell(coords.get("has_allergy"), "なし")
                elif has_allergy:
                    # その他の値の場合（念のため）
                    safe_write_cell(coords.get("has_allergy"), has_allergy)

                # 服薬情報（服薬状況、眠剤、下剤、服薬内容をまとめて記載）
                medication_parts = []

                # 服薬状況
                medication_status = health.get("medication_status", "")
                if medication_status:
                    medication_parts.append(
                        f"服薬状況：{get_choice_label(f'medication_{medication_status}')}"
                    )

                # 眠剤の有無
                has_sleeping = health.get("has_sleeping_medication", "")
                if has_sleeping == "yes":
                    medication_parts.append("眠剤：あり")
                elif has_sleeping == "no":
                    medication_parts.append("眠剤：なし")

                # 下剤の有無
                has_laxative = health.get("has_laxative", "")
                if has_laxative == "yes":
                    medication_parts.append("下剤：あり")
                elif has_laxative == "no":
                    medication_parts.append("下剤：なし")

                # 1行目を作成
                first_line = "　".join([str(p) for p in medication_parts if p]) if medication_parts else ""

                # 服薬内容
                medication_content = health.get("medication_content", "")

                # 結合して出力
                if first_line and medication_content:
                    full_medication_text = f"{first_line}\n{medication_content}"
                elif first_line:
                    full_medication_text = first_line
                elif medication_content:
                    full_medication_text = medication_content
                else:
                    full_medication_text = ""

                safe_write_cell(coords.get("medication_content"), full_medication_text)

            # サービス利用状況
            if assessment.services:
                service = assessment.services
                # 介護サービス
                safe_write_cell(coords.get("care_services"), convert_list_to_text(service.get("care_services", []), "service_"))
                # 福祉用具
                safe_write_cell(coords.get("welfare_equipment"), convert_list_to_text(service.get("welfare_equipment", []), "welfare_"))
                safe_write_cell(
                    coords.get("other_services"), service.get("other_services", "")
                )
                # インフォーマルサービス（リストの各要素にinformal_プレフィックスを追加）
                informal_list = service.get("informal_services", [])
                if informal_list:
                    informal_labels = []
                    for i in informal_list:
                        label = get_choice_label(f"informal_{i}")
                        # 「その他」の場合、詳細を追加
                        if i == "other_services":
                            other_detail = service.get("other_services_detail_text", "")
                            if other_detail:
                                label = other_detail
                        informal_labels.append(label)
                    safe_write_cell(
                        coords.get("informal_services"), "、".join([str(l) for l in informal_labels if l])
                    )
                safe_write_cell(
                    coords.get("social_relationships"),
                    service.get("social_relationships", ""),
                )

            # 本人・家族の希望等（基本情報に含まれる）
            if assessment.basic_info:
                info = assessment.basic_info
                safe_write_cell(
                    coords.get("personal_hopes"), info.get("personal_hopes", "")
                )
                safe_write_cell(
                    coords.get("family_hopes"), info.get("family_hopes", "")
                )
                safe_write_cell(
                    coords.get("life_history"), info.get("life_history", "")
                )
                # 特記・備考（住居備考があれば追記）
                notes_parts = []
                notes_text = info.get("notes", "")
                if notes_text:
                    notes_parts.append(notes_text)
                if assessment.living_situation:
                    housing_notes = assessment.living_situation.get("housing_notes", "")
                    if housing_notes:
                        notes_parts.append(f"【住居備考】{housing_notes}")
                safe_write_cell(
                    coords.get("notes"), "\n".join(notes_parts) if notes_parts else ""
                )

            # 身体状況
            if assessment.physical_status:
                physical = assessment.physical_status

                # 皮膚疾患
                skin_presence = assessment.health_status.get("has_skin_disease", "")
                if skin_presence == "no":
                    safe_write_cell(coords.get("skin_disease"), "なし")
                else:
                    skin_list = assessment.health_status.get("skin_disease", [])
                    s_labels = [str(get_choice_label(f"skin_{item}") if item != "other" else assessment.health_status.get("skin_disease_other", "")) for item in skin_list]
                    s_summary = "、".join([l for l in s_labels if l])
                    safe_write_cell(coords.get("skin_disease"), s_summary if s_summary else ("あり" if skin_presence == "yes" else ""))

                # 感染症
                infection_presence = assessment.health_status.get("infection_presence", "")
                if infection_presence == "no":
                    safe_write_cell(coords.get("infection_status"), "なし")
                else:
                    infection_list = assessment.health_status.get("infection", [])
                    i_labels = [str(get_choice_label(f"infection_{item}") if item != "other" else assessment.health_status.get("infection_other_text", "")) for item in infection_list]
                    i_summary = "、".join([l for l in i_labels if l])
                    safe_write_cell(coords.get("infection_status"), i_summary if i_summary else ("あり" if infection_presence == "yes" else ""))
                safe_write_cell(
                    coords.get("paralysis_location"),
                    physical.get("paralysis_location", ""),
                )
                safe_write_cell(
                    coords.get("pain_location"), physical.get("pain_location", "")
                )
                # 身長・体重は数値として書き込み
                height = physical.get("height", "")
                safe_write_cell(coords.get("height"), float(height) if height and str(height).replace('.','',1).isdigit() else height)
                
                weight = physical.get("weight", "")
                safe_write_cell(coords.get("weight"), float(weight) if weight and str(weight).replace('.','',1).isdigit() else weight)

                safe_write_cell(coords.get("vision"), get_choice_label(f"vision_{physical.get('vision', '')}"))
                safe_write_cell(coords.get("uses_glasses"), get_choice_label(f"glasses_{assessment.health_status.get('uses_glasses', '')}"))
                safe_write_cell(coords.get("hearing"), get_choice_label(f"hearing_{physical.get('hearing', '')}"))
                safe_write_cell(coords.get("uses_hearing_aid"), get_choice_label(f"hearing_aid_{assessment.health_status.get('uses_hearing_aid', '')}"))
                
                # 身体状況メモ（基本情報の詳細＋身体状況の詳細を合わせる）
                physical_notes_parts = []
                # 基本情報の詳細 (health_status.health_basic_info_details)
                if assessment.health_status:
                    b_notes = assessment.health_status.get("health_basic_info_details", "")
                    if b_notes:
                        physical_notes_parts.append(b_notes)
                # 身体状況の詳細 (physical_status.notes)
                p_notes = physical.get("notes", "")
                if p_notes:
                    physical_notes_parts.append(p_notes)
                safe_write_cell(coords.get("physical_notes"), "\n".join(physical_notes_parts))

            # 基本動作
            if assessment.basic_activities:
                activity = assessment.basic_activities
                safe_write_cell(coords.get("turning_over"), get_choice_label(f"basic_{activity.get('turning_over', '')}"))
                safe_write_cell(coords.get("getting_up"), get_choice_label(f"basic_{activity.get('getting_up', '')}"))
                safe_write_cell(coords.get("standing_up"), get_choice_label(f"basic_{activity.get('standing_up', '')}"))
                safe_write_cell(coords.get("sitting"), get_choice_label(f"sitting_{activity.get('sitting', '')}"))
                safe_write_cell(coords.get("standing"), get_choice_label(f"standing_{activity.get('standing', '')}"))
                safe_write_cell(coords.get("transfer"), get_choice_label(f"mobility_{activity.get('transfer', '')}"))
                safe_write_cell(coords.get("indoor_mobility"), get_choice_label(f"mobility_{activity.get('indoor_mobility', '')}"))
                safe_write_cell(coords.get("outdoor_mobility"), get_choice_label(f"mobility_{activity.get('outdoor_mobility', '')}"))
                safe_write_cell(coords.get("indoor_mobility_equipment"), get_choice_label(f"equipment_{activity.get('indoor_mobility_equipment', '')}"))
                safe_write_cell(coords.get("outdoor_mobility_equipment"), get_choice_label(f"equipment_{activity.get('outdoor_mobility_equipment', '')}"))
                safe_write_cell(coords.get("basic_activity_notes"), activity.get("basic_activity_notes", ""))

            # ADL
            if assessment.adl:
                adl = assessment.adl
                # 副食の変換マップ
                meal_side_labels = {
                    "normal": "普通",
                    "soft": "一口大",
                    "minced": "キザミ",
                    "paste": "ペースト",
                    "paste_form": "ペースト",
                }
                safe_write_cell(coords.get("eating_method"), get_choice_label(f"eating_method_{adl.get('eating_method', '')}"))
                safe_write_cell(coords.get("eating_assistance"), get_choice_label(f"eating_assistance_{adl.get('eating_assistance', '')}"))
                safe_write_cell(coords.get("swallowing"), get_choice_label(f"swallowing_{adl.get('swallowing', '')}"))
                safe_write_cell(coords.get("meal_form_main"), get_choice_label(f"meal_main_{adl.get('meal_form_main', '')}"))
                safe_write_cell(coords.get("meal_form_side"), meal_side_labels.get(adl.get("meal_form_side", ""), adl.get("meal_form_side", "")))
                safe_write_cell(coords.get("water_thickening"), get_choice_label(f"water_thickening_{adl.get('water_thickening', '')}"))
                safe_write_cell(coords.get("eating_restriction"), get_choice_label(f"eating_restriction_{adl.get('eating_restriction', '')}"))
                # 食事用具
                utensils_list = adl.get("eating_utensils", [])
                utensils_labels = []
                for u in utensils_list:
                    label = get_choice_label(f"utensil_{u}")
                    if u == "other":
                        other_detail = adl.get("eating_utensils_other_detail", "")
                        if other_detail: label = other_detail
                    utensils_labels.append(label)
                safe_write_cell(coords.get("eating_utensils"), "、".join([str(l) for l in utensils_labels if l]))
                safe_write_cell(coords.get("eating_notes"), adl.get("eating_notes", ""))

                # 口腔
                safe_write_cell(
                    coords.get("oral_hygiene_assistance"),
                    get_choice_label(f"oral_assistance_{adl.get('oral_hygiene_assistance', '')}"),
                )
                safe_write_cell(
                    coords.get("natural_teeth_presence"),
                    get_choice_label(f"teeth_{adl.get('natural_teeth_presence', '')}"),
                )
                safe_write_cell(
                    coords.get("denture_type"),
                    get_choice_label(f"denture_{adl.get('denture_type', '')}"),
                )
                safe_write_cell(
                    coords.get("denture_location"),
                    get_choice_label(f"denture_{adl.get('denture_location', '')}"),
                )
                safe_write_cell(coords.get("oral_notes"), adl.get("oral_notes", ""))

                # 入浴・更衣
                safe_write_cell(
                    coords.get("bathing_assistance"),
                    get_choice_label(f"bathing_assistance_{adl.get('bathing_assistance', '')}"),
                )
                safe_write_cell(
                    coords.get("bathing_form"),
                    get_choice_label(f"bathing_form_{adl.get('bathing_form', '')}"),
                )
                safe_write_cell(
                    coords.get("bathing_restriction"),
                    get_choice_label(f"bathing_restriction_{adl.get('bathing_restriction', '')}"),
                )
                safe_write_cell(
                    coords.get("dressing_upper"),
                    get_choice_label(f"dressing_{adl.get('dressing_upper', '')}"),
                )
                safe_write_cell(
                    coords.get("dressing_lower"),
                    get_choice_label(f"dressing_{adl.get('dressing_lower', '')}"),
                )
                safe_write_cell(
                    coords.get("bathing_notes"), adl.get("bathing_notes", "")
                )

                # 排泄
                safe_write_cell(
                    coords.get("excretion_assistance"),
                    get_choice_label(f"excretion_assistance_{adl.get('excretion_assistance', '')}"),
                )
                safe_write_cell(
                    coords.get("urination"),
                    get_choice_label(f"excretion_{adl.get('urination', '')}"),
                )
                
                urinary_incon = adl.get("urinary_incontinence", "")
                urinary_label = get_choice_label(f"incontinence_{urinary_incon}") if urinary_incon else ""
                if urinary_incon == "yes":
                    urinary_freq = adl.get("urinary_incontinence_frequency", "")
                    if urinary_freq:
                        urinary_label = f"{urinary_label}（{urinary_freq}）"
                safe_write_cell(coords.get("urinary_incontinence"), urinary_label)

                safe_write_cell(
                    coords.get("defecation"),
                    get_choice_label(f"excretion_{adl.get('defecation', '')}"),
                )
                
                fecal_incon = adl.get("fecal_incontinence", "")
                fecal_label = get_choice_label(f"incontinence_{fecal_incon}") if fecal_incon else ""
                if fecal_incon == "yes":
                    fecal_freq = adl.get("fecal_incontinence_frequency", "")
                    if fecal_freq:
                        fecal_label = f"{fecal_label}（{fecal_freq}）"
                safe_write_cell(coords.get("fecal_incontinence"), fecal_label)

                # 排泄場所
                daytime_labels = convert_list_to_text(adl.get("daytime_location", []), "location_")
                safe_write_cell(coords.get("daytime_location"), daytime_labels)

                nighttime_labels = convert_list_to_text(adl.get("nighttime_location", []), "location_")
                safe_write_cell(coords.get("nighttime_location"), nighttime_labels)

                # 排泄用品
                supplies_labels = convert_list_to_text(adl.get("excretion_supplies", []), "supply_")
                safe_write_cell(coords.get("excretion_supplies"), supplies_labels)
                safe_write_cell(
                    coords.get("excretion_notes"), adl.get("excretion_notes", "")
                )

            # IADL
            if assessment.iadl:
                iadl = assessment.iadl
                safe_write_cell(coords.get("cooking"), get_choice_label(f"iadl_{iadl.get('cooking', '')}"))
                safe_write_cell(coords.get("washing"), get_choice_label(f"iadl_{iadl.get('washing', '')}"))
                safe_write_cell(coords.get("money_management"), get_choice_label(f"iadl_{iadl.get('money_management', '')}"))
                safe_write_cell(coords.get("cleaning"), get_choice_label(f"iadl_{iadl.get('cleaning', '')}"))
                safe_write_cell(coords.get("shopping"), get_choice_label(f"iadl_{iadl.get('shopping', '')}"))
                safe_write_cell(coords.get("iadl_notes"), iadl.get("iadl_notes", ""))

            # 認知機能
            if assessment.cognitive_function:
                cognitive = assessment.cognitive_function
                # 認知機能
                safe_write_cell(
                    coords.get("dementia_presence"),
                    get_choice_label(f"dementia_{cognitive.get('dementia_presence', '')}"),
                )
                safe_write_cell(
                    coords.get("dementia_severity"),
                    get_choice_label(f"dementia_{cognitive.get('dementia_severity', '')}"),
                )
                # BPSD症状
                bpsd_presence = cognitive.get("bpsd_presence", "")
                if bpsd_presence == "no":
                    safe_write_cell(coords.get("bpsd_symptoms"), "なし")
                else:
                    bpsd_list = cognitive.get("bpsd_symptoms", [])
                    bpsd_labels = [str(get_choice_label(f"bpsd_{s}")) for s in bpsd_list]
                    safe_write_cell(coords.get("bpsd_symptoms"), "、".join([l for l in bpsd_labels if l]) if bpsd_labels else ("あり" if bpsd_presence == "yes" else ""))
                
                safe_write_cell(
                    coords.get("conversation"),
                    get_choice_label(f"conversation_{cognitive.get('conversation', '')}"),
                )
                safe_write_cell(
                    coords.get("communication"),
                    get_choice_label(f"communication_{cognitive.get('communication', '')}"),
                )
                
                # 認知機能メモ
                cognitive_notes_parts = []
                cog_notes = cognitive.get("cognitive_notes", "")
                if cog_notes:
                    cognitive_notes_parts.append(str(cog_notes))
                dem_details = cognitive.get("dementia_details", "")
                if dem_details:
                    cognitive_notes_parts.append(f"【認知症詳細】{dem_details}")
                safe_write_cell(coords.get("cognitive_notes"), "\n".join([str(p) for p in cognitive_notes_parts]))

            # --- 救急医療情報シートへの書き込み ---
            ei_coords = coordinates.get("emergency_info", {})
            ei_sheet_name = ei_coords.get("sheet_name", "救急医療情報")
            all_sheet_names = workbook.sheetnames
            if ei_sheet_name in all_sheet_names:
                ei_sheet = get_sheet(ei_sheet_name)

                # 既存データをクリア（空白データが前回の値を残さないよう）
                for _key, _cell_ref in ei_coords.items():
                    if _key != "sheet_name" and _cell_ref:
                        try:
                            write_cell(ei_sheet, _cell_ref, "")
                        except Exception:
                            pass

                def safe_write_ei(cell_ref, value):
                    try:
                        if cell_ref:
                            if value is None: value = ""
                            write_cell(ei_sheet, cell_ref, value)
                    except Exception as e:
                        print(f"DEBUG: 救急医療情報 セル {cell_ref} 書き込みエラー: {str(e)}")

                def safe_write_ei_wareki(cell_ref, value):
                    try:
                        if cell_ref:
                            if value is None: value = ""
                            write_cell(ei_sheet, cell_ref, str(value))
                    except Exception as e:
                        print(f"DEBUG: 救急医療情報 セル {cell_ref} 和暦書き込みエラー: {str(e)}")


                safe_write_ei(ei_coords.get("client_name"), assessment.client.name)
                safe_write_ei(ei_coords.get("client_furigana"), assessment.client.furigana)
                safe_write_ei(ei_coords.get("client_gender"), assessment.client.get_gender_display() if assessment.client.gender else "")
                safe_write_ei_wareki(ei_coords.get("birth_date"), to_wareki(assessment.client.birth_date) if assessment.client.birth_date else "")
                safe_write_ei(ei_coords.get("client_age"), client_age)
                safe_write_ei(ei_coords.get("client_address"), assessment.client.address)
                safe_write_ei(
                    ei_coords.get("care_level_official"),
                    assessment.client.get_care_level_display() if assessment.client.care_level else ""
                )
                # 事業所名・担当者
                from django.conf import settings as dj_settings
                office_name = getattr(dj_settings, "OFFICE_NAME", "居宅介護支援事業所")
                safe_write_ei(ei_coords.get("main_office_name"), office_name)
                safe_write_ei(ei_coords.get("assessor"), assessor_name)

                # 緊急連絡先①②
                client = assessment.client
                safe_write_ei(ei_coords.get("emergency_contact1_name"), client.family_name1)
                safe_write_ei(ei_coords.get("emergency_contact1_relationship"), client.full_family_relationship1)
                safe_write_ei(ei_coords.get("emergency_contact1_address"), client.family_address1)
                safe_write_ei(ei_coords.get("emergency_contact1_phone"), client.family_contact1)
                safe_write_ei(ei_coords.get("emergency_contact2_name"), client.family_name2)
                safe_write_ei(ei_coords.get("emergency_contact2_relationship"), client.full_family_relationship2)
                safe_write_ei(ei_coords.get("emergency_contact2_address"), client.family_address2)
                safe_write_ei(ei_coords.get("emergency_contact2_phone"), client.family_contact2)

                # 疾患名①〜⑥・既往歴
                health = assessment.health_status or {}
                for i in range(1, 7):
                    safe_write_ei(ei_coords.get(f"disease_name_{i}"), health.get(f"disease_name_{i}", ""))
                safe_write_ei(ei_coords.get("medical_history"), health.get("medical_history", ""))

                # 主治医
                safe_write_ei(ei_coords.get("main_doctor_hospital"), health.get("main_doctor_hospital", ""))
                safe_write_ei(ei_coords.get("main_doctor_name"), health.get("main_doctor_name", ""))

                # 往診医
                safe_write_ei(ei_coords.get("visiting_doctor_hospital"), health.get("visiting_doctor_hospital", ""))
                safe_write_ei(ei_coords.get("visiting_doctor_name"), health.get("visiting_doctor_name", ""))

                # かかりつけ医①②③
                for i in range(1, 4):
                    safe_write_ei(ei_coords.get(f"family_doctor_hospital_{i}"), health.get(f"family_doctor_hospital_{i}", ""))

                # アレルギー
                has_allergy_val = health.get("has_allergy", "")
                if has_allergy_val == "yes":
                    safe_write_ei(ei_coords.get("has_allergy"), "あり")
                elif has_allergy_val == "no":
                    safe_write_ei(ei_coords.get("has_allergy"), "なし")
                else:
                    safe_write_ei(ei_coords.get("has_allergy"), has_allergy_val)
                safe_write_ei(ei_coords.get("allergy_details"), health.get("allergy_details", ""))

                # 屋内外移動方法・介助量
                physical_ei = assessment.basic_activities or {}
                mobility_label = {
                    "independent": "自立", "supervision": "見守り",
                    "partial_assistance": "一部介助", "full_assistance": "全介助",
                }
                equipment_label = {
                    "none": "なし", "wheelchair": "車椅子", "walker": "歩行器",
                    "cane": "杖", "crutches": "松葉杖", "other": "その他",
                }
                indoor_equip_val = equipment_label.get(physical_ei.get("indoor_mobility_equipment", ""), physical_ei.get("indoor_mobility_equipment", ""))
                outdoor_equip_val = equipment_label.get(physical_ei.get("outdoor_mobility_equipment", ""), physical_ei.get("outdoor_mobility_equipment", ""))
                indoor_mob_val = mobility_label.get(physical_ei.get("indoor_mobility", ""), physical_ei.get("indoor_mobility", ""))
                outdoor_mob_val = mobility_label.get(physical_ei.get("outdoor_mobility", ""), physical_ei.get("outdoor_mobility", ""))
                safe_write_ei(ei_coords.get("indoor_mobility_equipment"), indoor_equip_val)
                safe_write_ei(ei_coords.get("outdoor_mobility_equipment"), outdoor_equip_val)
                safe_write_ei(ei_coords.get("indoor_mobility"), indoor_mob_val)
                safe_write_ei(ei_coords.get("outdoor_mobility"), outdoor_mob_val)

            # --- ひらがな→カタカナ変換関数 ---
            def hira_to_kata(text):
                if not text: return ""
                return "".join(chr(ord(c) + 0x60) if '\u3041' <= c <= '\u3096' else c for c in text)

            assessor_kana = hira_to_kata(assessor_kana_raw)

            # --- 医療・介護連携シートへの書き込み ---
            mcc_coords = coordinates.get("medical_care_coordination", {})
            mcc_sheet_name = mcc_coords.get("sheet_name", "医療・介護連携シート")
            if mcc_sheet_name in workbook.sheetnames:
                mcc_sheet = get_sheet(mcc_sheet_name)

                def safe_write_mcc(cell_ref, value):
                    try:
                        if cell_ref and value is not None:
                            write_cell(mcc_sheet, cell_ref, value)
                    except Exception as e:
                        print(f"DEBUG: {mcc_sheet_name} セル {cell_ref} 書き込みエラー: {e}")

                def safe_write_mcc_wareki(cell_ref, value):
                    try:
                        if cell_ref and value is not None:
                                                        write_cell(mcc_sheet, cell_ref, str(value))
                    except Exception as e:
                        print(f"DEBUG: {mcc_sheet_name} セル {cell_ref} 和暦書き込みエラー: {e}")

                safe_write_mcc(mcc_coords.get("care_manager"), assessor_name)
                safe_write_mcc(mcc_coords.get("care_manager_2"), assessor_name)
                safe_write_mcc(mcc_coords.get("client_name"), assessment.client.name)
                safe_write_mcc(mcc_coords.get("client_furigana"), assessment.client.furigana)
                safe_write_mcc_wareki(mcc_coords.get("birth_date"), to_wareki(assessment.client.birth_date) if assessment.client.birth_date else "")
                safe_write_mcc(mcc_coords.get("client_age"), client_age)
                safe_write_mcc(mcc_coords.get("care_level"), assessment.client.get_care_level_display() if assessment.client.care_level else "")
                safe_write_mcc_wareki(mcc_coords.get("certification_period_start"), to_wareki(assessment.client.certification_period_start) if assessment.client.certification_period_start else "")
                safe_write_mcc_wareki(mcc_coords.get("certification_period_end"), to_wareki(assessment.client.certification_period_end) if assessment.client.certification_period_end else "")
                safe_write_mcc(mcc_coords.get("client_address"), assessment.client.address)

            # --- (貼付作成)入院時情報提供シートへの書き込み ---
            hai_coords = coordinates.get("hospital_admission_info", {})
            hai_sheet_name = hai_coords.get("sheet_name", "(貼付作成)入院時情報提供")
            if hai_sheet_name in workbook.sheetnames:
                hai_sheet = get_sheet(hai_sheet_name)

                def safe_write_hai(cell_ref, value):
                    try:
                        if cell_ref and value is not None:
                            write_cell(hai_sheet, cell_ref, value)
                    except Exception as e:
                        print(f"DEBUG: {hai_sheet_name} セル {cell_ref} 書き込みエラー: {e}")

                def safe_write_hai_wareki(cell_ref, value):
                    try:
                        if cell_ref and value is not None:
                                                        write_cell(hai_sheet, cell_ref, str(value))
                    except Exception as e:
                        print(f"DEBUG: {hai_sheet_name} セル {cell_ref} 和暦書き込みエラー: {e}")

                # 住宅形態テキストの生成（アセスメントシートと同様のロジック）
                housing = assessment.living_situation or {}
                housing_type_val = housing.get("housing_type", "")
                housing_type_text = ""
                if housing_type_val:
                    if housing_type_val == "other":
                        housing_type_text = housing.get("other_housing_type_detail_text", "") or get_choice_label(f"housing_{housing_type_val}")
                    else:
                        housing_type_text = get_choice_label(f"housing_{housing_type_val}")
                    if housing_type_val != "detached_house":
                        has_elevator = housing.get("has_elevator", "")
                        if has_elevator == "yes":
                            housing_type_text = f"{housing_type_text}、エレベーター有"
                        elif has_elevator == "no":
                            housing_type_text = f"{housing_type_text}、エレベーター無"

                client = assessment.client
                safe_write_hai(hai_coords.get("client_name"), client.name)
                safe_write_hai(hai_coords.get("client_furigana_katakana"), hira_to_kata(client.furigana or ""))
                safe_write_hai(hai_coords.get("client_address"), client.address)
                safe_write_hai(hai_coords.get("client_gender"), client.get_gender_display() if client.gender else "")
                safe_write_hai_wareki(hai_coords.get("birth_date"), to_wareki(client.birth_date) if client.birth_date else "")
                safe_write_hai(hai_coords.get("family_name1"), client.family_name1)
                safe_write_hai(hai_coords.get("family_relationship1"), client.full_family_relationship1)
                safe_write_hai(hai_coords.get("family_living_status1"), client.get_family_living_status1_display() if client.family_living_status1 else "")
                safe_write_hai(hai_coords.get("family_address1"), client.family_address1)
                safe_write_hai(hai_coords.get("family_contact1"), client.family_contact1)
                safe_write_hai(hai_coords.get("housing_type"), housing_type_text)
                safe_write_hai(hai_coords.get("life_history"), (assessment.basic_info or {}).get("life_history", ""))
                safe_write_hai(hai_coords.get("care_level"), client.get_care_level_display() if client.care_level else "")
                safe_write_hai_wareki(hai_coords.get("certification_period_start"), to_wareki(client.certification_period_start) if client.certification_period_start else "")
                safe_write_hai_wareki(hai_coords.get("certification_period_end"), to_wareki(client.certification_period_end) if client.certification_period_end else "")
                safe_write_hai(hai_coords.get("insurance_number"), client.insurance_number)
                safe_write_hai(hai_coords.get("difficult_disease"), "あり" if client.difficult_disease else "なし")
                difficult_disease_display = client.difficult_disease_other if client.difficult_disease_name == 'other' else client.get_difficult_disease_name_display()
                safe_write_hai(hai_coords.get("difficult_disease_name"), difficult_disease_display)
                # I52・P52: 手帳の有無＋疾患名を組み合わせ
                if client.disability_handbook:
                    handbook_text = "あり"
                else:
                    handbook_text = "なし"
                safe_write_hai(hai_coords.get("disability_handbook_i52"), handbook_text)
                safe_write_hai(hai_coords.get("disability_handbook"), handbook_text)
                # V52: 生活保護
                safe_write_hai(hai_coords.get("life_protection"), "あり" if client.life_protection else "なし")
                # I54: 障がい高齢者の日常生活自立度
                safe_write_hai(hai_coords.get("disability_level"), client.get_disability_level_display() if client.disability_level else "")
                # I55: 認知症の日常生活自立度
                safe_write_hai(hai_coords.get("dementia_level"), client.get_dementia_level_display() if client.dementia_level else "")
                # I56: かかりつけ医1～4をカンマ区切り
                health_hai = assessment.health_status or {}
                family_doctors = []
                for i in range(1, 5):
                    hosp = health_hai.get(f"family_doctor_hospital_{i}", "")
                    name = health_hai.get(f"family_doctor_name_{i}", "")
                    if hosp or name:
                        family_doctors.append("、".join(filter(None, [hosp, name])))
                safe_write_hai(hai_coords.get("family_doctors"), "，".join(family_doctors))
                # 疾患名1～6
                for i in range(1, 7):
                    safe_write_hai(hai_coords.get(f"disease_name_{i}"), health_hai.get(f"disease_name_{i}", ""))
                # I59: 服薬状況，眠剤，下剤
                med_parts = []
                med_status = health_hai.get("medication_status", "")
                if med_status:
                    med_parts.append(get_choice_label(f"medication_{med_status}"))
                has_sleeping = health_hai.get("has_sleeping_medication", "")
                med_parts.append(f"眠剤：{'あり' if has_sleeping == 'yes' else 'なし'}" if has_sleeping else "")
                has_laxative = health_hai.get("has_laxative", "")
                med_parts.append(f"下剤：{'あり' if has_laxative == 'yes' else 'なし'}" if has_laxative else "")
                safe_write_hai(hai_coords.get("medication_summary"), "，".join(filter(None, med_parts)))

                # ADL
                adl_hai = assessment.adl or {}
                meal_side_hai = {"normal": "普通", "soft": "一口大", "minced": "キザミ", "paste": "ペースト", "paste_form": "ペースト"}
                # I60: 食事方法，食事動作
                eating_parts = [
                    get_choice_label(f"eating_method_{adl_hai.get('eating_method', '')}"),
                    get_choice_label(f"eating_assistance_{adl_hai.get('eating_assistance', '')}"),
                ]
                safe_write_hai(hai_coords.get("eating_summary"), "，".join(filter(None, eating_parts)))
                # K61: 主食  P61: 副食  U61: トロミ
                safe_write_hai(hai_coords.get("meal_form_main"), get_choice_label(f"meal_main_{adl_hai.get('meal_form_main', '')}"))
                safe_write_hai(hai_coords.get("meal_form_side"), meal_side_hai.get(adl_hai.get("meal_form_side", ""), adl_hai.get("meal_form_side", "")))
                safe_write_hai(hai_coords.get("water_thickening"), get_choice_label(f"water_thickening_{adl_hai.get('water_thickening', '')}"))
                # I62: 排泄動作
                safe_write_hai(hai_coords.get("excretion_assistance"), get_choice_label(f"excretion_assistance_{adl_hai.get('excretion_assistance', '')}"))
                # I63: 排尿失禁の有無，程度
                u_incon = adl_hai.get("urinary_incontinence", "")
                u_label = get_choice_label(f"incontinence_{u_incon}") if u_incon else ""
                if u_incon == "yes":
                    u_freq = adl_hai.get("urinary_incontinence_frequency", "")
                    if u_freq: u_label = f"{u_label}（{u_freq}）"
                safe_write_hai(hai_coords.get("urinary_incontinence"), u_label)
                # K64/S64: 排尿の日中・夜間の方法
                safe_write_hai(hai_coords.get("excretion_daytime"), convert_list_to_text(adl_hai.get("daytime_location", []), "location_"))
                safe_write_hai(hai_coords.get("excretion_nighttime"), convert_list_to_text(adl_hai.get("nighttime_location", []), "location_"))
                # I65: 排便失禁の有無，頻度
                f_incon = adl_hai.get("fecal_incontinence", "")
                f_label = get_choice_label(f"incontinence_{f_incon}") if f_incon else ""
                if f_incon == "yes":
                    f_freq = adl_hai.get("fecal_incontinence_frequency", "")
                    if f_freq: f_label = f"{f_label}（{f_freq}）"
                safe_write_hai(hai_coords.get("fecal_incontinence"), f_label)
                # K65/S65: 排便の日中・夜間の方法
                safe_write_hai(hai_coords.get("defecation_daytime"), convert_list_to_text(adl_hai.get("daytime_location", []), "location_"))
                safe_write_hai(hai_coords.get("defecation_nighttime"), convert_list_to_text(adl_hai.get("nighttime_location", []), "location_"))
                # I67: 更衣上衣  I68: 更衣下衣
                safe_write_hai(hai_coords.get("dressing_upper"), get_choice_label(f"dressing_{adl_hai.get('dressing_upper', '')}"))
                safe_write_hai(hai_coords.get("dressing_lower"), get_choice_label(f"dressing_{adl_hai.get('dressing_lower', '')}"))

                # 基本動作
                activity_hai = assessment.basic_activities or {}
                safe_write_hai(hai_coords.get("indoor_mobility"), get_choice_label(f"mobility_{activity_hai.get('indoor_mobility', '')}"))
                safe_write_hai(hai_coords.get("transfer"), get_choice_label(f"mobility_{activity_hai.get('transfer', '')}"))
                safe_write_hai(hai_coords.get("indoor_mobility_equipment"), get_choice_label(f"equipment_{activity_hai.get('indoor_mobility_equipment', '')}"))

                # I72: 認知症の有無，程度 / BPSD
                cognitive_hai = assessment.cognitive_function or {}
                dem_pres_val = cognitive_hai.get("dementia_presence", "")
                dem_pres_label = "あり" if dem_pres_val == "yes" else ("なし" if dem_pres_val == "no" else "")
                dem_sev = get_choice_label(f"dementia_{cognitive_hai.get('dementia_severity', '')}")
                dem_line = "，".join(filter(None, [f"認知症：{dem_pres_label}" if dem_pres_label else "", dem_sev]))
                bpsd_pres = cognitive_hai.get("bpsd_presence", "")
                if bpsd_pres == "no":
                    bpsd_line = "BPSD：なし"
                elif bpsd_pres:
                    bpsd_list = cognitive_hai.get("bpsd_symptoms", [])
                    bpsd_labels = [str(get_choice_label(f"bpsd_{s}")) for s in bpsd_list]
                    bpsd_symptoms_str = "、".join(filter(None, bpsd_labels)) or "あり"
                    bpsd_line = f"BPSD：{bpsd_symptoms_str}"
                else:
                    bpsd_line = ""
                safe_write_hai(hai_coords.get("dementia_summary"), "\n".join(filter(None, [dem_line, bpsd_line])))

            # --- (手入力)入院時情報提供シートへの書き込み ---
            hai_m_coords = coordinates.get("hospital_admission_info_manual", {})
            hai_m_sheet_name = hai_m_coords.get("sheet_name", "(手入力)入院時情報提供")
            if hai_m_sheet_name in workbook.sheetnames:
                hai_m_sheet = get_sheet(hai_m_sheet_name)

                def safe_write_haim(cell_ref, value):
                    try:
                        if cell_ref and value is not None:
                            write_cell(hai_m_sheet, cell_ref, value)
                    except Exception as e:
                        print(f"DEBUG: {hai_m_sheet_name} セル {cell_ref} 書き込みエラー: {e}")

                def safe_write_haim_wareki(cell_ref, value):
                    try:
                        if cell_ref and value is not None:
                                                        write_cell(hai_m_sheet, cell_ref, str(value))
                    except Exception as e:
                        print(f"DEBUG: {hai_m_sheet_name} セル {cell_ref} 和暦書き込みエラー: {e}")

                client = assessment.client
                # housing_type_textはすでに(貼付作成)ブロックで計算済みの変数を流用
                safe_write_haim(hai_m_coords.get("client_name"), client.name)
                safe_write_haim(hai_m_coords.get("client_furigana_katakana"), hira_to_kata(client.furigana or ""))
                safe_write_haim(hai_m_coords.get("client_gender"), client.get_gender_display() if client.gender else "")
                safe_write_haim_wareki(hai_m_coords.get("birth_date"), to_wareki(client.birth_date) if client.birth_date else "")
                safe_write_haim(hai_m_coords.get("client_address"), client.address)
                safe_write_haim(hai_m_coords.get("family_name1"), client.family_name1)
                safe_write_haim(hai_m_coords.get("family_relationship1"), client.full_family_relationship1)
                safe_write_haim(hai_m_coords.get("family_living_status1"), client.get_family_living_status1_display() if client.family_living_status1 else "")
                safe_write_haim(hai_m_coords.get("family_address1"), client.family_address1)
                safe_write_haim(hai_m_coords.get("family_contact1"), client.family_contact1)
                housing = assessment.living_situation or {}
                housing_type_val_m = housing.get("housing_type", "")
                housing_type_text_m = ""
                if housing_type_val_m:
                    if housing_type_val_m == "other":
                        housing_type_text_m = housing.get("other_housing_type_detail_text", "") or get_choice_label(f"housing_{housing_type_val_m}")
                    else:
                        housing_type_text_m = get_choice_label(f"housing_{housing_type_val_m}")
                    if housing_type_val_m != "detached_house":
                        has_elev = housing.get("has_elevator", "")
                        if has_elev == "yes":
                            housing_type_text_m = f"{housing_type_text_m}、エレベーター有"
                        elif has_elev == "no":
                            housing_type_text_m = f"{housing_type_text_m}、エレベーター無"
                safe_write_haim(hai_m_coords.get("housing_type"), housing_type_text_m)
                safe_write_haim(hai_m_coords.get("life_history"), (assessment.basic_info or {}).get("life_history", ""))
                safe_write_haim(hai_m_coords.get("care_level"), client.get_care_level_display() if client.care_level else "")
                safe_write_haim_wareki(hai_m_coords.get("certification_period_start"), to_wareki(client.certification_period_start) if client.certification_period_start else "")
                safe_write_haim_wareki(hai_m_coords.get("certification_period_end"), to_wareki(client.certification_period_end) if client.certification_period_end else "")
                safe_write_haim(hai_m_coords.get("insurance_number"), client.insurance_number)
                safe_write_haim(hai_m_coords.get("difficult_disease"), "あり" if client.difficult_disease else "なし")
                difficult_disease_display_m = client.difficult_disease_other if client.difficult_disease_name == 'other' else client.get_difficult_disease_name_display()
                safe_write_haim(hai_m_coords.get("difficult_disease_name"), difficult_disease_display_m)
                handbook_text_m = "あり" if client.disability_handbook else "なし"
                safe_write_haim(hai_m_coords.get("disability_handbook_i52"), handbook_text_m)
                safe_write_haim(hai_m_coords.get("disability_handbook"), handbook_text_m)
                safe_write_haim(hai_m_coords.get("life_protection"), "あり" if client.life_protection else "なし")

            # --- 更新・新規申請書シートへの書き込み ---
            ks_coords = coordinates.get("koshin_shinki_shinsei", {})
            ks_sheet_name = ks_coords.get("sheet_name", "更新・新規申請書")
            if ks_sheet_name in workbook.sheetnames:
                ks_sheet = get_sheet(ks_sheet_name)

                def safe_write_ks(cell_ref, value):
                    try:
                        if cell_ref and value is not None:
                            write_cell(ks_sheet, cell_ref, value)
                    except Exception as e:
                        print(f"DEBUG: 更新・新規申請書 セル {cell_ref} 書き込みエラー: {e}")

                def safe_write_ks_wareki(cell_ref, value):
                    try:
                        if cell_ref and value is not None:
                                                        write_cell(ks_sheet, cell_ref, str(value))
                    except Exception as e:
                        print(f"DEBUG: 更新・新規申請書 セル {cell_ref} 和暦書き込みエラー: {e}")

                safe_write_ks(ks_coords.get("care_manager"), assessor_name)
                safe_write_ks(ks_coords.get("care_manager_2"), assessor_name)
                safe_write_ks(ks_coords.get("care_manager_furigana"), assessor_kana)
                safe_write_ks(ks_coords.get("insurance_number"), assessment.client.insurance_number)
                safe_write_ks(ks_coords.get("client_name"), assessment.client.name)
                safe_write_ks(ks_coords.get("client_furigana"), hira_to_kata(assessment.client.furigana or ""))
                safe_write_ks(ks_coords.get("client_gender"), assessment.client.get_gender_display() if assessment.client.gender else "")
                safe_write_ks_wareki(ks_coords.get("birth_date"), to_wareki(assessment.client.birth_date) if assessment.client.birth_date else "")
                safe_write_ks(ks_coords.get("postal_code"), assessment.client.postal_code)
                safe_write_ks(ks_coords.get("client_address"), assessment.client.address)
                safe_write_ks(ks_coords.get("care_level"), assessment.client.get_care_level_display() if assessment.client.care_level else "")
                safe_write_ks_wareki(ks_coords.get("certification_period_start"), to_wareki(assessment.client.certification_period_start) if assessment.client.certification_period_start else "")
                safe_write_ks_wareki(ks_coords.get("certification_period_end"), to_wareki(assessment.client.certification_period_end) if assessment.client.certification_period_end else "")

                # M25: 医療保険の保険者名（国民健康保険の場合は「国民健康保険 〇市」形式）
                client_ks = assessment.client
                med_type = client_ks.medical_insurance_type
                insurer_name = client_ks.medical_insurer_name_issuer or ""
                if med_type == "national_health":
                    med_insurer_text = f"国民健康保険　{insurer_name}" if insurer_name else "国民健康保険"
                else:
                    med_insurer_text = insurer_name
                safe_write_ks(ks_coords.get("medical_insurer_name"), med_insurer_text)

                # AA25: 保険者番号
                safe_write_ks(ks_coords.get("medical_insurer_number"), client_ks.medical_insurer_number)

                # O26: 記号（枝番があれば「記号〇〇（枝番）〇〇」）
                symbol = client_ks.medical_insurance_symbol
                branch = client_ks.medical_insurance_branch
                if symbol and branch:
                    symbol_text = f"記号{symbol}（枝番）{branch}"
                elif symbol:
                    symbol_text = symbol
                else:
                    symbol_text = ""
                safe_write_ks(ks_coords.get("medical_insurance_symbol"), symbol_text)

                # Y26: 被保険者番号
                safe_write_ks(ks_coords.get("medical_insurance_number"), client_ks.medical_insurance_number)

                # G37: 主治医（医療機関名）  H39: 主治医名
                health_ks = assessment.health_status or {}
                safe_write_ks(ks_coords.get("main_doctor_hospital"), health_ks.get("main_doctor_hospital", ""))
                safe_write_ks(ks_coords.get("main_doctor_name"), health_ks.get("main_doctor_name", ""))

            # --- 区分変更申請書シートへの書き込み ---
            kb_coords = coordinates.get("kubun_henkou_shinsei", {})
            kb_sheet_name = kb_coords.get("sheet_name", "区分変更申請書")
            if kb_sheet_name in workbook.sheetnames:
                kb_sheet = get_sheet(kb_sheet_name)

                def safe_write_kb(cell_ref, value):
                    if cell_ref:
                        try:
                            write_cell(kb_sheet, cell_ref, value)
                        except Exception as e:
                            print(f"DEBUG: 区分変更申請書 セル {cell_ref} 書き込みエラー: {e}")

                def safe_write_kb_wareki(cell_ref, value):
                    if cell_ref:
                        try:
                                                        write_cell(kb_sheet, cell_ref, str(value))
                        except Exception as e:
                            print(f"DEBUG: 区分変更申請書 セル {cell_ref} 和暦書き込みエラー: {e}")

                safe_write_kb(kb_coords.get("care_manager"), assessor_name)
                safe_write_kb(kb_coords.get("insurance_number"), client.insurance_number)
                safe_write_kb(kb_coords.get("client_furigana"), hira_to_kata(client.furigana or ""))
                safe_write_kb(kb_coords.get("client_name"), client.name)
                safe_write_kb(kb_coords.get("client_gender"), client.get_gender_display() if client.gender else "")
                safe_write_kb_wareki(kb_coords.get("birth_date"), to_wareki(client.birth_date) if client.birth_date else "")
                safe_write_kb(kb_coords.get("postal_code"), client.postal_code)
                safe_write_kb(kb_coords.get("client_address"), client.address)
                safe_write_kb(kb_coords.get("care_level"), client.get_care_level_display() if client.care_level else "")
                safe_write_kb_wareki(kb_coords.get("certification_period_start"), to_wareki(client.certification_period_start) if client.certification_period_start else "")
                safe_write_kb_wareki(kb_coords.get("certification_period_end"), to_wareki(client.certification_period_end) if client.certification_period_end else "")
                # 医療保険情報（更新・新規申請書と同じロジック）
                kb_med_type = client.medical_insurance_type
                kb_insurer_name = client.medical_insurer_name_issuer or ""
                if kb_med_type == "national_health":
                    kb_med_insurer_text = f"国民健康保険　{kb_insurer_name}" if kb_insurer_name else "国民健康保険"
                else:
                    kb_med_insurer_text = kb_insurer_name
                safe_write_kb(kb_coords.get("medical_insurer_name"), kb_med_insurer_text)
                safe_write_kb(kb_coords.get("medical_insurer_number"), client.medical_insurer_number)
                kb_symbol = client.medical_insurance_symbol
                kb_branch = client.medical_insurance_branch
                if kb_symbol and kb_branch:
                    kb_symbol_text = f"記号{kb_symbol}（枝番）{kb_branch}"
                elif kb_symbol:
                    kb_symbol_text = kb_symbol
                else:
                    kb_symbol_text = ""
                safe_write_kb(kb_coords.get("medical_insurance_symbol"), kb_symbol_text)
                safe_write_kb(kb_coords.get("medical_insurance_number"), client.medical_insurance_number)
                safe_write_kb(kb_coords.get("care_manager_furigana"), assessor_kana)
                safe_write_kb(kb_coords.get("care_manager_name"), assessor_name)
                health_kb = assessment.health_status or {}
                safe_write_kb(kb_coords.get("main_doctor_hospital"), health_kb.get("main_doctor_hospital", ""))
                safe_write_kb(kb_coords.get("main_doctor_name"), health_kb.get("main_doctor_name", ""))

            # --- 資料提供申請書シートへの書き込み ---
            st_coords = coordinates.get("shiryo_teikyou", {})
            st_sheet_name = st_coords.get("sheet_name", "資料提供申請書")
            if st_sheet_name in workbook.sheetnames:
                st_sheet = get_sheet(st_sheet_name)

                def safe_write_st(cell_ref, value):
                    if cell_ref:
                        try:
                            write_cell(st_sheet, cell_ref, value)
                        except Exception as e:
                            print(f"DEBUG: 資料提供申請書 セル {cell_ref} 書き込みエラー: {e}")

                def safe_write_st_wareki(cell_ref, value):
                    if cell_ref:
                        try:
                                                        write_cell(st_sheet, cell_ref, str(value))
                        except Exception as e:
                            print(f"DEBUG: 資料提供申請書 セル {cell_ref} 和暦書き込みエラー: {e}")

                safe_write_st(st_coords.get("care_manager"), assessor_name)
                safe_write_st(st_coords.get("client_address"), client.address)
                safe_write_st_wareki(st_coords.get("birth_date"), to_wareki(client.birth_date) if client.birth_date else "")
                safe_write_st(st_coords.get("insurance_number"), client.insurance_number)

            # --- 入退去連絡シートへの書き込み ---
            ntr_coords = coordinates.get("nyutaikyo_renraku", {})
            ntr_sheet_name = ntr_coords.get("sheet_name", "入退去連絡")
            if ntr_sheet_name in workbook.sheetnames:
                ntr_sheet = get_sheet(ntr_sheet_name)
                try:
                    write_cell(ntr_sheet, ntr_coords.get("client_name"), client.name)
                except Exception as e:
                    print(f"DEBUG: 入退去連絡 セル {ntr_coords.get('client_name')} 書き込みエラー: {e}")

            # --- 確認書（原本）シートへの書き込み ---
            kg_coords = coordinates.get("kakuninsho_genpon", {})
            kg_sheet_name = kg_coords.get("sheet_name", "確認書（原本）")
            if kg_sheet_name in workbook.sheetnames:
                kg_sheet = get_sheet(kg_sheet_name)
                try:
                    write_cell(kg_sheet, kg_coords.get("certification_period_start"), to_wareki(client.certification_period_start) if client.certification_period_start else "")
                except Exception as e:
                    print(f"DEBUG: 確認書（原本） セル {kg_coords.get('certification_period_start')} 書き込みエラー: {e}")

            # --- 確認書シートへの書き込み（ヘルパー・デイ・福祉用具）---
            for kakuninsho_key in ["kakuninsho_helper", "kakuninsho_day", "kakuninsho_fukushi"]:
                kk_coords = coordinates.get(kakuninsho_key, {})
                kk_sheet_name = kk_coords.get("sheet_name", "")
                if kk_sheet_name and kk_sheet_name in workbook.sheetnames:
                    kk_sheet = get_sheet(kk_sheet_name)
                    try:
                        write_cell(kk_sheet, kk_coords.get("care_manager"), assessor_name)
                        write_cell(kk_sheet, kk_coords.get("client_name"), client.name)
                    except Exception as e:
                        print(f"DEBUG: {kk_sheet_name} 書き込みエラー: {e}")

            # --- モニタリングシートへの書き込み ---
            mon_coords = coordinates.get("monitoring_sheet", {})
            mon_sheet_name = mon_coords.get("sheet_name", "モニタリングシート")
            if mon_sheet_name in workbook.sheetnames:
                mon_sheet = get_sheet(mon_sheet_name)
                try:
                    write_cell(mon_sheet, mon_coords["client_name"], assessment.client.name)
                    write_cell(mon_sheet, mon_coords["care_manager"], assessor_name)
                except Exception as e:
                    print(f"DEBUG: モニタリングシート 書き込みエラー: {e}")

            # --- 担会案内シートへの書き込み ---
            from django.conf import settings as dj_settings
            tk_coords = coordinates.get("tankai_annai", {})
            tk_sheet_name = tk_coords.get("sheet_name", "担会案内")
            if tk_sheet_name in workbook.sheetnames:
                tk_sheet = get_sheet(tk_sheet_name)
                office_name = getattr(dj_settings, "OFFICE_NAME", "居宅介護支援事業所")
                try:
                    write_cell(tk_sheet, tk_coords["care_manager"], assessor_name)
                    write_cell(tk_sheet, tk_coords["client_name"], assessment.client.name)
                    write_cell(tk_sheet, tk_coords["office_name"], office_name)
                except Exception as e:
                    print(f"DEBUG: 担会案内 書き込みエラー: {e}")

            # --- チェックシートへの書き込み ---
            cs_coords = coordinates.get("check_sheet", {})
            cs_sheet_name = cs_coords.get("sheet_name", "チェックシート")
            if cs_sheet_name in workbook.sheetnames:
                cs_sheet = get_sheet(cs_sheet_name)
                write_cell(cs_sheet, cs_coords["client_name"], assessment.client.name)

            # --- HOMEシートへの書き込み ---
            home_coords = coordinates.get("home_sheet", {})
            home_sheet_name = home_coords.get("sheet_name", "HOME")
            if home_sheet_name in workbook.sheetnames:
                home_sheet = get_sheet(home_sheet_name)
                write_cell(home_sheet, home_coords["client_name"], assessment.client.name)

            # --- 居宅届シートへの書き込み ---
            def write_kyotaku(sheet_key):
                kt_coords = coordinates.get(sheet_key, {})
                kt_sheet_name = kt_coords.get("sheet_name", "")
                if not kt_sheet_name or kt_sheet_name not in workbook.sheetnames:
                    return
                kt_sheet = get_sheet(kt_sheet_name)

                def sw(cell_ref, value):
                    try:
                        if cell_ref and value is not None:
                            write_cell(kt_sheet, cell_ref, value)
                    except Exception as e:
                        print(f"DEBUG: {kt_sheet_name} セル {cell_ref} 書き込みエラー: {e}")

                def sw_wareki(cell_ref, value):
                    try:
                        if cell_ref and value is not None:
                                                        write_cell(kt_sheet, cell_ref, str(value))
                    except Exception as e:
                        print(f"DEBUG: {kt_sheet_name} セル {cell_ref} 和暦書き込みエラー: {e}")

                sw(kt_coords.get("client_name"), assessment.client.name)
                sw(kt_coords.get("client_furigana_katakana"), hira_to_kata(assessment.client.furigana or ""))
                sw(kt_coords.get("insurance_number"), assessment.client.insurance_number)
                sw_wareki(kt_coords.get("birth_date"), to_wareki(assessment.client.birth_date) if assessment.client.birth_date else "")
                sw(kt_coords.get("care_manager"), assessor_name)
                sw(kt_coords.get("client_address"), assessment.client.address)

            write_kyotaku("kyotaku_todoke")
            write_kyotaku("kyotaku_todoke_yobou")

            # Excelファイルを保存
            workbook.save(temp_filepath)

            # --- ネットワークドライブへの自動保存ロジック ---
            try:
                if os.path.exists(network_base_path):
                    # サブフォルダが存在しない場合は作成（target_dirは関数冒頭で特定済み）
                    if not os.path.exists(target_dir):
                        try:
                            os.makedirs(target_dir)
                            print(f"SUCCESS: Created directory {target_dir}")
                        except Exception as e:
                            print(f"WARNING: Failed to create directory {target_dir}: {str(e)}")
                            target_dir = network_base_path

                    # target_pathを最終決定
                    target_path = os.path.join(target_dir, save_filename)
                    import shutil

                    shutil.copy2(temp_filepath, target_path)
                    print(f"SUCCESS: Network save (overwrite/create) completed to {target_path}")
                else:
                    print(f"WARNING: Network base path not found: {network_base_path}")
            except Exception as e:
                print(f"WARNING: Network save failed: {str(e)}")
            # --------------------------------------------
            # HTTPレスポンスの準備
            with open(temp_filepath, "rb") as f:
                excel_data = f.read()

            # 一時ファイルの削除
            try:
                os.unlink(temp_filepath)
            except OSError:
                pass

            response = HttpResponse(
                excel_data,
                content_type="application/vnd.ms-excel.sheet.macroEnabled.12",
            )
            # クライアントのふりがな（苗字）と氏名を組み合わせてファイル名を生成
            # 例: やまかわ_山川夏楓様.xlsm
            from urllib.parse import quote

            furigana = assessment.client.furigana or ""
            # スペースで分割して最初の要素（苗字）を取得
            last_name_kana = furigana.split()[0] if furigana.split() else ""
            # 名前から空白を削除
            full_name = (assessment.client.name or "").replace(" ", "").replace("　", "")

            filename = f"{last_name_kana}_{full_name}様.xlsm"
            # 日本語ファイル名を正しく扱うためのエンコーディング
            response["Content-Disposition"] = (
                f"attachment; filename*=UTF-8''{quote(filename)}"
            )

            # AJAXリクエスト（fetchなど）からの呼び出しの場合は、バイナリを返さず成功通知のみ返す（B案: ネットワーク保存のみ）
            if request and request.headers.get("x-requested-with") == "XMLHttpRequest":
                return JsonResponse(
                    {
                        "success": True,
                        "message": "共有ドライブの各担当の利用者フォルダに保存しました。",
                        "filename": filename,
                    }
                )

            return response
    except Exception as e:
        print(f"ERROR: Excel出力エラー: {str(e)}")
        import traceback

        traceback.print_exc()
        raise


@login_required
def assessment_excel_export(request, pk):
    """アセスメント詳細画面からのExcel出力"""
    assessment = get_object_or_404(Assessment, pk=pk)

    try:
        return generate_assessment_excel(assessment, request)
    except Exception as e:
        messages.error(request, f"Excel出力中にエラーが発生しました: {str(e)}")
        return redirect("client_detail", pk=assessment.client.pk)


# ── Excel インポート用モジュールレベル定数 ──────────────────────────────────

_COORD_TO_JSON_FIELD = {
    'personal_hopes': 'basic_info', 'family_hopes': 'basic_info',
    'life_history': 'basic_info', 'notes': 'basic_info',
    'medical_insurance': 'insurance_info', 'disability_handbook': 'insurance_info',
    'disability_type': 'insurance_info', 'difficult_disease': 'insurance_info',
    'life_protection': 'insurance_info', 'disability_level': 'insurance_info',
    'dementia_level': 'insurance_info', 'burden_ratio': 'insurance_info',
    'family_member1_name': 'family_situation', 'family_member1_relation': 'family_situation',
    'family_member1_address': 'family_situation', 'family_member1_contact': 'family_situation',
    'family_member1_care': 'family_situation', 'family_member1_living_together': 'family_situation',
    'family_member1_job': 'family_situation', 'family_member1_notes': 'family_situation',
    'family_member2_name': 'family_situation', 'family_member2_relation': 'family_situation',
    'family_member2_address': 'family_situation', 'family_member2_contact': 'family_situation',
    'family_member2_care': 'family_situation', 'family_member2_living_together': 'family_situation',
    'family_member2_job': 'family_situation', 'family_member2_notes': 'family_situation',
    'home_environment': 'living_situation', 'housing_type': 'living_situation',
    'housing_ownership': 'living_situation', 'private_room': 'living_situation',
    'air_conditioning': 'living_situation', 'toilet_type': 'living_situation',
    'bathroom': 'living_situation', 'sleeping_arrangement': 'living_situation',
    'room_level_difference': 'living_situation', 'housing_modification': 'living_situation',
    'housing_modification_need': 'living_situation', 'special_circumstances': 'living_situation',
    'care_services': 'services', 'welfare_equipment': 'services',
    'other_services': 'services', 'informal_services': 'services',
    'social_relationships': 'services',
    'disease_name_1': 'health_status', 'onset_date_1': 'health_status',
    'disease_name_2': 'health_status', 'onset_date_2': 'health_status',
    'disease_name_3': 'health_status', 'onset_date_3': 'health_status',
    'disease_name_4': 'health_status', 'onset_date_4': 'health_status',
    'disease_name_5': 'health_status', 'onset_date_5': 'health_status',
    'disease_name_6': 'health_status', 'onset_date_6': 'health_status',
    'medical_history': 'health_status', 'special_medical_treatment': 'health_status',
    'main_doctor_hospital': 'health_status', 'main_doctor_name': 'health_status',
    'visiting_doctor_hospital': 'health_status', 'visiting_doctor_name': 'health_status',
    'family_doctor_1': 'health_status', 'family_doctor_2': 'health_status',
    'family_doctor_3': 'health_status', 'family_doctor_4': 'health_status',
    'hospital_visit_info': 'health_status', 'medication_content': 'health_status',
    'has_allergy': 'health_status', 'allergy_details': 'health_status',
    'skin_disease': 'physical_status', 'infection_status': 'physical_status',
    'special_medical_treatment_summary': 'physical_status',
    'paralysis_location': 'physical_status', 'pain_location': 'physical_status',
    'height': 'physical_status', 'weight': 'physical_status',
    'vision': 'physical_status', 'uses_glasses': 'physical_status',
    'hearing': 'physical_status', 'uses_hearing_aid': 'physical_status',
    'physical_notes': 'physical_status',
    'turning_over': 'basic_activities', 'getting_up': 'basic_activities',
    'sitting': 'basic_activities', 'standing_up': 'basic_activities',
    'standing': 'basic_activities', 'transfer': 'basic_activities',
    'indoor_mobility': 'basic_activities', 'indoor_mobility_equipment': 'basic_activities',
    'outdoor_mobility': 'basic_activities', 'outdoor_mobility_equipment': 'basic_activities',
    'basic_activity_notes': 'basic_activities',
    'eating_method': 'adl', 'eating_assistance': 'adl', 'swallowing': 'adl',
    'meal_form_main': 'adl', 'meal_form_side': 'adl', 'water_thickening': 'adl',
    'eating_restriction': 'adl', 'eating_utensils': 'adl', 'eating_notes': 'adl',
    'oral_hygiene_assistance': 'adl', 'natural_teeth_presence': 'adl',
    'denture_type': 'adl', 'denture_location': 'adl', 'oral_notes': 'adl',
    'bathing_assistance': 'adl', 'bathing_form': 'adl', 'bathing_restriction': 'adl',
    'dressing_upper': 'adl', 'dressing_lower': 'adl', 'bathing_notes': 'adl',
    'excretion_assistance': 'adl', 'urination': 'adl', 'urinary_incontinence': 'adl',
    'defecation': 'adl', 'fecal_incontinence': 'adl',
    'daytime_location': 'adl', 'nighttime_location': 'adl',
    'excretion_supplies': 'adl', 'excretion_notes': 'adl',
    'cooking': 'iadl', 'washing': 'iadl', 'money_management': 'iadl',
    'cleaning': 'iadl', 'shopping': 'iadl', 'iadl_notes': 'iadl',
    'dementia_presence': 'cognitive_function', 'dementia_severity': 'cognitive_function',
    'bpsd_symptoms': 'cognitive_function', 'conversation': 'cognitive_function',
    'communication': 'cognitive_function', 'cognitive_notes': 'cognitive_function',
}


# 単一選択フィールド：表示テキスト → choice key（フォームの選択肢コードに合わせる）
_MOBIL = {'自立': 'independent', '見守り': 'supervision', '一部介助': 'partial_assistance', '全介助': 'full_assistance'}
_FIELD_IMPORT_CHOICES = {
    'turning_over':  {'つかまらないでできる': 'no_assistance', '何かにつかまればできる': 'with_assistance', 'できない': 'cannot'},
    'getting_up':    {'つかまらないでできる': 'no_assistance', '何かにつかまればできる': 'with_assistance', 'できない': 'cannot'},
    'standing_up':   {'つかまらないでできる': 'no_assistance', '何かにつかまればできる': 'with_assistance', 'できない': 'cannot'},
    'sitting':       {'できる': 'can_do', '自分の手で支えればできる': 'self_support', '支えてもらえればできる': 'support_needed', 'できない': 'cannot'},
    'standing':      {'支えなしでできる': 'no_support', '何か支えがあればできる': 'with_something', 'できない': 'cannot'},
    'transfer':             _MOBIL,
    'indoor_mobility':      _MOBIL,
    'outdoor_mobility':     _MOBIL,
    'indoor_mobility_equipment':  {'なし': 'none', '車椅子': 'wheelchair', '歩行器': 'walker', '杖': 'cane', '松葉杖': 'crutches', 'その他': 'other'},
    'outdoor_mobility_equipment': {'なし': 'none', '車椅子': 'wheelchair', '歩行器': 'walker', '杖': 'cane', '松葉杖': 'crutches', 'その他': 'other'},
    'eating_method':     {'経口摂取': 'oral', '経管栄養+経口摂取': 'tube_oral', '経管栄養（胃ろう）': 'tube_only', '経管栄養（経鼻）': 'tube_nasal', '経管栄養（腸ろう）': 'tube_gastronomy'},
    'eating_assistance': _MOBIL,
    'swallowing':        {'できる': 'can_do', '見守り等が必要': 'supervision_needed', 'できない': 'cannot'},
    'meal_form_main':    {'普通': 'normal', '軟飯': 'soft', '全粥': 'porridge', 'ペースト': 'paste'},
    'meal_form_side':    {'普通': 'normal', '一口大': 'soft', '粗きザミ': 'minced', 'キザミ': 'paste', 'ペースト': 'paste_form'},
    'water_thickening':  {'あり': 'yes', 'なし': 'no', '有': 'yes', '無': 'no'},
    'eating_restriction':{'あり': 'yes', 'なし': 'no', '不明': 'unknown', '有': 'yes', '無': 'no'},
    'oral_hygiene_assistance': _MOBIL,
    'natural_teeth_presence': {'あり': 'yes', 'なし': 'no', '有': 'yes', '無': 'no'},
    'denture_type':     {'総義歯': 'complete', '部分義歯': 'partial'},
    'denture_location': {'上顎': 'upper', '下顎': 'lower', '上下': 'both'},
    'bathing_assistance':  _MOBIL,
    'bathing_form':        {'一般浴': 'regular_bath', '寝台浴': 'sitting_bath', 'シャワー浴': 'shower_bath', 'チェアー浴': 'chair_bath'},
    'bathing_restriction': {'あり': 'yes', 'なし': 'no', '有': 'yes', '無': 'no'},
    'dressing_upper': _MOBIL,
    'dressing_lower': _MOBIL,
    'excretion_assistance': _MOBIL,
    'urination':             {'あり': 'yes', 'なし': 'no', '有': 'yes', '無': 'no'},
    'urinary_incontinence':  {'あり': 'yes', 'なし': 'no', '有': 'yes', '無': 'no'},
    'defecation':            {'あり': 'yes', 'なし': 'no', '有': 'yes', '無': 'no'},
    'fecal_incontinence':    {'あり': 'yes', 'なし': 'no', '有': 'yes', '無': 'no'},
    'daytime_location':   {'トイレ': 'toilet', 'Pトイレ': 'portable_toilet', 'ベッド': 'bed'},
    'nighttime_location': {'トイレ': 'toilet', 'Pトイレ': 'portable_toilet', 'ベッド': 'bed'},
    'cooking':         _MOBIL,
    'washing':         _MOBIL,
    'money_management':_MOBIL,
    'cleaning':        _MOBIL,
    'shopping':        _MOBIL,
    'dementia_presence': {'あり': 'yes', 'なし': 'no', '有': 'yes', '無': 'no'},
    'dementia_severity': {'軽度': 'mild', '中等度': 'moderate', '重度': 'severe'},
    'conversation':    {'可能': 'possible', '不明瞭': 'unclear', 'やや不自由': 'somewhat_difficult', '全くできない': 'impossible'},
    'communication':   {'可能': 'possible', 'その場のみ可能': 'only_sometimes', 'やや不自由': 'somewhat_difficult', '全くできない': 'impossible'},
    'home_environment':   {'家族と同居': 'family_cohabitation', '一人暮らし': 'living_alone', '高齢世帯': 'elderly_household', '高齢所帯': 'elderly_household', '日中独居': 'daytime_alone', '二世帯住宅': 'two_generation_house', 'その他': 'other'},
    'housing_type':       {'一戸建て': 'detached_house', '集合住宅': 'apartment_complex', '公営住宅': 'public_housing', 'マンション': 'condominium', 'その他': 'other'},
    'housing_ownership':  {'所有': 'owned', '賃貸': 'rental', '間借り': 'lodging', 'その他': 'other'},
    'private_room':           {'あり': 'available', 'なし': 'not_available', '有': 'available', '無': 'not_available'},
    'air_conditioning':       {'あり': 'available', 'なし': 'not_available', '有': 'available', '無': 'not_available'},
    'toilet_type':            {'洋式': 'western', '和式': 'japanese'},
    'bathroom':               {'あり': 'available', 'なし': 'not_available', '有': 'available', '無': 'not_available'},
    'sleeping_arrangement':   {'畳・床': 'tatami_floor', 'ベッド': 'regular_bed', '介護用ベッド': 'care_bed', 'その他': 'other'},
    'room_level_difference':  {'あり': 'available', 'なし': 'not_available', '有': 'available', '無': 'not_available'},
    'housing_modification':   {'あり': 'completed', 'なし': 'not_completed', '有': 'completed', '無': 'not_completed'},
    'housing_modification_need': {'あり': 'needed', 'なし': 'not_needed', '有': 'needed', '無': 'not_needed'},
    'vision':          {'正常': 'normal', '大きい字は可': 'large_letters_ok', 'ほぼ見えない': 'barely_visible', '失明': 'blind'},
    'uses_glasses':    {'使用': 'yes', '未使用': 'no', '有': 'yes', '無': 'no'},
    'hearing':         {'正常': 'normal', '大きい声は可': 'loud_voice_ok', 'ほぼ聞こえない': 'barely_audible', '聞こえない': 'deaf', '失聴': 'deaf'},
    'uses_hearing_aid':{'使用': 'yes', '未使用': 'no', '有': 'yes', '無': 'no'},
}

# 複数選択フィールド（「、」区切り）→ choice key リスト（フォームの選択肢コードに合わせる）
_LIST_IMPORT_CHOICES = {
    'bpsd_symptoms': {
        'なし': 'none', '被害妄想': 'persecution_delusion', '作話': 'confabulation',
        '感情の不安定': 'mood_instability', '昼夜逆転': 'day_night_reversal',
        '帰宅願望': 'home_return_desire', '大声・奇声': 'loud_voice',
        '暴力・暴言': 'violence', '収集癖': 'collection_habit',
        '介護抵抗': 'care_resistance', '落ち着きがない': 'restlessness',
        '徘徊': 'wandering', '破壊行為': 'destructive_behavior',
        'ひどい物忘れ': 'severe_forgetfulness', '自分勝手な行動': 'selfish_behavior',
        '不穏': 'agitation', 'うつ傾向': 'depression_tendency',
    },
    'care_services': {
        '訪問介護': 'home_help', '訪問入浴': 'visit_bath',
        '訪問看護': 'visit_nursing', '訪問リハビリ': 'visit_rehab', '訪問リハ': 'visit_rehab',
        '通所介護': 'day_service', '通所リハビリ': 'day_rehab', '通所リハ': 'day_rehab',
        'ショートステイ': 'short_stay', '小規模多機能': 'small_scale_multi',
    },
    'welfare_equipment': {
        '車いす': 'wheelchair', '車いす付属品': 'walker',
        '特殊寝台': 'special_bed', '特殊寝台付属品': 'special_bed_accessories',
        '床ずれ防止用具': 'fall_prevention', '体位変換器': 'position_changer',
        '手すり': 'walking_aid', 'スロープ': 'slope',
        '歩行器': 'walking_frame', '歩行補助つえ': 'walking_support',
        '排個感知機器': 'detect_sensor', '移動用リフト': 'mobility_lift',
        '自動排泄処理装置': 'automatic_drainage',
    },
    'informal_services': {
        '家族による支援': 'family_support', '近隣による支援': 'neighbor_support',
        'ボランティア': 'volunteer', '地域活動グループ': 'community_group',
        '友人による支援': 'friend_support', 'NPO団体支援': 'npo_support',
        '宗教団体支援': 'religious_support', '食事支援': 'meal_support',
        '排泄支援': 'excretion_support', 'リネンリース': 'linen_lease',
        '洗濯': 'laundry', '居室清掃': 'room_cleaning',
        '喀痰吸引': 'sputum_suction', 'オムツ支給': 'diaper_supply',
        'タクシー券': 'taxi_voucher', 'マッサージ券': 'massage_voucher',
        '火災報知器': 'fire_alarm', '自動消火器': 'auto_fire_extinguisher',
        '老人用電話': 'elderly_phone', '寝具乾燥消毒': 'bedding_disinfection',
        '電磁調理器': 'induction_cooker', '緊急通報装置': 'emergency_alert',
        '配食サービス': 'meal_delivery', 'その他': 'other_services',
    },
    'excretion_supplies': {
        'リハビリパンツ': 'rehabilitation_pants', '紙おむつ': 'paper_diaper',
        '小パット': 'small_pad', '大パット': 'large_pad',
    },
    'eating_utensils': {
        '箸': 'chopsticks', 'スプーン': 'spoon',
        'エプロン': 'apron', '補助具': 'assistive', 'その他': 'other',
    },
    'skin_disease': {
        '床ずれ': 'bedsore', '湿疹': 'eczema', 'かゆみ': 'itching',
        '水虫': 'athletes_foot', '帯状疱疹': 'shingles', 'その他': 'other',
    },
    'infection_status': {
        '結核': 'tuberculosis', '肺炎': 'pneumonia',
        '肝炎': 'hepatitis', '疥癬': 'scabies', 'MRSA': 'mrsa', 'その他': 'other',
    },
}


def _map_import_value(key, val):
    """Excel表示テキスト → DBのchoice key に変換。複数選択はリストで返す"""
    if not val:
        return val
    if key in _LIST_IMPORT_CHOICES:
        choices = _LIST_IMPORT_CHOICES[key]
        # '、' と ' , ' と ',' の両方に対応
        import re as _re
        items = [item.strip() for item in _re.split(r'[、,]', str(val)) if item.strip()]
        return [choices.get(item, item) for item in items]
    if key in _FIELD_IMPORT_CHOICES:
        return _FIELD_IMPORT_CHOICES[key].get(val, val)
    return val


def _parse_wareki(text):
    """和暦/西暦テキスト → date オブジェクト"""
    import re
    from datetime import datetime as dt
    if not text:
        return None
    if isinstance(text, date):
        return text
    if hasattr(text, 'date'):
        return text.date()
    text = str(text).strip()
    for pattern, base in [
        (r'令和(\d+)年(\d+)月(\d+)日', 2018),
        (r'平成(\d+)年(\d+)月(\d+)日', 1988),
        (r'昭和(\d+)年(\d+)月(\d+)日', 1925),
    ]:
        m = re.match(pattern, text)
        if m:
            try:
                return date(base + int(m.group(1)), int(m.group(2)), int(m.group(3)))
            except ValueError:
                pass
    for fmt in ('%Y/%m/%d', '%Y-%m-%d', '%Y年%m月%d日'):
        try:
            return dt.strptime(text, fmt).date()
        except ValueError:
            pass
    return None


def _read_cell(ws, cell_ref):
    """セルの値を取得（MergedCell対応）"""
    try:
        cell = ws[cell_ref]
        if isinstance(cell, MergedCell):
            for rng in ws.merged_cells.ranges:
                if cell_ref in rng:
                    return ws.cell(rng.min_row, rng.min_col).value
            return None
        return cell.value
    except Exception:
        return None


@login_required
def assessment_import(request):
    """ExcelファイルからAssessmentを取り込む"""
    clients = Client.objects.all().order_by('furigana')

    if request.method == 'POST':
        step = request.POST.get('step', 'upload')

        if step == 'upload':
            client_id = request.POST.get('client_id')
            excel_file = request.FILES.get('excel_file')

            if not client_id or not excel_file:
                messages.error(request, '利用者とファイルを選択してください。')
                return render(request, 'assessments/assessment_import.html', {'clients': clients})

            fname = excel_file.name.lower()
            if not (fname.endswith('.xlsx') or fname.endswith('.xlsm')):
                messages.error(request, '.xlsx または .xlsm ファイルを選択してください。')
                return render(request, 'assessments/assessment_import.html', {'clients': clients})

            try:
                client = get_object_or_404(Client, pk=client_id)

                coords_path = os.path.join(
                    settings.BASE_DIR, 'static', 'config', 'excel_coordinates_assessment.json'
                )
                with open(coords_path, 'r', encoding='utf-8') as f:
                    coords = json.load(f)['assessment_sheet']

                wb = load_workbook(excel_file, data_only=True)
                sheet_name = '（最新）アセスメントシート'
                ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.worksheets[0]

                raw = {}
                for key, cell_ref in coords.items():
                    val = _read_cell(ws, cell_ref)
                    raw[key] = str(val).strip() if val is not None else ''
                wb.close()

                # assessment_type 逆マッピング
                type_map = {
                    '新規': 'new', '更新': 'update', '区分変更': 'change',
                    '退院': 'discharge', '情報提供': 'information', 'その他': 'other',
                }
                assessment_type_raw = raw.get('assessment_type', '')
                assessment_type = type_map.get(assessment_type_raw, 'other')
                assessment_type_other = (
                    assessment_type_raw if assessment_type == 'other' and assessment_type_raw else ''
                )

                # interview_location 逆マッピング
                loc_map = {
                    '居室': 'room', '自宅': 'home', '病院': 'hospital',
                    '施設': 'facility', 'その他': 'other',
                }
                interview_location_raw = raw.get('interview_location', '')
                interview_location = loc_map.get(interview_location_raw, 'other')
                interview_location_other = (
                    interview_location_raw if interview_location == 'other' and interview_location_raw else ''
                )

                assessment_date_obj = _parse_wareki(raw.get('assessment_date', ''))

                request.session['assessment_import_data'] = {
                    'client_id': int(client_id),
                    'raw': raw,
                    'assessment_date': assessment_date_obj.isoformat() if assessment_date_obj else None,
                    'assessment_type': assessment_type,
                    'assessment_type_other': assessment_type_other,
                    'interview_location': interview_location,
                    'interview_location_other': interview_location_other,
                }

                # デバッグ用：全フィールドの生の値とセル座標を確認
                debug_raw = {k: f"{coords.get(k, '?')} → {repr(v)}" for k, v in raw.items() if v}

                preview = {
                    'assessment_date': assessment_date_obj.strftime('%Y年%m月%d日') if assessment_date_obj else raw.get('assessment_date', ''),
                    'assessment_type': assessment_type_raw,
                    'interview_location': interview_location_raw,
                    'assessor': raw.get('assessor', ''),
                    'care_level': raw.get('care_level_official', ''),
                    'disease_1': raw.get('disease_name_1', ''),
                    'personal_hopes': raw.get('personal_hopes', '')[:100],
                    'family_hopes': raw.get('family_hopes', '')[:100],
                    'life_history': raw.get('life_history', '')[:100],
                    'debug_raw': debug_raw,
                }

                return render(request, 'assessments/assessment_import.html', {
                    'clients': clients,
                    'selected_client': client,
                    'preview': preview,
                    'step': 'preview',
                })

            except Exception as e:
                messages.error(request, f'ファイルの読み込みに失敗しました: {str(e)}')
                return render(request, 'assessments/assessment_import.html', {'clients': clients})

        elif step == 'confirm':
            import_data = request.session.pop('assessment_import_data', None)
            if not import_data:
                messages.error(request, 'セッションが切れました。再度アップロードしてください。')
                return redirect('assessment_import')

            client = get_object_or_404(Client, pk=import_data['client_id'])
            raw = import_data['raw']

            json_fields = {
                'basic_info': {}, 'insurance_info': {}, 'family_situation': {},
                'living_situation': {}, 'services': {}, 'health_status': {},
                'physical_status': {}, 'basic_activities': {}, 'adl': {},
                'iadl': {}, 'cognitive_function': {},
            }
            for key, val in raw.items():
                field = _COORD_TO_JSON_FIELD.get(key)
                if field and val:
                    json_fields[field][key] = _map_import_value(key, val)

            assessment_date = None
            if import_data.get('assessment_date'):
                try:
                    from datetime import date as date_cls
                    assessment_date = date_cls.fromisoformat(import_data['assessment_date'])
                except ValueError:
                    pass
            if not assessment_date:
                assessment_date = date.today()

            assessment = Assessment(
                client=client,
                assessor=request.user,
                assessment_date=assessment_date,
                assessment_type=import_data['assessment_type'],
                assessment_type_other=import_data.get('assessment_type_other', ''),
                interview_location=import_data['interview_location'],
                interview_location_other=import_data.get('interview_location_other', ''),
                status='draft',
                basic_info=json_fields['basic_info'],
                insurance_info=json_fields['insurance_info'],
                family_situation=json_fields['family_situation'],
                living_situation=json_fields['living_situation'],
                services=json_fields['services'],
                health_status=json_fields['health_status'],
                physical_status=json_fields['physical_status'],
                basic_activities=json_fields['basic_activities'],
                adl=json_fields['adl'],
                iadl=json_fields['iadl'],
                cognitive_function=json_fields['cognitive_function'],
            )
            assessment.save()

            messages.success(request, f'「{client.name}」のアセスメントを取り込みました。内容を確認・編集してください。')
            return redirect('assessment_edit', pk=assessment.pk)

    return render(request, 'assessments/assessment_import.html', {'clients': clients})

@login_required
def assessment_delete(request, pk):
    """アセスメント削除"""
    if request.method == "POST":
        assessment = get_object_or_404(Assessment, pk=pk)
        client_pk = assessment.client.pk
        assessment.delete()
        # 削除成功メッセージをセッションに保存（警告色で表示）
        messages.warning(request, "アセスメントを削除しました")
        return JsonResponse({"success": True, "message": "アセスメントを削除しました"})
    return JsonResponse(
        {"success": False, "message": "不正なリクエストです"}, status=400
    )
