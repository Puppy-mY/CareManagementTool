from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required, user_passes_test

def staff_required(user):
    return user.is_active and (user.is_staff or user.is_superuser)
from django.contrib import messages
from django.core.paginator import Paginator
from django.db.models import Q
from django.http import JsonResponse, HttpResponse, HttpResponseRedirect
from django.urls import reverse
from django.template.loader import render_to_string
from django.conf import settings
from datetime import datetime
import base64
import os
import json
import logging

logger = logging.getLogger(__name__)

# PDF生成ライブラリのインポート
try:
    from weasyprint import HTML, CSS
    WEASYPRINT_AVAILABLE = True
except ImportError:
    WEASYPRINT_AVAILABLE = False

try:
    from reportlab.pdfgen import canvas
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.cidfonts import UnicodeCIDFont
    from PIL import Image
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

# Excel処理ライブラリのインポート
try:
    from openpyxl import load_workbook
    import tempfile
    import json
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
from .models import Client, LimitCalculation, AdditionalService, ServiceType, ServiceProvider, ClientDementiaStatus, ServiceAddOn, ProviderAddOnSetting, DocumentCreationHistory, Feedback, FeedbackReply, HomeCareSupportOffice, RegionalSupportCenter
from .forms import ClientForm, FeedbackForm, FeedbackEditForm, FeedbackReplyForm, CareInsuranceForm, DisabilityWelfareForm, MedicalCertForm


@login_required
def client_list(request):
    from django.contrib.auth.models import User

    # ログインしているケアマネ自身のクライアントのみを取得（クライアントサイドでフィルタリング）
    clients = Client.objects.select_related('created_by').filter(created_by=request.user)

    # 各利用者の今月の限度額試算情報を取得
    now = datetime.now()
    for client in clients:
        try:
            current_calculation = LimitCalculation.objects.get(
                client=client,
                year=now.year,
                month=now.month
            )
            client.current_calculation = current_calculation
        except LimitCalculation.DoesNotExist:
            client.current_calculation = None
    
    # ケアマネージャー一覧を取得（安全な方法）
    care_managers = User.objects.all().order_by('username')

    context = {
        'clients': list(clients), # list()に変換して長さなどをテンプレートで扱いやすくする
        'care_level_choices': Client.CARE_LEVEL_CHOICES,
        'status_choices': [
            ('', '全て'),
            ('over_limit', '限度額超過'),
            ('within_limit', '限度額内'),
            ('no_calculation', '試算未作成'),
        ],
        'care_managers': care_managers,
        'current_year_month': f'{now.year}年{now.month}月',
        'is_all_clients': False, # テンプレートでの表示切り替え用フラグ
    }

    return render(request, 'clients/client_list.html', context)


@login_required
def all_client_list(request):
    """すべての利用者一覧（全ケアマネージャーの利用者を対象）"""
    from django.contrib.auth.models import User

    # 全クライアントを取得（クライアントサイドでフィルタリング）
    clients = Client.objects.select_related('created_by').all()

    # 各利用者の今月の限度額試算情報を取得
    now = datetime.now()
    for client in clients:
        try:
            current_calculation = LimitCalculation.objects.get(
                client=client,
                year=now.year,
                month=now.month
            )
            client.current_calculation = current_calculation
        except LimitCalculation.DoesNotExist:
            client.current_calculation = None
    
    # ケアマネージャー一覧を取得（安全な方法）
    care_managers = User.objects.all().order_by('username')

    context = {
        'clients': list(clients),
        'care_level_choices': Client.CARE_LEVEL_CHOICES,
        'status_choices': [
            ('', '全て'),
            ('over_limit', '限度額超過'),
            ('within_limit', '限度額内'),
            ('no_calculation', '試算未作成'),
        ],
        'care_managers': care_managers,
        'current_year_month': f'{now.year}年{now.month}月',
        'is_all_clients': True, # テンプレートでの表示切り替え用フラグ
    }

    return render(request, 'clients/client_list.html', context)


@login_required
def client_detail(request, pk):
    client = get_object_or_404(Client, pk=pk)
    recent_assessments = client.assessments.order_by('-assessment_date', '-id')

    # 書類作成履歴を取得
    document_histories = client.document_histories.order_by('-created_at')

    context = {
        'client': client,
        'recent_assessments': recent_assessments,
        'document_histories': document_histories,
    }

    return render(request, 'clients/client_detail.html', context)


@login_required
def client_create(request):
    if request.method == 'POST':
        form = ClientForm(request.POST)
        if form.is_valid():
            client = form.save(commit=False)
            client.created_by = request.user
            client.save()
            
            # 自動で当月の限度額試算を作成
            now = datetime.now()
            calculation, created = LimitCalculation.objects.get_or_create(
                client=client,
                year=now.year,
                month=now.month,
                defaults={'care_manager_units': 0}
            )
            if created:
                messages.success(request, f'{client.name}様の利用者情報を保存し、限度額試算を作成しました。')
            else:
                messages.success(request, f'{client.name}様の利用者情報を保存しました。')
                
            return redirect('client_detail', pk=client.pk)
    else:
        form = ClientForm()
    
    context = {
        'form': form,
        'title': '新規利用者登録',
    }
    
    return render(request, 'clients/client_form.html', context)


@login_required
def client_edit(request, pk):
    client = get_object_or_404(Client, pk=pk)
    
    if request.method == 'POST':
        form = ClientForm(request.POST, instance=client)
        if form.is_valid():
            form.save()
            messages.success(request, f'{client.name}様の利用者情報を保存しました。')
            return redirect('client_detail', pk=client.pk)
    else:
        form = ClientForm(instance=client)
    
    context = {
        'form': form,
        'client': client,
        'title': '利用者情報編集',
    }
    
    return render(request, 'clients/client_form.html', context)


@login_required
def client_delete(request, pk):
    client = get_object_or_404(Client, pk=pk)
    
    if request.method == 'POST':
        client_name = client.name
        client.delete()
        messages.success(request, f'利用者「{client_name}」を削除しました。')
        return redirect('client_list')
    
    context = {
        'client': client,
    }
    
    return render(request, 'clients/client_confirm_delete.html', context)


@login_required
def schedule_management(request, pk):
    """保険・受給者証管理（有効期間等の詳細確認・更新）"""
    client = get_object_or_404(Client, pk=pk)

    ITEM_CONFIG = {
        'medical_insurance': {'form': CareInsuranceForm, 'label': '医療保険証'},
        'daily_living': {'form': CareInsuranceForm, 'label': '日常生活の自立度'},
        'certification': {'form': CareInsuranceForm, 'label': '介護保険被保険者証'},
        'burden': {'form': CareInsuranceForm, 'label': '介護保険負担割合証'},
        'limit_cert': {'form': CareInsuranceForm, 'label': '負担限度額認定証'},
        'high_cost_care': {'form': CareInsuranceForm, 'label': '高額介護サービス費'},
        'care_public': {'form': CareInsuranceForm, 'label': '介護保険関係（公的制度）'},
        'disability_handbook': {'form': DisabilityWelfareForm, 'label': '身体障がい者手帳'},
        'disability': {'form': DisabilityWelfareForm, 'label': '障害福祉サービス受給者証'},
        'disability_all': {'form': DisabilityWelfareForm, 'label': '障害福祉関係'},
        'specific_medical': {'form': MedicalCertForm, 'label': '特定医療費受給者証'},
        'welfare_medical': {'form': MedicalCertForm, 'label': '福祉医療費受給者証'},
        'nhi_limit_cert': {'form': MedicalCertForm, 'label': '国保限度額適用認定証'},
        'high_cost_combined': {'form': MedicalCertForm, 'label': '高額医療・高額介護合算療養費'},
        'medical_all': {'form': MedicalCertForm, 'label': '医療関係'},
        'life_protection': {'form': CareInsuranceForm, 'label': '生活保護'},
        'care_insurance_all': {'form': CareInsuranceForm, 'label': '介護保険関係'},
    }

    if request.method == 'POST':
        action = request.POST.get('action', 'save')
        item = request.POST.get('item', '')

        if action == 'quick_add':
            if item == 'limit_cert':
                client.limit_cert = 'yes'
            elif item == 'high_cost_care':
                client.high_cost_care = 'yes'
            elif item == 'disability':
                client.disability_welfare = 'yes'
            elif item == 'specific_medical':
                client.specific_medical = 'yes'
            elif item == 'welfare_medical':
                client.welfare_medical = 'yes'
            elif item == 'nhi_limit_cert':
                client.nhi_limit_cert = 'yes'
            elif item == 'high_cost_combined':
                client.high_cost_combined = 'yes'
            client.save()
            messages.success(request, f'{client.name}様の{ITEM_CONFIG[item]["label"]}を追加しました。詳細は編集ボタンから設定してください。')
            return redirect('schedule_management', pk=client.pk)

        elif action == 'quick_remove':
            if item == 'medical_insurance':
                client.medical_insurance_start = None
                client.medical_insurance_end = None
            elif item == 'limit_cert':
                client.limit_cert = 'no'
                client.limit_cert_start = None
                client.limit_cert_end = None
            elif item == 'high_cost_care':
                client.high_cost_care = 'no'
            elif item == 'disability':
                client.disability_welfare = 'no'
                client.disability_welfare_cert_start = None
                client.disability_welfare_cert_end = None
                client.disability_welfare_decision_start = None
                client.disability_welfare_decision_end = None
            elif item == 'specific_medical':
                client.specific_medical = 'no'
                client.specific_medical_start = None
                client.specific_medical_end = None
                client.difficult_disease_name = ''
            elif item == 'welfare_medical':
                client.welfare_medical = 'no'
                client.welfare_medical_start = None
                client.welfare_medical_end = None
            elif item == 'nhi_limit_cert':
                client.nhi_limit_cert = 'no'
                client.nhi_limit_cert_start = None
                client.nhi_limit_cert_end = None
            elif item == 'high_cost_combined':
                client.high_cost_combined = 'no'
            elif item == 'certification':
                client.care_level = ''
                client.certification_date = None
                client.certification_period_start = None
                client.certification_period_end = None
            elif item == 'burden':
                client.care_burden = ''
                client.burden_period_start = None
                client.burden_period_end = None
            
            client.save()
            label = ITEM_CONFIG[item]['label'] if item in ITEM_CONFIG else item
            messages.success(request, f'{client.name}様の{label}を削除しました。')
            return redirect('schedule_management', pk=client.pk)

        elif action == 'save':
            config = ITEM_CONFIG.get(item)
            if config:
                # 必要なフィールドのみを更新するための定義
                edit_fields = {
                    'medical_insurance': ['medical_insurance_type', 'medical_burden', 'medical_insurer_name_issuer', 'medical_insurer_number', 'medical_insurance_symbol', 'medical_insurance_number', 'medical_insurance_branch', 'medical_insurance_start', 'medical_insurance_end'],
                    'certification': ['care_level', 'certification_date', 'certification_period_start', 'certification_period_end'],
                    'burden': ['care_burden', 'burden_period_start', 'burden_period_end'],
                    'limit_cert': ['limit_cert', 'limit_cert_start', 'limit_cert_end'],
                    'high_cost_care': ['high_cost_care'],
                    'care_public': ['limit_cert', 'limit_cert_start', 'limit_cert_end', 'high_cost_care'],
                    'disability_handbook': ['disability_handbook', 'disability_handbook_type', 'disability_handbook_grade'],
                    'disability': ['disability_welfare', 'disability_welfare_decision_start', 'disability_welfare_decision_end'],
                    'disability_all': ['disability_handbook', 'disability_handbook_type', 'disability_handbook_grade', 'disability_welfare', 'disability_welfare_decision_start', 'disability_welfare_decision_end'],
                    'specific_medical': ['specific_medical', 'specific_medical_start', 'specific_medical_end'],
                    'welfare_medical': ['welfare_medical', 'welfare_medical_start', 'welfare_medical_end'],
                    'nhi_limit_cert': ['nhi_limit_cert', 'nhi_limit_cert_start', 'nhi_limit_cert_end'],
                    'high_cost_combined': ['high_cost_combined'],
                    'medical_all': ['specific_medical', 'specific_medical_start', 'specific_medical_end', 'welfare_medical', 'welfare_medical_start', 'welfare_medical_end', 'nhi_limit_cert', 'nhi_limit_cert_start', 'nhi_limit_cert_end', 'high_cost_combined'],
                    'life_protection': ['life_protection'],
                    'care_insurance_all': ['insurance_number', 'care_level', 'certification_date', 'certification_period_start', 'certification_period_end', 'care_burden', 'burden_period_start', 'burden_period_end'],
                    'daily_living': ['disability_level', 'dementia_level'],
                }

                fields_to_update = edit_fields.get(item, [])
                if fields_to_update:
                    from django.forms import modelform_factory
                    widgets = config['form'].Meta.widgets
                    # 編集対象のフィールドのみを持つ動的フォームを作成（他フィールドの上書きバグ防止）
                    DynamicForm = modelform_factory(
                        Client, 
                        fields=fields_to_update,
                        widgets={k: v for k, v in widgets.items() if k in fields_to_update}
                    )
                    # POSTの認定期間とDBの現在値を文字列で比較（date型変換前）
                    prev_cert_start_str = client.certification_period_start.strftime('%Y-%m-%d') if client.certification_period_start else ''
                    prev_cert_end_str   = client.certification_period_end.strftime('%Y-%m-%d')   if client.certification_period_end   else ''
                    posted_cert_start_str = request.POST.get('certification_period_start', '')
                    posted_cert_end_str   = request.POST.get('certification_period_end', '')
                    form = DynamicForm(request.POST, instance=client)
                    if form.is_valid():
                        form.save()
                        if item == 'daily_living':
                            client.refresh_from_db()
                            Client.objects.filter(pk=client.pk).update(
                                daily_living_assessed_cert_end=client.certification_period_end,
                                prev_disability_level='',
                                prev_dementia_level='',
                            )
                        elif item in ('certification', 'care_insurance_all'):
                            # 認定期間（開始・終了いずれか）が変わった場合は自立度をクリア（前回値を保持）
                            # daily_living_assessed_cert_endに旧終了日をセット→JSが新旧差異を検出して要更新バッジを表示する
                            cert_period_changed = (posted_cert_end_str != prev_cert_end_str) or (posted_cert_start_str != prev_cert_start_str)
                            if cert_period_changed:
                                client.refresh_from_db()
                                import datetime
                                old_cert_end_date = datetime.date.fromisoformat(prev_cert_end_str) if prev_cert_end_str else None
                                Client.objects.filter(pk=client.pk).update(
                                    prev_disability_level=client.disability_level,
                                    prev_dementia_level=client.dementia_level,
                                    disability_level='',
                                    dementia_level='',
                                    daily_living_assessed_cert_end=old_cert_end_date,
                                )
                        messages.success(request, f'{client.name}様の{config["label"]}を更新しました。')
                    else:
                        messages.error(request, f'{config["label"]}の更新中にエラーが発生しました。入力内容を確認してください。')
                return redirect('schedule_management', pk=client.pk)

    care_form = CareInsuranceForm(instance=client)
    disability_form = DisabilityWelfareForm(instance=client)
    medical_form = MedicalCertForm(instance=client)

    from datetime import date, timedelta
    today = date.today()
    warn_date = today + timedelta(days=30)

    context = {
        'care_form': care_form,
        'disability_form': disability_form,
        'medical_form': medical_form,
        'client': client,
        'title': '保険・受給者証管理',
        'today': today,
        'warn_date': warn_date,
        'disability_level_choices': Client.DISABILITY_LEVEL_CHOICES,
        'dementia_level_choices': Client.DEMENTIA_LEVEL_CHOICES,
        'daily_living_needs_update': bool(client.prev_disability_level or client.prev_dementia_level),
    }

    return render(request, 'clients/schedule_management.html', context)


@login_required
def limit_calculation_list(request):
    """限度額試算一覧"""
    from django.contrib.auth.models import User
    
    calculations = LimitCalculation.objects.select_related('client', 'client__created_by').all()
    
    search_query = request.GET.get('search', '')
    if search_query:
        calculations = calculations.filter(
            Q(client__name__icontains=search_query) |
            Q(client__insurance_number__icontains=search_query)
        )
    
    # フィルター機能
    care_level_filter = request.GET.get('care_level', '')
    if care_level_filter:
        calculations = calculations.filter(client__care_level=care_level_filter)
    
    status_filter = request.GET.get('status', '')
    if status_filter == 'over_limit':
        calculations = [c for c in calculations if c.is_over_limit]
    elif status_filter == 'within_limit':
        calculations = [c for c in calculations if not c.is_over_limit]
    
    care_manager_filter = request.GET.get('care_manager', '')
    if care_manager_filter:
        if isinstance(calculations, list):
            calculations = [c for c in calculations if c.client.created_by_id == int(care_manager_filter)]
        else:
            calculations = calculations.filter(client__created_by_id=care_manager_filter)
    
    # 今月のデータのみに絞り込み（デフォルト）
    now = datetime.now()
    show_current_only = request.GET.get('show_current_only', 'true') == 'true'
    if show_current_only:
        if isinstance(calculations, list):
            calculations = [c for c in calculations if c.year == now.year and c.month == now.month]
        else:
            calculations = calculations.filter(year=now.year, month=now.month)
    
    paginator = Paginator(calculations, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # ケアマネージャー一覧を取得（安全な方法）
    care_managers = User.objects.all().order_by('username')
    
    context = {
        'page_obj': page_obj,
        'search_query': search_query,
        'care_level_filter': care_level_filter,
        'status_filter': status_filter,
        'care_manager_filter': care_manager_filter,
        'show_current_only': show_current_only,
        'care_level_choices': Client.CARE_LEVEL_CHOICES,
        'status_choices': [
            ('', '全て'),
            ('over_limit', '限度額超過'),
            ('within_limit', '限度額内'),
        ],
        'care_managers': care_managers,
        'current_year_month': f'{now.year}年{now.month}月',
    }
    
    return render(request, 'clients/limit_calculation_list.html', context)


@login_required
def limit_calculation_detail(request, pk):
    """限度額試算詳細・編集"""
    calculation = get_object_or_404(LimitCalculation, pk=pk)
    service_types = ServiceType.objects.filter(is_active=True).order_by('category', 'name')
    
    # カテゴリ別にグループ化
    service_categories = {}
    for service_type in service_types:
        category = service_type.get_category_display()
        if category not in service_categories:
            service_categories[category] = []
        service_categories[category].append(service_type)
    
    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'update_care_manager_units':
            care_manager_units = request.POST.get('care_manager_units', '0')
            try:
                calculation.care_manager_units = float(care_manager_units)
                calculation.save()
                messages.success(request, 'ケアマネ設定単位を更新しました。')
            except ValueError:
                messages.error(request, '有効な数値を入力してください。')
        
        elif action == 'add_service':
            service_type_id = request.POST.get('service_type')
            date_provided = request.POST.get('date_provided')
            provider_name = request.POST.get('provider_name', '')
            notes = request.POST.get('notes', '')
            
            # 手動時間入力（後方互換性のため残す）
            minutes = request.POST.get('minutes')
            
            # 加算チェックボックス
            has_functional_training = request.POST.get('has_functional_training') == 'on'
            has_bathing = request.POST.get('has_bathing') == 'on'
            has_cognitive_function = request.POST.get('has_cognitive_function') == 'on'
            
            try:
                service_type = ServiceType.objects.get(id=service_type_id)
                
                # デイサービス事業所を取得（最初に見つかった事業所を使用）
                provider = None
                if service_type.category == 'day_service':
                    provider = ServiceProvider.objects.filter(
                        service_category='day_service', 
                        is_active=True
                    ).first()
                
                # サービス種別に時間が設定されていない場合のみ手動入力を使用
                if not service_type.duration_minutes and minutes:
                    minutes = int(minutes)
                else:
                    minutes = None
                
                service = AdditionalService.objects.create(
                    limit_calculation=calculation,
                    service_type=service_type,
                    provider=provider,
                    minutes=minutes,
                    date_provided=date_provided,
                    provider_name=provider_name or (provider.name if provider else ''),
                    has_functional_training=has_functional_training,
                    has_bathing=has_bathing,
                    has_cognitive_function=has_cognitive_function,
                    notes=notes
                )
                
                addon_info = []
                if has_functional_training:
                    addon_info.append('機能訓練加算')
                if has_bathing:
                    addon_info.append('入浴加算')
                if has_cognitive_function:
                    addon_info.append('認知機能加算')
                
                addon_text = f"（{', '.join(addon_info)}付き）" if addon_info else ""
                messages.success(request, f'{service_type.name}{addon_text}を追加しました。')
            except (ValueError, ServiceType.DoesNotExist) as e:
                messages.error(request, 'サービス追加に失敗しました。')
        
        return redirect('limit_calculation_detail', pk=pk)
    
    additional_services = calculation.additional_services.all()
    additional_services_total = sum(float(service.total_units) for service in additional_services)
    
    context = {
        'calculation': calculation,
        'additional_services': additional_services,
        'additional_services_total': additional_services_total,
        'service_categories': service_categories,
        'service_types': service_types,  # テンプレートのJavaScript用に残す
    }
    
    return render(request, 'clients/limit_calculation_detail.html', context)


@login_required
def limit_calculation_create(request, client_id):
    """限度額試算作成"""
    client = get_object_or_404(Client, pk=client_id)
    
    if request.method == 'POST':
        year = request.POST.get('year')
        month = request.POST.get('month')
        care_manager_units = request.POST.get('care_manager_units', '0')
        
        try:
            calculation, created = LimitCalculation.objects.get_or_create(
                client=client,
                year=int(year),
                month=int(month),
                defaults={'care_manager_units': float(care_manager_units)}
            )
            
            if created:
                messages.success(request, f'{client.name}さんの{year}年{month}月の限度額試算を作成しました。')
                return redirect('limit_calculation_detail', pk=calculation.pk)
            else:
                messages.warning(request, 'その期間の試算は既に存在します。')
                return redirect('limit_calculation_detail', pk=calculation.pk)
                
        except ValueError:
            messages.error(request, '有効な値を入力してください。')
    
    now = datetime.now()
    years = list(range(now.year - 1, now.year + 2))
    months = list(range(1, 13))
    
    context = {
        'client': client,
        'years': years,
        'months': months,
        'current_year': now.year,
        'current_month': now.month,
    }
    
    return render(request, 'clients/limit_calculation_create.html', context)


@login_required
def delete_additional_service(request, pk):
    """追加サービス削除（Ajax）"""
    if request.method == 'POST':
        try:
            service = AdditionalService.objects.get(pk=pk)
            service.delete()
            return JsonResponse({'success': True})
        except AdditionalService.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'サービスが見つかりません'})
    
    return JsonResponse({'success': False, 'error': '無効なリクエストです'})


@login_required
def update_dementia_status(request, pk):
    """利用者の認知症状況を更新"""
    client = get_object_or_404(Client, pk=pk)
    
    if request.method == 'POST':
        requires_dementia_care = request.POST.get('requires_dementia_care') == 'on'
        dementia_level = request.POST.get('dementia_level', '')
        notes = request.POST.get('notes', '')
        
        # ClientDementiaStatusレコードを取得または作成
        dementia_status, created = ClientDementiaStatus.objects.get_or_create(
            client=client,
            defaults={
                'requires_dementia_care': requires_dementia_care,
                'dementia_level': dementia_level if dementia_level else '',
                'notes': notes
            }
        )
        
        # 既存レコードの場合は更新
        if not created:
            dementia_status.requires_dementia_care = requires_dementia_care
            dementia_status.dementia_level = dementia_level if dementia_level else ''
            dementia_status.notes = notes
            dementia_status.save()
        
        status_text = "要対応" if requires_dementia_care else "対応不要"
        level_text = f"（レベル{dementia_level}）" if dementia_level else ""
        messages.success(request, f'{client.name}さんの認知症対応設定を「{status_text}」{level_text}に更新しました。')
        
    return redirect('client_detail', pk=pk)


@login_required
def provider_settings(request):
    """事業所別加算設定画面"""
    providers = ServiceProvider.objects.filter(is_active=True)
    
    context = {
        'providers': providers,
    }
    
    return render(request, 'clients/provider_settings.html', context)


@login_required
def provider_addon_settings(request, provider_id):
    """特定事業所の加算設定画面"""
    provider = get_object_or_404(ServiceProvider, pk=provider_id)
    addons = ServiceAddOn.objects.filter(is_active=True).order_by('addon_type', 'name')
    
    # 現在の設定を取得
    current_settings = {}
    for setting in ProviderAddOnSetting.objects.filter(provider=provider):
        current_settings[setting.addon.pk] = setting.is_enabled
    
    if request.method == 'POST':
        # 既存の設定をすべて削除
        ProviderAddOnSetting.objects.filter(provider=provider).delete()
        
        # 新しい設定を作成
        updated_count = 0
        for addon in addons:
            is_enabled = request.POST.get(f'addon_{addon.pk}') == 'on'
            if is_enabled:  # チェックされた加算のみ保存
                ProviderAddOnSetting.objects.create(
                    provider=provider,
                    addon=addon,
                    is_enabled=True
                )
                updated_count += 1
        
        messages.success(request, f'{provider.name}の加算設定を更新しました（{updated_count}件の加算を有効化）。')
        return redirect('provider_addon_settings', provider_id=provider.pk)
    
    context = {
        'provider': provider,
        'addons': addons,
        'current_settings': current_settings,
    }
    
    return render(request, 'clients/provider_addon_settings.html', context)


@login_required
def document_create(request, client_id, document_type):
    """書類作成画面"""
    client = get_object_or_404(Client, pk=client_id)

    # 書類タイプの検証
    document_types = {
        'kyotaku_service_plan_request': '居宅サービス計画作成依頼（変更）届出書',
        'kyotaku_preventive_service_plan_request': '介護予防サービス計画作成・介護予防ケアマネジメント依頼（変更）届出書',
        'careplan_info_request': '介護サービス計画作成に係る資料提供申請書',
    }

    if document_type not in document_types:
        messages.error(request, '指定された書類タイプが見つかりません。')
        return redirect('client_detail', pk=client_id)

    document_name = document_types[document_type]

    # 履歴からの編集データ読み込み
    initial_data = {}
    history_id = request.GET.get('history_id')
    if history_id:
        try:
            history = DocumentCreationHistory.objects.get(pk=history_id, client=client)
            initial_data = history.form_data

            # 編集時に最新のマスタデータを反映
            # 地域包括支援センターのマスタデータを更新
            if document_type == 'kyotaku_preventive_service_plan_request' and initial_data.get('preventive_support_provider'):
                from .models import RegionalSupportCenter
                try:
                    center = RegionalSupportCenter.objects.get(
                        name=initial_data.get('preventive_support_provider'),
                        is_active=True
                    )
                    initial_data['houkatsu_office_number'] = center.office_number
                    initial_data['houkatsu_postal_code'] = center.postal_code or ''
                    initial_data['houkatsu_address'] = center.address or ''
                    initial_data['houkatsu_phone'] = center.phone or ''
                except RegionalSupportCenter.DoesNotExist:
                    pass  # マスタデータが見つからない場合は履歴データをそのまま使用

            # 居宅介護支援事業所のマスタデータを更新（介護予防版の受託事業所セクション）
            if document_type == 'kyotaku_preventive_service_plan_request':
                from .models import HomeCareSupportOffice

                # provider_category が 'anoutsu' の場合、ユーザーの所属事業所を取得
                provider_category = initial_data.get('provider_category')
                logger.info(f'[DEBUG] provider_category: {provider_category}')
                logger.info(f'[DEBUG] jutaku_provider: {initial_data.get("jutaku_provider")}')

                if provider_category == 'anoutsu' and request.user.profile.home_care_office:
                    # データベースから最新の値を取得
                    office = HomeCareSupportOffice.objects.get(pk=request.user.profile.home_care_office.pk)
                    initial_data['jutaku_office_number'] = office.office_number
                    initial_data['jutaku_postal_code'] = office.postal_code or ''
                    initial_data['jutaku_address'] = office.address or ''
                    initial_data['jutaku_phone'] = office.phone or ''
                    logger.info(f'[DEBUG] Updated jutaku info from user office: {office.name}')
                    logger.info(f'[DEBUG] jutaku_office_number: {office.office_number}')
                    logger.info(f'[DEBUG] jutaku_postal_code: {office.postal_code}')
                    logger.info(f'[DEBUG] jutaku_address: {office.address}')
                    logger.info(f'[DEBUG] jutaku_phone: {office.phone}')
                elif initial_data.get('jutaku_provider'):
                    # jutaku_providerの値を確認
                    jutaku_provider_value = initial_data.get('jutaku_provider')

                    # 'anoutsu_roman'という固定値の場合も、ユーザーの所属事業所を取得
                    if jutaku_provider_value == 'anoutsu_roman' and request.user.profile.home_care_office:
                        # データベースから最新の値を取得
                        office = HomeCareSupportOffice.objects.get(pk=request.user.profile.home_care_office.pk)
                        initial_data['jutaku_office_number'] = office.office_number
                        initial_data['jutaku_postal_code'] = office.postal_code or ''
                        initial_data['jutaku_address'] = office.address or ''
                        initial_data['jutaku_phone'] = office.phone or ''
                        logger.info(f'[DEBUG] Updated jutaku info from anoutsu_roman: {office.name}')
                        logger.info(f'[DEBUG] jutaku_office_number: {office.office_number}')
                        logger.info(f'[DEBUG] jutaku_postal_code: {office.postal_code}')
                        logger.info(f'[DEBUG] jutaku_address: {office.address}')
                        logger.info(f'[DEBUG] jutaku_phone: {office.phone}')
                    else:
                        # 事業所名で検索する場合
                        try:
                            office = HomeCareSupportOffice.objects.get(
                                name=jutaku_provider_value,
                                is_active=True
                            )
                            initial_data['jutaku_office_number'] = office.office_number
                            initial_data['jutaku_postal_code'] = office.postal_code or ''
                            initial_data['jutaku_address'] = office.address or ''
                            initial_data['jutaku_phone'] = office.phone or ''
                            logger.info(f'[DEBUG] Updated jutaku info from office name: {office.name}')
                        except HomeCareSupportOffice.DoesNotExist:
                            logger.warning(f'[DEBUG] Office not found: {jutaku_provider_value}')
                            pass  # マスタデータが見つからない場合は履歴データをそのまま使用

            # 居宅介護支援事業所のマスタデータを更新（居宅版）
            if document_type == 'kyotaku_service_plan_request' and request.user.profile.home_care_office:
                office = request.user.profile.home_care_office
                initial_data['kyotaku_office_number'] = office.office_number
                initial_data['kyotaku_postal_code'] = office.postal_code or ''
                initial_data['kyotaku_address'] = office.address or ''
                initial_data['kyotaku_phone'] = office.phone or ''

        except DocumentCreationHistory.DoesNotExist:
            messages.warning(request, '指定された履歴データが見つかりません。')

    # initial_dataがある場合でも介護保険情報の表示用フィールドを最新クライアントデータで補完
    # （古い履歴データには _display / _iso フィールドがない場合がある）
    if initial_data and not initial_data.get('client_care_level_display'):
        initial_data['client_care_level'] = client.care_level or ''
        initial_data['client_care_level_display'] = client.get_care_level_display() if client.care_level else ''
        initial_data['client_care_burden'] = client.care_burden or ''
        # 日付変換は _to_wareki 定義後に行うため後で補完

    # initial_dataがない場合のデフォルト値を設定
    def _to_wareki(date_obj):
        if not date_obj:
            return '', ''
        iso = date_obj.strftime('%Y-%m-%d')
        y, m, d = date_obj.year, date_obj.month, date_obj.day
        for era_name, era_start in [('令和', 2019), ('平成', 1989), ('昭和', 1926), ('大正', 1912), ('明治', 1868)]:
            if y >= era_start:
                return era_name + str(y - era_start + 1) + '年' + str(m) + '月' + str(d) + '日', iso
        return iso, iso

    # 履歴データで日付 _display / _iso が欠けている場合の補完
    if initial_data and not initial_data.get('client_certification_date_iso'):
        cert_date_disp, cert_date_iso = _to_wareki(client.certification_date)
        cert_start_disp, cert_start_iso = _to_wareki(client.certification_period_start)
        cert_end_disp, cert_end_iso = _to_wareki(client.certification_period_end)
        initial_data['client_certification_date_display'] = cert_date_disp
        initial_data['client_certification_date_iso'] = cert_date_iso
        initial_data['client_certification_period_start_display'] = cert_start_disp
        initial_data['client_certification_period_start_iso'] = cert_start_iso
        initial_data['client_certification_period_end_display'] = cert_end_disp
        initial_data['client_certification_period_end_iso'] = cert_end_iso

    if not initial_data:
        cert_date_disp, cert_date_iso = _to_wareki(client.certification_date)
        cert_start_disp, cert_start_iso = _to_wareki(client.certification_period_start)
        cert_end_disp, cert_end_iso = _to_wareki(client.certification_period_end)
        initial_data = {
            'client_name': client.name or '',
            'client_furigana': client.furigana or '',
            'client_birth_date': client.birth_date.strftime('%Y-%m-%d') if client.birth_date else '',
            'client_postal_code': client.postal_code or '',
            'client_address': client.address or '',
            'client_phone': client.phone or '',
            'client_gender': '男' if client.gender == 'male' else ('女' if client.gender == 'female' else ''),
            'client_insurance_number': client.insurance_number or '',
            'client_care_level': client.care_level or '',
            'client_care_level_display': client.get_care_level_display() if client.care_level else '',
            'client_certification_date_display': cert_date_disp,
            'client_certification_date_iso': cert_date_iso,
            'client_certification_period_start_display': cert_start_disp,
            'client_certification_period_start_iso': cert_start_iso,
            'client_certification_period_end_display': cert_end_disp,
            'client_certification_period_end_iso': cert_end_iso,
            'client_care_burden': client.care_burden or '',
        }

    # 事業所情報を initial_data に設定
    # ユーザープロフィールの所属事業所、なければDBの最初の有効事業所を使用
    if document_type in ('kyotaku_service_plan_request', 'kyotaku_preventive_service_plan_request'):
        from .models import HomeCareSupportOffice
        office = getattr(getattr(request.user, 'profile', None), 'home_care_office', None)
        if not office:
            office = HomeCareSupportOffice.objects.filter(is_active=True).order_by('name').first()
        if office:
            # 常にセット（表示名・モーダルAPI呼び出しに必要）
            initial_data['kyotaku_office_id'] = office.pk
            initial_data['kyotaku_office_name'] = office.name or ''
            initial_data['kyotaku_office_furigana'] = office.furigana or ''
            initial_data['kyotaku_office_fax'] = office.fax or ''
            initial_data['kyotaku_office_manager'] = office.manager_name or ''
            # 他のフィールドは未設定の場合のみセット（履歴データを優先）
            if not initial_data.get('kyotaku_office_number'):
                initial_data['kyotaku_office_number'] = office.office_number
                initial_data['kyotaku_postal_code'] = office.postal_code or ''
                initial_data['kyotaku_address'] = office.address or ''
                initial_data['kyotaku_phone'] = office.phone or ''

    if request.method == 'POST':
        # フォームデータの取得
        form_data = {
            # 基本情報
            'application_date': request.POST.get('application_date', ''),
            'kubun': request.POST.get('kubun'),
            'client_name': request.POST.get('client_name'),
            'client_insurance_number': request.POST.get('client_insurance_number'),
            'client_furigana': request.POST.get('client_furigana'),
            'personal_number': request.POST.get('personal_number'),
            'client_birth_date': request.POST.get('client_birth_date'),
            'client_postal_code': request.POST.get('client_postal_code'),
            'client_address': request.POST.get('client_address'),
            'client_phone': request.POST.get('client_phone'),
            'client_gender': request.POST.get('client_gender'),
            # 介護保険情報
            'client_care_level': request.POST.get('client_care_level', ''),
            'client_care_level_display': request.POST.get('client_care_level_display', ''),
            'client_certification_date_display': request.POST.get('client_certification_date_display', ''),
            'client_certification_date_iso': request.POST.get('client_certification_date_iso', ''),
            'client_certification_period_start_display': request.POST.get('client_certification_period_start_display', ''),
            'client_certification_period_start_iso': request.POST.get('client_certification_period_start_iso', ''),
            'client_certification_period_end_display': request.POST.get('client_certification_period_end_display', ''),
            'client_certification_period_end_iso': request.POST.get('client_certification_period_end_iso', ''),
            'client_care_burden': request.POST.get('client_care_burden', ''),

            # 介護予防支援事業者（介護予防版のみ）
            'provider_category': request.POST.get('provider_category'),
            'preventive_support_provider': request.POST.get('preventive_support_provider'),
            # 地域包括支援センター詳細情報
            'houkatsu_office_number': request.POST.get('houkatsu_office_number'),
            'houkatsu_postal_code': request.POST.get('houkatsu_postal_code'),
            'houkatsu_address': request.POST.get('houkatsu_address'),
            'houkatsu_phone': request.POST.get('houkatsu_phone'),
            'houkatsu_staff_name': request.POST.get('houkatsu_staff_name'),

            # 介護予防支援を受託する居宅介護支援事業者
            'jutaku_provider': request.POST.get('jutaku_provider'),
            'jutaku_provider_name': request.POST.get('jutaku_provider_name', ''),
            'jutaku_office_number': request.POST.get('jutaku_office_number'),
            'jutaku_postal_code': request.POST.get('jutaku_postal_code'),
            'jutaku_address': request.POST.get('jutaku_address'),
            'jutaku_phone': request.POST.get('jutaku_phone'),
            'jutaku_staff_name': request.POST.get('jutaku_staff_name'),

            # 事業者情報
            'office_name': request.POST.get('office_name'),
            'office_address': request.POST.get('office_address'),
            'office_phone': request.POST.get('office_phone'),
            'care_manager': request.POST.get('care_manager'),

            # 居宅介護支援事業者（居宅版）
            'kyotaku_provider': request.POST.get('kyotaku_provider'),
            'kyotaku_office_number': request.POST.get('kyotaku_office_number'),
            'kyotaku_postal_code': request.POST.get('kyotaku_postal_code'),
            'kyotaku_address': request.POST.get('kyotaku_address'),
            'kyotaku_phone': request.POST.get('kyotaku_phone'),
            'kyotaku_staff_name': request.POST.get('kyotaku_staff_name'),

            # 情報提供依頼書（申請者・担当者情報）
            'office_name_type': request.POST.get('office_name_type', 'anoutsu'),
            'staff_name': request.POST.get('staff_name'),
            'staff_job_title': request.POST.get('staff_job_title'),
            'manager_name': request.POST.get('manager_name'),
            'request_doc_chousasho': request.POST.get('request_doc_chousasho'),
            'request_doc_ikensho': request.POST.get('request_doc_ikensho'),

            # 事業所番号
            'office_number_1': request.POST.get('office_number_1'),
            'office_number_2': request.POST.get('office_number_2'),
            'office_number_3': request.POST.get('office_number_3'),
            'office_number_4': request.POST.get('office_number_4'),
            'office_number_5': request.POST.get('office_number_5'),
            'office_number_6': request.POST.get('office_number_6'),
            'office_number_7': request.POST.get('office_number_7'),
            'office_number_8': request.POST.get('office_number_8'),
            'office_number_9': request.POST.get('office_number_9'),
            'office_number_10': request.POST.get('office_number_10'),

            # 事由（チェックボックス配列から判定）
            'reason_new_or_change': 'new_or_change' in request.POST.getlist('request_reason'),
            'reason_level_change': 'level_change' in request.POST.getlist('request_reason'),
            'reason_provisional_different': 'provisional_different' in request.POST.getlist('request_reason'),
            'reason_other': 'other' in request.POST.getlist('request_reason'),
            'other_reason_detail': request.POST.get('other_reason_detail'),
            'effective_start_date': request.POST.get('effective_start_date'),

            # 届出区分
            'notification_type': request.POST.getlist('notification_type'),

            # ケアマネ担当者名
            'care_manager_name': request.POST.get('care_manager_name'),

            # 負担割合証
            'burden_delivery_method': request.POST.get('burden_delivery_method', ''),

            # 届出者情報
            'application_era': request.POST.get('application_era'),
            'application_year': request.POST.get('application_year'),
            'application_month': request.POST.get('application_month'),
            'application_day': request.POST.get('application_day'),
            'applicant_address': request.POST.get('applicant_address'),
            'applicant_name': request.POST.get('applicant_name'),
            'applicant_phone': request.POST.get('applicant_phone'),

            # 確認物
            'application_method': request.POST.get('application_method'),
            'confirmation_2points': request.POST.get('confirmation_2points') == 'on',
            'confirmation_1point': request.POST.get('confirmation_1point') == 'on',
            'confirmation_private_only': request.POST.get('confirmation_private_only') == 'on',
            'insurance_other': request.POST.get('insurance_other'),
            'id_other': request.POST.get('id_other'),
        }

        # 介護予防支援事業者が「安濃津ろまん」の場合、受託事業者のデータをクリア
        if form_data.get('provider_category') == 'anoutsu':
            form_data['jutaku_provider'] = ''
            form_data['jutaku_office_number'] = ''
            form_data['jutaku_postal_code'] = ''
            form_data['jutaku_address'] = ''
            form_data['jutaku_phone'] = ''
            form_data['jutaku_staff_name'] = ''

        # 作成履歴の保存（「保存」アクション時のみ）
        action = request.POST.get('action', 'create')

        if action == 'save':
            # 保存時のみDocumentCreationHistoryに保存
            # 編集の場合は既存の履歴を更新、新規の場合は新規作成
            if history_id:
                try:
                    history = DocumentCreationHistory.objects.get(pk=history_id, client=client)
                    history.form_data = form_data
                    history.save()
                    messages.success(request, '書類を更新しました。')
                except DocumentCreationHistory.DoesNotExist:
                    messages.error(request, '編集対象の履歴が見つかりません。')
                    url = reverse('client_detail', kwargs={'pk': client_id}) + '#documents'
                    return HttpResponseRedirect(url)
            else:
                history = DocumentCreationHistory.objects.create(
                    client=client,
                    document_type=document_type,
                    document_name=document_name,
                    form_data=form_data,
                    status='draft',
                    created_by=request.user
                )
                messages.success(request, '書類を保存しました。')

            url = reverse('client_detail', kwargs={'pk': client_id}) + '#documents'
            return HttpResponseRedirect(url)

        # Excel出力処理
        if action == 'excel':
            # Excel出力時も履歴を保存・更新
            if history_id:
                try:
                    history = DocumentCreationHistory.objects.get(pk=history_id, client=client)
                    history.form_data = form_data
                    history.save()
                except DocumentCreationHistory.DoesNotExist:
                    # 履歴が見つからない場合は新規作成
                    DocumentCreationHistory.objects.create(
                        client=client,
                        document_type=document_type,
                        document_name=document_name,
                        form_data=form_data,
                        status='draft',
                        created_by=request.user
                    )
            else:
                DocumentCreationHistory.objects.create(
                    client=client,
                    document_type=document_type,
                    document_name=document_name,
                    form_data=form_data,
                    status='draft',
                    created_by=request.user
                )
            return generate_document_excel(request, client, document_type, document_name, form_data)
        else:
            messages.error(request, '無効なアクションです。')
            return redirect('client_detail', pk=client_id)
    
    # 介護予防版の場合、地域包括支援センターデータを取得
    support_centers_data = []
    if document_type == 'kyotaku_preventive_service_plan_request':
        from .models import RegionalSupportCenter
        support_centers = RegionalSupportCenter.objects.filter(is_active=True).order_by('area', 'name')
        for center in support_centers:
            support_centers_data.append({
                'id': center.id,
                'name': center.name,
                'office_number': center.office_number,
                'postal_code': center.postal_code or '',
                'address': center.address or '',
                'phone': center.phone or '',
                'area': center.area,
            })

    context = {
        'client': client,
        'document_type': document_type,
        'document_name': document_name,
        'initial_data': initial_data,
        'support_centers_json': json.dumps(support_centers_data, ensure_ascii=False),
    }

    # ドキュメントタイプに応じてテンプレートを選択
    if document_type == 'kyotaku_preventive_service_plan_request':
        template_name = 'clients/document_create_preventive.html'
    elif document_type == 'careplan_info_request':
        template_name = 'clients/document_create_careplan_info_request.html'
    else:
        template_name = 'clients/document_create_kyotaku.html'

    return render(request, template_name, context)


def generate_document_excel(request, client, document_type, document_name, form_data):
    """Excel出力機能"""
    if not OPENPYXL_AVAILABLE:
        messages.error(request, 'Excel出力機能は現在利用できません。openpyxlライブラリのインストールが必要です。')
        return redirect('client_detail', pk=client.pk)

    try:
        # 座標設定ファイルを読み込み
        coordinates_path = os.path.join(settings.BASE_DIR, 'static', 'config', 'excel_coordinates.json')
        with open(coordinates_path, 'r', encoding='utf-8') as f:
            coordinates = json.load(f)

        if document_type not in coordinates:
            messages.error(request, f'Excel座標設定が見つかりません: {document_type}')
            return redirect('client_detail', pk=client.pk)

        coords = coordinates[document_type]

        # テンプレートファイルのパス（.xlsxと.xlsmの両方をサポート）
        template_path_xlsx = os.path.join(settings.BASE_DIR, 'templates', 'forms', f'{document_type}.xlsx')
        template_path_xlsm = os.path.join(settings.BASE_DIR, 'templates', 'forms', f'{document_type}.xlsm')

        if os.path.exists(template_path_xlsm):
            template_path = template_path_xlsm
            is_macro_enabled = True
        elif os.path.exists(template_path_xlsx):
            template_path = template_path_xlsx
            is_macro_enabled = False
        else:
            messages.error(request, f'Excelテンプレートが見つかりません')
            return redirect('client_detail', pk=client.pk)

        # Excelファイルを読み込み（マクロ付きの場合のみkeep_vba=True）
        workbook = load_workbook(template_path, keep_vba=is_macro_enabled)
        worksheet = workbook.active

        # 外部リンクを削除（保存時のエラー回避）
        if hasattr(workbook, '_external_links'):
            workbook._external_links = []

        # セルへの安全な書き込み関数（MergedCell対応）
        def safe_write_cell(cell_ref, value, force_number=False):
            try:
                if cell_ref and value is not None:
                    from openpyxl.cell.cell import MergedCell
                    cell = worksheet[cell_ref]

                    # MergedCellの場合、マージ範囲の最初のセルを見つけて書き込み
                    if isinstance(cell, MergedCell):
                        # マージされたセル範囲を取得
                        for merged_range in worksheet.merged_cells.ranges:
                            if cell_ref in merged_range:
                                # マージ範囲の左上セルを取得
                                min_col, min_row, max_col, max_row = merged_range.bounds
                                target_cell = worksheet.cell(row=min_row, column=min_col)
                                target_cell.value = value
                                if force_number and isinstance(value, (int, float)):
                                    target_cell.number_format = '0'
                                return
                    else:
                        # 通常のセル
                        cell.value = value
                        if force_number and isinstance(value, (int, float)):
                            cell.number_format = '0'
            except Exception as e:
                pass  # セル書き込みエラーは無視

        # ふりがな（ひらがな）をカタカナに変換する関数
        def hiragana_to_katakana(text):
            if not text:
                return text
            result = ""
            for char in text:
                # ひらがなをカタカナに変換（ひらがな: U+3041-U+3096, カタカナ: U+30A1-U+30F6）
                if 'あ' <= char <= 'ん':
                    result += chr(ord(char) + 0x60)
                else:
                    result += char
            return result

        # 和暦変換関数
        def to_wareki(date):
            if not date:
                return ""
            from datetime import date as date_class
            year = date.year
            month = date.month
            day = date.day

            # 令和: 2019年5月1日〜
            if date >= date_class(2019, 5, 1):
                wareki_year = year - 2018
                era = "令和"
            # 平成: 1989年1月8日〜2019年4月30日
            elif date >= date_class(1989, 1, 8):
                wareki_year = year - 1988
                era = "平成"
            # 昭和: 1926年12月25日〜1989年1月7日
            elif date >= date_class(1926, 12, 25):
                wareki_year = year - 1925
                era = "昭和"
            # 大正: 1912年7月30日〜1926年12月24日
            elif date >= date_class(1912, 7, 30):
                wareki_year = year - 1911
                era = "大正"
            # 明治
            else:
                wareki_year = year - 1867
                era = "明治"

            if wareki_year == 1:
                return f"{era}元年{month}月{day}日"
            return f"{era}{wareki_year}年{month}月{day}日"

        # 基本情報を書き込み
        safe_write_cell(coords.get('client_name'), client.name)
        safe_write_cell(coords.get('client_furigana'), hiragana_to_katakana(client.furigana))
        # 保険番号の書き込み（フォームで編集された場合はそれを使用）
        insurance_number = form_data.get('client_insurance_number') or client.insurance_number
        if insurance_number:
            # 1桁ずつ書き込み（P5, R5, T5, V5, X5, Z5, AB5, AD5, AF5, AH5）
            insurance_str = str(insurance_number).strip()
            # 10桁分のセルに右寄せで配置
            total_cells = 10
            # 右寄せのため、左側に空白を追加
            padded_insurance = insurance_str.rjust(total_cells)
            for i, char in enumerate(padded_insurance):
                cell_key = f'insurance_number_{i + 1}'
                if char and char.isdigit():
                    safe_write_cell(coords.get(cell_key), int(char), force_number=True)
        if client.birth_date:
            safe_write_cell(coords.get('birth_date'), to_wareki(client.birth_date))
        # 住所（フォームで編集された場合はそれを使用、そうでなければクライアントの住所）
        address = form_data.get('address') if form_data.get('address') else client.address
        safe_write_cell(coords.get('client_address'), address)

        # フォームデータを書き込み
        safe_write_cell(coords.get('care_manager_name'), form_data.get('care_manager_name', ''))

        # 有効開始日を和暦に変換
        effective_date_str = form_data.get('effective_start_date', '')
        if effective_date_str:
            try:
                from datetime import datetime
                effective_date = datetime.strptime(effective_date_str, '%Y-%m-%d').date()
                safe_write_cell(coords.get('effective_start_date'), to_wareki(effective_date))
            except ValueError:
                safe_write_cell(coords.get('effective_start_date'), effective_date_str)

        # その他の事由の詳細（その他にチェックが入っている場合のみ）
        if form_data.get('reason_other') and form_data.get('other_reason_detail'):
            safe_write_cell(coords.get('reason_other_detail'), form_data.get('other_reason_detail', ''))

        # 届出区分を文字列で書き込み（AF3）
        notification_type = form_data.get('notification_type', [])
        if isinstance(notification_type, list) and len(notification_type) > 0:
            notification_type = notification_type[0]  # 配列の最初の要素を使用
        elif not isinstance(notification_type, str):
            notification_type = ''

        notification_text = ""
        if notification_type == 'new':
            notification_text = "新規"
        elif notification_type == 'change':
            notification_text = "変更"

        if notification_text:
            safe_write_cell(coords.get('care_level'), notification_text)

        # チェックボックス項目の処理
        if form_data.get('reason_new_or_change'):
            safe_write_cell(coords.get('reason_new_or_change'), '☑')
        if form_data.get('reason_level_change'):
            safe_write_cell(coords.get('reason_level_change'), '☑')
        if form_data.get('reason_provisional_different'):
            safe_write_cell(coords.get('reason_provisional_different'), '☑')
        if form_data.get('reason_other'):
            safe_write_cell(coords.get('reason_other'), '☑')

        # 負担割合のチェックボックス項目の処理
        burden_delivery = form_data.get('burden_delivery_method', '')

        if burden_delivery == 'care_manager':
            safe_write_cell(coords.get('burden_agree_provider'), '☑')
        elif burden_delivery == 'mail':
            safe_write_cell(coords.get('burden_agree_mail'), '☑')
        elif burden_delivery == 'with_result':
            safe_write_cell(coords.get('burden_agree_simultaneous'), '☑')

        elif document_type == 'careplan_info_request':
            # 介護サービス計画作成に係る資料提供申請書専用ロジック
            app_date_iso = form_data.get('application_date', '')
            if app_date_iso:
                from datetime import date as _date
                try:
                    y, m, d = app_date_iso.split('-')
                    app_date_wareki = to_wareki(_date(int(y), int(m), int(d)))
                except Exception:
                    app_date_wareki = app_date_iso
            else:
                from datetime import datetime
                app_date_wareki = to_wareki(datetime.now().date())
            safe_write_cell(coords.get('request_date'), app_date_wareki)

            # 申請者情報
            if form_data.get('office_name_type', 'anoutsu') == 'anoutsu':
                office_name_val = '居宅介護支援事業所　安濃津ろまん'
            else:
                office_name_val = form_data.get('office_name', '')
            safe_write_cell(coords.get('office_name'), office_name_val)
            safe_write_cell(coords.get('manager_name'), form_data.get('manager_name', ''))
            safe_write_cell(coords.get('office_phone'), form_data.get('office_phone', ''))

            # 提供依頼資料のチェックボックス表現
            if form_data.get('request_doc_chousasho') == 'yes':
                safe_write_cell(coords.get('request_doc_chousasho'), '■\u3000介護保険の認定に係る認定調査票')
            if form_data.get('request_doc_ikensho') == 'yes':
                safe_write_cell(coords.get('request_doc_ikensho'), '■\u3000介護保険の認定に係る主治医意見書')

            # 提供申請者情報
            safe_write_cell(coords.get('staff_name'), form_data.get('staff_name', ''))
            safe_write_cell(coords.get('staff_job_title'), form_data.get('staff_job_title', ''))

            # 被保険者情報（同意欄）
            safe_write_cell(coords.get('client_address'), form_data.get('client_address', client.address))
            if client.birth_date:
                safe_write_cell(coords.get('client_birth_date'), to_wareki(client.birth_date))
            ins_num_raw = form_data.get('client_insurance_number') or client.insurance_number or ''
            try:
                ins_num_val = int(str(ins_num_raw).replace('-', '').strip())
            except (ValueError, TypeError):
                ins_num_val = ins_num_raw
            safe_write_cell(coords.get('client_insurance_number'), ins_num_val, force_number=True)
            safe_write_cell(coords.get('client_name'), form_data.get('client_name', client.name))

        # 介護予防支援事業者の情報（介護予防版のみ）
        if document_type == 'kyotaku_preventive_service_plan_request':
            # 利用者住所（フォームで編集された場合はそれを使用、そうでなければクライアントの住所）
            client_address = form_data.get('client_address') if form_data.get('client_address') else client.address
            safe_write_cell(coords.get('client_address'), client_address)

            provider_category = form_data.get('provider_category', '')

            # 事業所名
            if provider_category == 'anoutsu':
                provider_name = '居宅介護支援事業所　安濃津ろまん'
            elif provider_category == 'kengai':
                provider_name = form_data.get('preventive_support_provider', '') or form_data.get('preventive_support_provider_manual', '')
            else:
                provider_name = form_data.get('preventive_support_provider', '')
            safe_write_cell(coords.get('preventive_provider_name'), provider_name)

            # 郵便番号
            postal_code = form_data.get('houkatsu_postal_code', '')
            safe_write_cell(coords.get('preventive_provider_postal_code'), postal_code)

            # 住所
            address = form_data.get('houkatsu_address', '')
            safe_write_cell(coords.get('preventive_provider_address'), address)

            # 電話番号
            phone = form_data.get('houkatsu_phone', '')
            safe_write_cell(coords.get('preventive_provider_phone'), phone)

            # 担当者氏名
            staff_name = form_data.get('houkatsu_staff_name', '')
            safe_write_cell(coords.get('preventive_provider_staff_name'), staff_name)

            # 事業所番号（10桁を1文字ずつ分割して数値として書き込み）
            office_number = form_data.get('houkatsu_office_number', '')
            # 10桁に満たない場合は左側を空白で埋める
            office_number = office_number.zfill(10) if office_number else ''
            for i, digit in enumerate(office_number):
                cell_key = f'preventive_provider_office_number_{i + 1}'
                # 数字がない場合は空白、ある場合は数値として書き込み
                if digit and digit.isdigit():
                    safe_write_cell(coords.get(cell_key), int(digit), force_number=True)

            # 介護予防支援を受託する居宅介護支援事業者の情報（介護予防版のみ）
            # jutaku_providerが選択されている場合のみ書き込む
            jutaku_provider = form_data.get('jutaku_provider', '')
            if jutaku_provider:
                # 事業所名：その他（手入力）の場合は入力値、それ以外はマスタ名
                if jutaku_provider == 'other':
                    jutaku_name = form_data.get('jutaku_provider_name', '')
                elif request.user.profile.home_care_office:
                    jutaku_name = request.user.profile.home_care_office.name
                else:
                    jutaku_name = '居宅介護支援事業所　安濃津ろまん'  # フォールバック
                safe_write_cell(coords.get('jutaku_provider_name'), jutaku_name)

                # 郵便番号
                jutaku_postal_code = form_data.get('jutaku_postal_code', '')
                safe_write_cell(coords.get('jutaku_provider_postal_code'), jutaku_postal_code)

                # 住所
                jutaku_address = form_data.get('jutaku_address', '')
                safe_write_cell(coords.get('jutaku_provider_address'), jutaku_address)

                # 電話番号
                jutaku_phone = form_data.get('jutaku_phone', '')
                safe_write_cell(coords.get('jutaku_provider_phone'), jutaku_phone)

                # 担当者氏名
                jutaku_staff_name = form_data.get('jutaku_staff_name', '')
                safe_write_cell(coords.get('jutaku_provider_staff_name'), jutaku_staff_name)

                # 事業所番号（10桁を1文字ずつ分割して数値として書き込み）
                jutaku_office_number = form_data.get('jutaku_office_number', '')
                # 10桁に満たない場合は左側を空白で埋める
                jutaku_office_number = jutaku_office_number.zfill(10) if jutaku_office_number else ''
                for i, digit in enumerate(jutaku_office_number):
                    cell_key = f'jutaku_provider_office_number_{i + 1}'
                    # 数字がない場合は空白、ある場合は数値として書き込み
                    if digit and digit.isdigit():
                        safe_write_cell(coords.get(cell_key), int(digit), force_number=True)

        # 居宅介護支援事業者の情報（居宅版のみ）
        if document_type == 'kyotaku_service_plan_request':
            # 事業所名
            safe_write_cell(coords.get('kyotaku_provider_name'), '居宅介護支援事業所　安濃津ろまん')

            # 郵便番号
            postal_code = form_data.get('kyotaku_postal_code', '')
            safe_write_cell(coords.get('kyotaku_provider_postal_code'), postal_code)

            # 住所
            address = form_data.get('kyotaku_address', '')
            safe_write_cell(coords.get('kyotaku_provider_address'), address)

            # 電話番号
            phone = form_data.get('kyotaku_phone', '')
            safe_write_cell(coords.get('kyotaku_provider_phone'), phone)

            # 担当者氏名
            staff_name = form_data.get('kyotaku_staff_name', '')
            safe_write_cell(coords.get('kyotaku_provider_staff_name'), staff_name)

            # 事業所番号（10桁を1文字ずつ分割して数値として書き込み）
            office_number = form_data.get('kyotaku_office_number', '')
            # 10桁に満たない場合は左側を空白で埋める
            office_number = office_number.zfill(10) if office_number else ''
            for i, digit in enumerate(office_number):
                cell_key = f'kyotaku_provider_office_number_{i + 1}'
                # 数字がない場合は空白、ある場合は数値として書き込み
                if digit and digit.isdigit():
                    safe_write_cell(coords.get(cell_key), int(digit), force_number=True)

        # 一時ファイルに保存（拡張子はテンプレートに合わせる）
        file_ext = '.xlsm' if is_macro_enabled else '.xlsx'
        with tempfile.NamedTemporaryFile(delete=False, suffix=file_ext) as tmp:
            workbook.save(tmp.name)
            workbook.close()  # ワークブックを明示的に閉じる

            # ファイルを読み込んでレスポンスに返す
            with open(tmp.name, 'rb') as f:
                if is_macro_enabled:
                    content_type = 'application/vnd.ms-excel.sheet.macroEnabled.12'
                else:
                    content_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                response = HttpResponse(f.read(), content_type=content_type)
                from urllib.parse import quote
                dl_name = _make_dl_filename(client, document_name, file_ext)
                response['Content-Disposition'] = f"attachment; filename=\"download.xlsx\"; filename*=UTF-8''{quote(dl_name, safe='')}"

        # 一時ファイルを削除
        os.unlink(tmp.name)
        return response

    except Exception as e:
        messages.error(request, f'Excel生成中にエラーが発生しました: {str(e)}')
        return redirect('client_detail', pk=client.pk)


@login_required
@user_passes_test(staff_required)
def color_theme_settings(request):
    """カラーテーマ設定画面"""
    from .models import UserProfile

    # ユーザープロファイルを取得または作成
    profile, created = UserProfile.objects.get_or_create(user=request.user)

    if request.method == 'POST':
        theme = request.POST.get('color_theme')
        if theme in dict(UserProfile.THEME_CHOICES):
            profile.color_theme = theme
            profile.save()
            messages.success(request, 'カラーテーマを変更しました。')
            return redirect('color_theme_settings')

    return render(request, 'clients/color_theme_settings.html', {
        'profile': profile,
        'theme_choices': UserProfile.THEME_CHOICES,
    })


@login_required
def document_delete(request, pk):
    """書類削除"""
    if request.method == 'POST':
        document = get_object_or_404(DocumentCreationHistory, pk=pk)
        document.delete()
        # 削除成功メッセージをセッションに保存（警告色で表示）
        messages.warning(request, '書類を削除しました')
        return JsonResponse({'success': True, 'message': '書類を削除しました'})
    return JsonResponse({'success': False, 'message': '不正なリクエストです'}, status=400)


def _make_dl_filename(client, document_name, ext='.xlsx'):
    """統一ダウンロードファイル名: ふりがな苗字_氏名_（書類名）_日付YYYYMMDD"""
    from datetime import date
    furigana = (client.furigana or '').split()[0] if client.furigana else ''
    today = date.today().strftime('%Y%m%d')
    return f"{furigana}_{client.name}_{document_name}_{today}{ext}"


def _generate_ltc_renewal_excel_bytes(client, form_data):
    """更新認定申請書のExcelバイト列を生成して返す"""
    from openpyxl import load_workbook as _load_wb
    from openpyxl.cell import MergedCell as _MergedCell

    filename = getattr(settings, 'LTC_RENEWAL_FILENAME', 'LTC_Certification_Renewal_R8.4-.xlsx')
    template_path = os.path.join(settings.BASE_DIR, 'templates', 'forms', filename)

    workbook = _load_wb(template_path)
    ws = workbook.active

    def w(cell_ref, value, as_text=False):
        if not cell_ref or value is None:
            return
        try:
            cell = ws[cell_ref]
            if isinstance(cell, _MergedCell):
                for mr in ws.merged_cells.ranges:
                    if cell_ref in mr:
                        min_col, min_row, _, _ = mr.bounds
                        target = ws.cell(row=min_row, column=min_col)
                        target.value = str(value) if as_text else value
                        if as_text:
                            target.number_format = '@'
                        return
            else:
                cell.value = str(value) if as_text else value
                if as_text:
                    cell.number_format = '@'
        except Exception:
            pass

    def to_wareki(date_str):
        if not date_str:
            return ''
        try:
            from datetime import date as date_cls, datetime
            d = datetime.strptime(date_str, '%Y-%m-%d').date()
            y, m, day = d.year, d.month, d.day
            if d >= date_cls(2019, 5, 1):
                return f"令和{y-2018:02d}年{m:02d}月{day:02d}日"
            elif d >= date_cls(1989, 1, 8):
                return f"平成{y-1988:02d}年{m:02d}月{day:02d}日"
            elif d >= date_cls(1926, 12, 25):
                return f"昭和{y-1925:02d}年{m:02d}月{day:02d}日"
            else:
                return f"大正{y-1911:02d}年{m:02d}月{day:02d}日"
        except Exception:
            return date_str

    # ① 申請年月日
    w('X7', to_wareki(form_data.get('application_date', '')))

    # ② 申請書提出者（事業所）
    w('G8',  form_data.get('office_furigana', ''))
    w('G9',  form_data.get('office_name', ''))
    w('X9',  form_data.get('relation', '介護支援専門員'))
    w('I11', f"〒{form_data.get('office_postal_code', '')}" if form_data.get('office_postal_code') else '')
    w('G12', form_data.get('office_address', ''))
    w('X12', form_data.get('office_phone', ''))
    w('G14', form_data.get('staff_name', ''))
    w('W14', int(form_data['office_number']) if form_data.get('office_number', '').isdigit() else form_data.get('office_number', ''))

    # ③ 被保険者
    w('G16', int(form_data['insurance_number']) if form_data.get('insurance_number', '').isdigit() else form_data.get('insurance_number', ''))
    w('G17', ''.join(chr(ord(c) + 0x60) if '\u3041' <= c <= '\u3096' else c for c in form_data.get('client_furigana', '')))
    w('G18', form_data.get('client_name', ''))
    w('U17', form_data.get('client_gender', ''))
    w('U18', to_wareki(form_data.get('birth_date', '')))
    w('H19', f"〒{form_data.get('postal_code', '')}" if form_data.get('postal_code') else '')
    w('G20', form_data.get('client_address', ''))
    w('Y21', form_data.get('client_phone', ''))

    # ④ 前回の認定
    w('N22', form_data.get('care_level', ''))
    w('N23', to_wareki(form_data.get('cert_start', '')))
    w('W23', to_wareki(form_data.get('cert_end', '')))

    # ⑤ 転入関係
    if form_data.get('transfer_from_municipality'):
        w('S24', f"[{form_data.get('transfer_from_municipality', '')}]")
    if form_data.get('transfer_applied') == 'yes':
        transfer_date = to_wareki(form_data.get('transfer_applied_date', ''))
        w('W25', f"[はい（申請日：{transfer_date}）]" if transfer_date else '[はい（申請日：　年　月　日）・いいえ]')
    else:
        w('W25', '[はい（申請日：　年　月　日）・いいえ]')

    # ⑥ 医療保険
    w('M26', form_data.get('medical_insurer_name', ''))
    w('AA26', int(form_data['medical_insurer_number']) if form_data.get('medical_insurer_number', '').isdigit() else form_data.get('medical_insurer_number', ''))
    w('O27', form_data.get('medical_insurance_symbol', ''))
    w('Y27', int(form_data['medical_insurance_number']) if form_data.get('medical_insurance_number', '').isdigit() else form_data.get('medical_insurance_number', ''))

    # ⑦ 特定疾病名
    if form_data.get('specific_disease'):
        w('G28', form_data.get('specific_disease', ''))

    # ⑧ 認定調査 場所
    w('G32', form_data.get('survey_location', ''))
    w('S31', f"〒{form_data.get('survey_location_postal', '')}" if form_data.get('survey_location_postal') else '')
    w('O32', form_data.get('survey_location_address', ''))
    w('W33', form_data.get('survey_location_phone', ''))

    # ⑨ 認定調査 連絡先
    w('I34', form_data.get('survey_contact_furigana', ''))
    w('I35', form_data.get('survey_contact_name', ''))
    w('O35', form_data.get('survey_contact_relation', ''))
    w('T35', form_data.get('survey_contact_phone', ''))
    w('G36', form_data.get('survey_contact_preferred_time', ''))
    w('G37', form_data.get('survey_contact_notes', ''))

    # ⑩ 主治医意見書依頼先
    w('G40', form_data.get('doctor_hospital', ''))
    if form_data.get('doctor_outside_city') == 'yes':
        w('S41', f"〒{form_data.get('doctor_postal_code', '')}" if form_data.get('doctor_postal_code') else '')
        w('Q42', form_data.get('doctor_address', ''))
    w('G42', form_data.get('doctor_name', ''))
    w('U43', form_data.get('doctor_phone', ''))
    w('G46', form_data.get('doctor_notes', ''))

    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
        workbook.save(tmp.name)
        workbook.close()
        with open(tmp.name, 'rb') as f:
            content = f.read()
    os.unlink(tmp.name)
    return content


def _add_oval_to_xlsx_bytes(xlsx_bytes, cell_ref, padding=30000):
    """xlsxバイト列のcell_refセルに楕円を重ねて返す（ZIP直接操作）。"""
    import io, zipfile
    from lxml import etree
    from openpyxl.utils.cell import coordinate_from_string, column_index_from_string

    col_str, row_num = coordinate_from_string(cell_ref)
    col_num = column_index_from_string(col_str)

    XDR = 'http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing'
    A_NS = 'http://schemas.openxmlformats.org/drawingml/2006/main'
    RELS_NS = 'http://schemas.openxmlformats.org/package/2006/relationships'
    SS_NS = 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'
    R_NS = 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'
    DRAWING_REL_TYPE = 'http://schemas.openxmlformats.org/officeDocument/2006/relationships/drawing'
    CT_DRAWING = 'application/vnd.openxmlformats-officedocument.drawing+xml'

    anchor_xml = (
        f'<xdr:twoCellAnchor xmlns:xdr="{XDR}" xmlns:a="{A_NS}" editAs="oneCell">'
        f'<xdr:from><xdr:col>{col_num - 1}</xdr:col><xdr:colOff>{padding}</xdr:colOff>'
        f'<xdr:row>{row_num - 1}</xdr:row><xdr:rowOff>{padding}</xdr:rowOff></xdr:from>'
        f'<xdr:to><xdr:col>{col_num}</xdr:col><xdr:colOff>-{padding}</xdr:colOff>'
        f'<xdr:row>{row_num}</xdr:row><xdr:rowOff>-{padding}</xdr:rowOff></xdr:to>'
        '<xdr:sp macro="" textlink="">'
        '<xdr:nvSpPr>'
        '<xdr:cNvPr id="2" name="Oval 1"/>'
        '<xdr:cNvSpPr><a:spLocks noGrp="1"/></xdr:cNvSpPr>'
        '</xdr:nvSpPr>'
        '<xdr:spPr>'
        '<a:prstGeom prst="ellipse"><a:avLst/></a:prstGeom>'
        '<a:noFill/>'
        '<a:ln w="19050"><a:solidFill><a:srgbClr val="000000"/></a:solidFill></a:ln>'
        '</xdr:spPr>'
        '<xdr:txBody><a:bodyPr/><a:lstStyle/><a:p/></xdr:txBody>'
        '</xdr:sp>'
        '<xdr:clientData/>'
        '</xdr:twoCellAnchor>'
    )

    buf_in = io.BytesIO(xlsx_bytes)
    buf_out = io.BytesIO()

    with zipfile.ZipFile(buf_in, 'r') as zin:
        names = set(zin.namelist())
        sheet_path = 'xl/worksheets/sheet1.xml'
        sheet_rels_path = 'xl/worksheets/_rels/sheet1.xml.rels'

        # 既存のdrawingを探す
        existing_draw_path = None
        if sheet_rels_path in names:
            rels_root = etree.fromstring(zin.read(sheet_rels_path))
            for rel in rels_root:
                if rel.get('Type') == DRAWING_REL_TYPE:
                    target = rel.get('Target', '')
                    existing_draw_path = 'xl/' + target.lstrip('../')
                    break

        if existing_draw_path:
            new_draw_path = existing_draw_path
            new_rel_id = None
            draw_num = None
        else:
            draw_num = sum(1 for n in names if n.startswith('xl/drawings/drawing')) + 1
            new_draw_path = f'xl/drawings/drawing{draw_num}.xml'
            new_rel_id = f'rIdOval{draw_num}'

        with zipfile.ZipFile(buf_out, 'w', zipfile.ZIP_DEFLATED) as zout:
            for info in zin.infolist():
                data = zin.read(info.filename)

                if info.filename == '[Content_Types].xml' and new_rel_id:
                    root = etree.fromstring(data)
                    if CT_DRAWING not in {el.get('ContentType') for el in root}:
                        etree.SubElement(root, 'Override', {
                            'PartName': f'/{new_draw_path}',
                            'ContentType': CT_DRAWING,
                        })
                    data = etree.tostring(root, xml_declaration=True, encoding='UTF-8', standalone=True)

                elif info.filename == sheet_path and new_rel_id:
                    root = etree.fromstring(data)
                    if root.find(f'{{{SS_NS}}}drawing') is None:
                        el = etree.SubElement(root, f'{{{SS_NS}}}drawing')
                        el.set(f'{{{R_NS}}}id', new_rel_id)
                    data = etree.tostring(root, xml_declaration=True, encoding='UTF-8', standalone=True)

                elif info.filename == sheet_rels_path and new_rel_id:
                    root = etree.fromstring(data)
                    etree.SubElement(root, 'Relationship', {
                        'Id': new_rel_id,
                        'Type': DRAWING_REL_TYPE,
                        'Target': f'../drawings/drawing{draw_num}.xml',
                    })
                    data = etree.tostring(root, xml_declaration=True, encoding='UTF-8', standalone=True)

                elif info.filename == new_draw_path and not new_rel_id:
                    # 既存drawingに楕円を追記
                    root = etree.fromstring(data)
                    anchor_el = etree.fromstring(anchor_xml)
                    existing_ids = [int(e.get('id', 0)) for e in root.iter(f'{{{XDR}}}cNvPr')]
                    max_id = max(existing_ids) if existing_ids else 1
                    for e in anchor_el.iter(f'{{{XDR}}}cNvPr'):
                        e.set('id', str(max_id + 1))
                    root.append(anchor_el)
                    data = etree.tostring(root, xml_declaration=True, encoding='UTF-8', standalone=True)

                zout.writestr(info, data)

            if new_rel_id:
                if sheet_rels_path not in names:
                    zout.writestr(sheet_rels_path, (
                        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
                        f'<Relationships xmlns="{RELS_NS}">'
                        f'<Relationship Id="{new_rel_id}" Type="{DRAWING_REL_TYPE}"'
                        f' Target="../drawings/drawing{draw_num}.xml"/>'
                        '</Relationships>'
                    ).encode('utf-8'))
                zout.writestr(new_draw_path, (
                    '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
                    f'<xdr:wsDr xmlns:xdr="{XDR}" xmlns:a="{A_NS}">'
                    + anchor_xml +
                    '</xdr:wsDr>'
                ).encode('utf-8'))
                zout.writestr(f'xl/drawings/_rels/drawing{draw_num}.xml.rels', (
                    '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
                    f'<Relationships xmlns="{RELS_NS}"/>'
                ).encode('utf-8'))

    return buf_out.getvalue()


def _generate_ltc_change_excel_bytes(client, form_data):
    """区分変更申請書のExcelバイト列を生成して返す"""
    from openpyxl import load_workbook as _load_wb
    from openpyxl.cell import MergedCell as _MergedCell

    filename = getattr(settings, 'LTC_CHANGE_FILENAME', 'LTC_Certification_Change_R8.4-.xlsx')
    template_path = os.path.join(settings.BASE_DIR, 'templates', 'forms', filename)

    workbook = _load_wb(template_path)
    ws = workbook.active

    def w(cell_ref, value, as_text=False):
        if not cell_ref or value is None:
            return
        try:
            cell = ws[cell_ref]
            if isinstance(cell, _MergedCell):
                for mr in ws.merged_cells.ranges:
                    if cell_ref in mr:
                        min_col, min_row, _, _ = mr.bounds
                        target = ws.cell(row=min_row, column=min_col)
                        target.value = str(value) if as_text else value
                        if as_text:
                            target.number_format = '@'
                        return
            else:
                cell.value = str(value) if as_text else value
                if as_text:
                    cell.number_format = '@'
        except Exception:
            pass

    def to_wareki(date_str):
        if not date_str:
            return ''
        try:
            from datetime import date as date_cls, datetime
            d = datetime.strptime(date_str, '%Y-%m-%d').date()
            y, m, day = d.year, d.month, d.day
            if d >= date_cls(2019, 5, 1):
                return f"令和{y-2018:02d}年{m:02d}月{day:02d}日"
            elif d >= date_cls(1989, 1, 8):
                return f"平成{y-1988:02d}年{m:02d}月{day:02d}日"
            elif d >= date_cls(1926, 12, 25):
                return f"昭和{y-1925:02d}年{m:02d}月{day:02d}日"
            else:
                return f"大正{y-1911:02d}年{m:02d}月{day:02d}日"
        except Exception:
            return date_str

    # ① 申請年月日
    w('X7', to_wareki(form_data.get('application_date', '')))

    # ② 申請書提出者（事業所）
    w('G8',  form_data.get('office_furigana', ''))
    w('G9',  form_data.get('office_name', ''))
    w('X9',  form_data.get('relation', '介護支援専門員'))
    w('I11', form_data.get('office_postal_code', ''))
    w('G12', form_data.get('office_address', ''))
    w('X12', form_data.get('office_phone', ''))
    w('G14', form_data.get('staff_name', ''))
    w('W14', int(form_data['office_number']) if form_data.get('office_number', '').isdigit() else form_data.get('office_number', ''))

    # ③ 被保険者
    w('G16', int(form_data['insurance_number']) if form_data.get('insurance_number', '').isdigit() else form_data.get('insurance_number', ''))
    w('G17', ''.join(chr(ord(c) + 0x60) if '\u3041' <= c <= '\u3096' else c for c in form_data.get('client_furigana', '')))
    w('G18', form_data.get('client_name', ''))
    w('U17', form_data.get('client_gender', ''))
    w('U18', to_wareki(form_data.get('birth_date', '')))
    w('H19', form_data.get('postal_code', ''))
    w('G20', form_data.get('client_address', ''))
    w('Y21', form_data.get('client_phone', ''))

    # ④ 前回の認定
    w('J22', form_data.get('care_level', ''))
    w('J23', to_wareki(form_data.get('cert_start', '')))
    w('T23', to_wareki(form_data.get('cert_end', '')))

    # ⑤ 申請の理由
    w('G24', form_data.get('change_reason', ''))

    # ⑥ 医療保険
    w('M26', form_data.get('medical_insurer_name', ''))
    w('AA26', int(form_data['medical_insurer_number']) if form_data.get('medical_insurer_number', '').isdigit() else form_data.get('medical_insurer_number', ''))
    w('O27', form_data.get('medical_insurance_symbol', ''))
    w('Y27', int(form_data['medical_insurance_number']) if form_data.get('medical_insurance_number', '').isdigit() else form_data.get('medical_insurance_number', ''))

    # ⑦ 特定疾病名
    if form_data.get('specific_disease'):
        w('G28', form_data.get('specific_disease', ''))

    # ⑧ 認定調査 場所
    w('G32', form_data.get('survey_location', ''))
    w('S31', form_data.get('survey_location_postal', ''))
    w('O32', form_data.get('survey_location_address', ''))
    w('W33', form_data.get('survey_location_phone', ''))

    # ⑨ 認定調査 連絡先
    w('I34', form_data.get('survey_contact_furigana', ''))
    w('I35', form_data.get('survey_contact_name', ''))
    w('O35', form_data.get('survey_contact_relation', ''))
    w('T35', form_data.get('survey_contact_phone', ''))
    w('G36', form_data.get('survey_contact_preferred_time', ''))
    w('G38', form_data.get('survey_contact_notes', ''))

    # ⑩ 主治医意見書依頼先
    w('G40', form_data.get('doctor_hospital', ''))
    if form_data.get('doctor_outside_city') == 'yes':
        w('S41', form_data.get('doctor_postal_code', ''))
        w('Q42', form_data.get('doctor_address', ''))
    w('G42', form_data.get('doctor_name', ''))
    w('U43', form_data.get('doctor_phone', ''))
    w('G46', form_data.get('doctor_notes', ''))

    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
        workbook.save(tmp.name)
        workbook.close()
        with open(tmp.name, 'rb') as f:
            content = f.read()
    os.unlink(tmp.name)
    return content


@login_required
def document_history_excel(request, pk):
    """書類履歴からExcelファイルをダウンロード"""
    history = get_object_or_404(DocumentCreationHistory, pk=pk)
    client = history.client

    if history.document_type in ('ltc_renewal', 'ltc_change'):
        try:
            if history.document_type == 'ltc_change':
                content = _generate_ltc_change_excel_bytes(client, history.form_data)
                dl_label = '区分変更申請書'
            else:
                content = _generate_ltc_renewal_excel_bytes(client, history.form_data)
                dl_label = '更新認定申請書'
        except Exception as e:
            messages.error(request, f'Excel生成中にエラーが発生しました: {str(e)}')
            return redirect('client_detail', pk=client.pk)
        from urllib.parse import quote
        dl_name = _make_dl_filename(client, dl_label)
        response = HttpResponse(
            content,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f"attachment; filename=\"download.xlsx\"; filename*=UTF-8''{quote(dl_name, safe='')}"
        return response

    return generate_document_excel(
        request,
        client,
        history.document_type,
        history.document_name,
        history.form_data
    )


@login_required
@user_passes_test(staff_required)
@login_required
@user_passes_test(staff_required)
def document_file_management(request):
    """書類ファイル名管理画面"""
    documents = [
        {
            'section': '介護保険更新及び変更申請',
            'items': [
                {'name': '更新認定申請書',           'filename': getattr(settings, 'LTC_RENEWAL_FILENAME', 'LTC_Certification_Renewal_R8.4-.xlsx')},
                {'name': '区分変更申請書',           'filename': getattr(settings, 'LTC_CHANGE_FILENAME',   'LTC_Certification_Change_R8.4-.xlsx')},
                {'name': '認定申請主治医変更届出書', 'filename': '介護保険要介護（更新）・要支援（更新）申請書主治医変更届出書.xls'},
                {'name': '認定申請取下書',           'filename': '介護保険要介護認定・要支援認定取下書.xls'},
            ],
        },
        {
            'section': '送付先変更・再交付関係',
            'items': [
                {'name': '介護保険関係・再交付申請書',              'filename': 'LTC_Reissue_Application.xlsx'},
                {'name': '介護保険被保険者証の送付先変更届',        'filename': '（未設定）'},
                {'name': '介護保険負担限度額・割合証送付先変更届',  'filename': '（未設定）'},
            ],
        },
        {
            'section': '居宅の届出及び資料請求',
            'items': [
                {'name': '居宅サービス計画作成依頼（変更）届出書',                                       'filename': 'kyotaku_service_plan_request.xlsx'},
                {'name': '介護予防サービス計画作成・介護予防ケアマネジメント依頼（変更）届出書', 'filename': 'kyotaku_preventive_service_plan_request.xlsx'},
                {'name': '介護サービス計画作成に係る資料提供申請書',                              'filename': 'CarePlan_Info_Request.xlsx'},
            ],
        },
        {
            'section': 'アセスメント・記録',
            'items': [
                {'name': 'アセスメントシート',  'filename': 'assessment_sheet.xlsx / assessment_sheet_genpon.xlsm'},
                {'name': 'フェイスシート',      'filename': 'Face Sheet.xlsx'},
                {'name': '医療連携シート',      'filename': 'medical_care_coordination_sheet.xlsx'},
                {'name': '入院時情報提供書',    'filename': 'hospital_admission_info_form.xlsx'},
            ],
        },
        {
            'section': '計画作成及び担当者会議',
            'items': [
                {'name': 'ケアサービス担当者会議通知書',  'filename': 'care_service_meeting_notice.xlsx'},
                {'name': '居宅サービス事業所選択確認書',  'filename': 'care_provider_selection_confirm.xlsx'},
            ],
        },
    ]

    return render(request, 'clients/document_file_management.html', {'documents': documents})


def color_reference(request):
    """カラー参照画面"""
    # ベースの色データ
    base_colors = [
        {
            'category': 'ヘッダー',
            'items': [
                {'name': 'ヘッダー背景', 'bg_color': '#4a5568', 'border_color': '#4a5568', 'text_color': '#ffffff'},
            ]
        },
        {
            'category': 'フッター',
            'items': [
                {'name': 'フッター背景', 'bg_color': '#f8f9fa', 'border_color': '#dee2e6', 'text_color': '#2d3748'},
            ]
        },
        {
            'category': '表のヘッダー',
            'items': [
                {'name': 'ヘッダー行', 'bg_color': '#f8f9fa', 'border_color': '#dee2e6', 'text_color': '#2d3748'},
            ]
        },
        {
            'category': 'サブタブ',
            'items': [
                {'name': 'アクティブ時', 'bg_color': '#f5f5f5', 'border_color': '#999999', 'text_color': '#333333'},
                {'name': '非アクティブ時', 'bg_color': '#ffffff', 'border_color': '#dee2e6', 'text_color': '#6c757d'},
            ]
        },
        {
            'category': '操作ボタン',
            'items': [
                {'name': '未選択', 'bg_color': '#ffffff', 'border_color': '#6c757d', 'text_color': '#6c757d'},
                {'name': '戻る', 'bg_color': '#f5f5f5', 'border_color': '#999999', 'text_color': '#333333'},
                {'name': 'Excel', 'bg_color': '#d1f2d1', 'border_color': '#198754', 'text_color': '#198754'},
                {'name': '編集', 'bg_color': '#ffe5cc', 'border_color': '#fd7e14', 'text_color': '#fd7e14'},
                {'name': '削除', 'bg_color': '#f8d7da', 'border_color': '#dc3545', 'text_color': '#dc3545'},
                {'name': 'プレビュー', 'bg_color': '#e3f2fd', 'border_color': '#0d6efd', 'text_color': '#0d6efd'},
                {'name': '保存', 'bg_color': '#f8d7da', 'border_color': '#dc3545', 'text_color': '#dc3545'},
            ]
        },
    ]

    # 利用者一覧の色データ
    client_list_colors = [
        {
            'category': '利用者一覧 - 新規登録ボタン',
            'items': [
                {'name': '通常時', 'bg_color': '#ffffff', 'border_color': '#6c757d', 'text_color': '#6c757d'},
                {'name': 'ホバー時', 'bg_color': '#e3f2fd', 'border_color': '#0d6efd', 'text_color': '#0d6efd'},
            ]
        },
        {
            'category': '利用者一覧 - 件数表示',
            'items': [
                {'name': '件数バッジ', 'bg_color': '#f5f5f5', 'border_color': '#999999', 'text_color': '#000000'},
            ]
        },
        {
            'category': '利用者一覧 - テーブル行',
            'items': [
                {'name': '一覧行にカーソルを合わせた時', 'bg_color': '#f8f9ff', 'border_color': '#dee2e6', 'text_color': '#000000'},
            ]
        },
        {
            'category': '利用者一覧 - 介護度バッジ',
            'items': [
                {'name': '要支援1', 'bg_color': '#f5f5f5', 'border_color': '#999999', 'text_color': '#000000'},
                {'name': '要支援2', 'bg_color': '#f5f0e8', 'border_color': '#999999', 'text_color': '#000000'},
                {'name': '要介護1', 'bg_color': '#e3f2fd', 'border_color': '#999999', 'text_color': '#000000'},
                {'name': '要介護2', 'bg_color': '#e8f5e8', 'border_color': '#999999', 'text_color': '#000000'},
                {'name': '要介護3', 'bg_color': '#fff3e0', 'border_color': '#999999', 'text_color': '#000000'},
                {'name': '要介護4', 'bg_color': '#f3e5f5', 'border_color': '#999999', 'text_color': '#000000'},
                {'name': '要介護5', 'bg_color': '#ffebee', 'border_color': '#999999', 'text_color': '#000000'},
            ]
        },
        {
            'category': '利用者一覧 - カウントダウンバッジ',
            'items': [
                {'name': '期限切れ（×）', 'bg_color': '#ffebee', 'border_color': '#999999', 'text_color': '#000000'},
                {'name': '30日以内', 'bg_color': '#fff3e0', 'border_color': '#999999', 'text_color': '#000000'},
                {'name': '31日以上', 'bg_color': '#fff9c4', 'border_color': '#999999', 'text_color': '#000000'},
            ]
        },
        {
            'category': '利用者一覧 - 操作ボタン',
            'items': [
                {'name': '通常時', 'bg_color': '#ffffff', 'border_color': '#6c757d', 'text_color': '#6c757d'},
                {'name': 'ホバー・アクティブ時', 'bg_color': '#f5f5f5', 'border_color': '#999999', 'text_color': '#000000'},
            ]
        },
        {
            'category': '利用者一覧 - 操作アイコン',
            'items': [
                {'name': '詳細表示', 'bg_color': 'transparent', 'border_color': 'なし', 'text_color': 'rgba(153, 153, 153, 0.5)'},
                {'name': '情報編集', 'bg_color': 'transparent', 'border_color': 'なし', 'text_color': 'rgba(13, 110, 253, 0.5)'},
                {'name': '限度額試算', 'bg_color': 'transparent', 'border_color': 'なし', 'text_color': 'rgba(25, 135, 84, 0.5)'},
                {'name': 'アセスメント', 'bg_color': 'transparent', 'border_color': 'なし', 'text_color': 'rgba(253, 126, 20, 0.5)'},
                {'name': '書類作成', 'bg_color': 'transparent', 'border_color': 'なし', 'text_color': 'rgba(111, 66, 193, 0.5)'},
                {'name': '保険・受給者証管理', 'bg_color': 'transparent', 'border_color': 'なし', 'text_color': 'rgba(214, 51, 132, 0.5)'},
                {'name': '履歴・記録', 'bg_color': 'transparent', 'border_color': 'なし', 'text_color': 'rgba(255, 193, 7, 0.5)'},
                {'name': '印刷・出力', 'bg_color': 'transparent', 'border_color': 'なし', 'text_color': 'rgba(210, 180, 140, 0.5)'},
            ]
        },
        {
            'category': '利用者一覧 - 操作メニューボタン',
            'items': [
                {'name': '詳細表示', 'bg_color': '#f5f5f5', 'border_color': '#999999', 'text_color': '#000000', 'muted_bg_color': '#e8e8e8', 'muted_border_color': '#999999', 'muted_text_color': '#666666'},
                {'name': '情報編集', 'bg_color': '#e3f2fd', 'border_color': '#0d6efd', 'text_color': '#0d6efd', 'muted_bg_color': '#c8d8e8', 'muted_border_color': '#7a9cc6', 'muted_text_color': '#4a6fa5'},
                {'name': '限度額試算', 'bg_color': '#e8f5e8', 'border_color': '#198754', 'text_color': '#198754', 'muted_bg_color': '#c8e0c8', 'muted_border_color': '#7aaa7a', 'muted_text_color': '#4a854a'},
                {'name': 'アセスメント', 'bg_color': '#fff3e0', 'border_color': '#fd7e14', 'text_color': '#fd7e14', 'muted_bg_color': '#e8d0b8', 'muted_border_color': '#c18a5a', 'muted_text_color': '#9a6a3a'},
                {'name': '書類作成', 'bg_color': '#f3e5f5', 'border_color': '#6f42c1', 'text_color': '#6f42c1', 'muted_bg_color': '#d8c8e0', 'muted_border_color': '#9a7ab3', 'muted_text_color': '#704a8f'},
                {'name': '保険・受給者証管理', 'bg_color': '#ffebee', 'border_color': '#d63384', 'text_color': '#d63384', 'muted_bg_color': '#e8c8d8', 'muted_border_color': '#c17a9a', 'muted_text_color': '#9a4a70'},
                {'name': '履歴・記録', 'bg_color': '#fff9c4', 'border_color': '#ffc107', 'text_color': '#856404', 'muted_bg_color': '#e8e0c0', 'muted_border_color': '#c1b370', 'muted_text_color': '#9a8a4a'},
                {'name': '印刷・出力', 'bg_color': '#f5f0e8', 'border_color': '#d2b48c', 'text_color': '#8b7355', 'muted_bg_color': '#e0d4c8', 'muted_border_color': '#b39a7a', 'muted_text_color': '#8a6f50'},
            ]
        },
    ]

    # アセスメント作成の色データ
    assessment_colors = [
        {
            'category': 'アセスメント - 選択ボタン',
            'items': [
                {'name': '未選択', 'bg_color': '#ffffff', 'border_color': '#6c757d', 'text_color': '#6c757d'},
                {'name': 'カーソル', 'bg_color': '#e9ecef', 'border_color': '#6c757d', 'text_color': '#495057'},
                {'name': 'ポジティブ', 'bg_color': '#e3f2fd', 'border_color': '#0d6efd', 'text_color': '#0d6efd'},
                {'name': 'ネガティブ', 'bg_color': '#ffc8dd', 'border_color': '#d63384', 'text_color': '#d63384'},
                {'name': 'ニュートラル', 'bg_color': '#d1f2d1', 'border_color': '#198754', 'text_color': '#198754'},
            ]
        },
        {
            'category': 'アセスメント - メインタブ',
            'items': [
                {'name': '基本情報タブ', 'bg_color': 'rgba(153, 153, 153, 0.5)', 'border_color': 'rgba(153, 153, 153, 0.5)', 'text_color': '#000000'},
                {'name': '健康状態タブ', 'bg_color': 'rgba(13, 110, 253, 0.5)', 'border_color': 'rgba(13, 110, 253, 0.5)', 'text_color': '#000000'},
                {'name': '身体機能タブ', 'bg_color': 'rgba(25, 135, 84, 0.5)', 'border_color': 'rgba(25, 135, 84, 0.5)', 'text_color': '#000000'},
                {'name': 'ADLタブ', 'bg_color': 'rgba(253, 126, 20, 0.5)', 'border_color': 'rgba(253, 126, 20, 0.5)', 'text_color': '#000000'},
                {'name': 'IADLタブ', 'bg_color': 'rgba(111, 66, 193, 0.5)', 'border_color': 'rgba(111, 66, 193, 0.5)', 'text_color': '#000000'},
                {'name': '認知機能タブ', 'bg_color': 'rgba(214, 51, 132, 0.5)', 'border_color': 'rgba(214, 51, 132, 0.5)', 'text_color': '#000000'},
                {'name': 'サービスタブ', 'bg_color': 'rgba(255, 193, 7, 0.5)', 'border_color': 'rgba(255, 193, 7, 0.5)', 'text_color': '#000000'},
                {'name': '住宅環境タブ', 'bg_color': 'rgba(210, 180, 140, 0.5)', 'border_color': 'rgba(210, 180, 140, 0.5)', 'text_color': '#000000'},
            ]
        },
        {
            'category': 'アセスメント - チェックボックス',
            'items': [
                {'name': 'チェックマーク（記号）', 'bg_color': '#6eb3ff', 'border_color': '#6eb3ff', 'text_color': '#ffffff'},
            ]
        },
        {
            'category': 'アセスメント - セクション背景色',
            'items': [
                {'name': '各項目の背景', 'bg_color': '#f8f9fa', 'border_color': '#e9ecef', 'text_color': '#495057'},
            ]
        },
    ]

    # 利用者詳細の色データ
    client_detail_colors = [
        {
            'category': '利用者詳細 - メインタブ',
            'items': [
                {'name': '基本情報タブ', 'bg_color': 'rgba(153, 153, 153, 0.5)', 'border_color': 'rgba(153, 153, 153, 0.5)', 'text_color': '#000000'},
                {'name': 'アセスメント履歴タブ', 'bg_color': 'rgba(13, 110, 253, 0.5)', 'border_color': 'rgba(13, 110, 253, 0.5)', 'text_color': '#000000'},
                {'name': '書類作成履歴タブ', 'bg_color': 'rgba(25, 135, 84, 0.5)', 'border_color': 'rgba(25, 135, 84, 0.5)', 'text_color': '#000000'},
            ]
        },
        {
            'category': '利用者詳細 - 件数バッジ',
            'items': [
                {'name': '件数表示', 'bg_color': '#f5f5f5', 'border_color': '#999999', 'text_color': '#333333'},
            ]
        },
    ]

    # カラーパターン（番号で統一管理）
    color_patterns = [
        # パターン0: 通常時のボタン
        {'number': 0, 'text_color': '#6c757d', 'border_color': '#6c757d', 'bg_color': '#ffffff', 'muted_text_color': '#6c757d', 'muted_border_color': '#6c757d', 'muted_bg_color': '#ffffff'},

        # パターン1～8: グレー枠線固定 + 背景色8色
        {'number': 1, 'text_color': '#000000', 'border_color': '#999999', 'bg_color': '#f5f5f5', 'muted_text_color': '#000000', 'muted_border_color': '#999999', 'muted_bg_color': '#e8e8e8'},
        {'number': 2, 'text_color': '#000000', 'border_color': '#999999', 'bg_color': '#e3f2fd', 'muted_text_color': '#000000', 'muted_border_color': '#7a9cc6', 'muted_bg_color': '#c8d8e8'},
        {'number': 3, 'text_color': '#000000', 'border_color': '#999999', 'bg_color': '#e8f5e8', 'muted_text_color': '#000000', 'muted_border_color': '#7aaa7a', 'muted_bg_color': '#c8e0c8'},
        {'number': 4, 'text_color': '#000000', 'border_color': '#999999', 'bg_color': '#fff3e0', 'muted_text_color': '#000000', 'muted_border_color': '#c18a5a', 'muted_bg_color': '#e8d0b8'},
        {'number': 5, 'text_color': '#000000', 'border_color': '#999999', 'bg_color': '#f3e5f5', 'muted_text_color': '#000000', 'muted_border_color': '#9a7ab3', 'muted_bg_color': '#d8c8e0'},
        {'number': 6, 'text_color': '#000000', 'border_color': '#999999', 'bg_color': '#ffebee', 'muted_text_color': '#000000', 'muted_border_color': '#c17a9a', 'muted_bg_color': '#e8c8d8'},
        {'number': 7, 'text_color': '#000000', 'border_color': '#999999', 'bg_color': '#fff9c4', 'muted_text_color': '#000000', 'muted_border_color': '#c1b370', 'muted_bg_color': '#e8e0c0'},
        {'number': 8, 'text_color': '#000000', 'border_color': '#999999', 'bg_color': '#f5f0e8', 'muted_text_color': '#000000', 'muted_border_color': '#b39a7a', 'muted_bg_color': '#e0d4c8'},

        # パターン9～16: 色付き枠線 + 背景色セット
        {'number': 9, 'text_color': '#6c757d', 'border_color': '#6c757d', 'bg_color': '#f5f5f5', 'muted_text_color': '#666666', 'muted_border_color': '#999999', 'muted_bg_color': '#e8e8e8'},
        {'number': 10, 'text_color': '#0d6efd', 'border_color': '#0d6efd', 'bg_color': '#e3f2fd', 'muted_text_color': '#4a6fa5', 'muted_border_color': '#7a9cc6', 'muted_bg_color': '#c8d8e8'},
        {'number': 11, 'text_color': '#198754', 'border_color': '#198754', 'bg_color': '#e8f5e8', 'muted_text_color': '#4a854a', 'muted_border_color': '#7aaa7a', 'muted_bg_color': '#c8e0c8'},
        {'number': 12, 'text_color': '#fd7e14', 'border_color': '#fd7e14', 'bg_color': '#fff3e0', 'muted_text_color': '#9a6a3a', 'muted_border_color': '#c18a5a', 'muted_bg_color': '#e8d0b8'},
        {'number': 13, 'text_color': '#6f42c1', 'border_color': '#6f42c1', 'bg_color': '#f3e5f5', 'muted_text_color': '#704a8f', 'muted_border_color': '#9a7ab3', 'muted_bg_color': '#d8c8e0'},
        {'number': 14, 'text_color': '#d63384', 'border_color': '#d63384', 'bg_color': '#ffebee', 'muted_text_color': '#9a4a70', 'muted_border_color': '#c17a9a', 'muted_bg_color': '#e8c8d8'},
        {'number': 15, 'text_color': '#ffc107', 'border_color': '#ffc107', 'bg_color': '#fff9c4', 'muted_text_color': '#9a8a4a', 'muted_border_color': '#c1b370', 'muted_bg_color': '#e8e0c0'},
        {'number': 16, 'text_color': '#d2b48c', 'border_color': '#d2b48c', 'bg_color': '#f5f0e8', 'muted_text_color': '#8a6f50', 'muted_border_color': '#b39a7a', 'muted_bg_color': '#e0d4c8'},

        # パターン17～24: 背景色のみ（枠線なし）
        {'number': 17, 'text_color': '#000000', 'border_color': 'none', 'bg_color': '#f5f5f5', 'muted_text_color': '#000000', 'muted_border_color': 'none', 'muted_bg_color': '#e8e8e8'},
        {'number': 18, 'text_color': '#000000', 'border_color': 'none', 'bg_color': '#e3f2fd', 'muted_text_color': '#000000', 'muted_border_color': 'none', 'muted_bg_color': '#c8d8e8'},
        {'number': 19, 'text_color': '#000000', 'border_color': 'none', 'bg_color': '#e8f5e8', 'muted_text_color': '#000000', 'muted_border_color': 'none', 'muted_bg_color': '#c8e0c8'},
        {'number': 20, 'text_color': '#000000', 'border_color': 'none', 'bg_color': '#fff3e0', 'muted_text_color': '#000000', 'muted_border_color': 'none', 'muted_bg_color': '#e8d0b8'},
        {'number': 21, 'text_color': '#000000', 'border_color': 'none', 'bg_color': '#f3e5f5', 'muted_text_color': '#000000', 'muted_border_color': 'none', 'muted_bg_color': '#d8c8e0'},
        {'number': 22, 'text_color': '#000000', 'border_color': 'none', 'bg_color': '#ffebee', 'muted_text_color': '#000000', 'muted_border_color': 'none', 'muted_bg_color': '#e8c8d8'},
        {'number': 23, 'text_color': '#000000', 'border_color': 'none', 'bg_color': '#fff9c4', 'muted_text_color': '#000000', 'muted_border_color': 'none', 'muted_bg_color': '#e8e0c0'},
        {'number': 24, 'text_color': '#000000', 'border_color': 'none', 'bg_color': '#f5f0e8', 'muted_text_color': '#000000', 'muted_border_color': 'none', 'muted_bg_color': '#e0d4c8'},

        # パターン25: テーブルヘッダー背景色
        {'number': 25, 'text_color': '#000000', 'border_color': 'none', 'bg_color': '#f8f9fa', 'muted_text_color': '#000000', 'muted_border_color': 'none', 'muted_bg_color': '#e8e8e8'},
    ]

    # 暖色パターン（赤、オレンジ、ピンク、黄色系）
    warm_color_patterns = [
        # パターン0: 通常時のボタン
        {'number': 0, 'text_color': '#6c757d', 'border_color': '#6c757d', 'bg_color': '#ffffff'},

        # パターン1～8: グレー枠線固定 + 暖色背景8色
        {'number': 1, 'text_color': '#000000', 'border_color': '#999999', 'bg_color': '#ffebee'},  # 淡いレッド
        {'number': 2, 'text_color': '#000000', 'border_color': '#999999', 'bg_color': '#fff3e0'},  # 淡いオレンジ
        {'number': 3, 'text_color': '#000000', 'border_color': '#999999', 'bg_color': '#fce4ec'},  # 淡いピンク
        {'number': 4, 'text_color': '#000000', 'border_color': '#999999', 'bg_color': '#fffde7'},  # 淡いイエロー
        {'number': 5, 'text_color': '#000000', 'border_color': '#999999', 'bg_color': '#ffcdd2'},  # レッド
        {'number': 6, 'text_color': '#000000', 'border_color': '#999999', 'bg_color': '#ffe0b2'},  # オレンジ
        {'number': 7, 'text_color': '#000000', 'border_color': '#999999', 'bg_color': '#f8bbd0'},  # ピンク
        {'number': 8, 'text_color': '#000000', 'border_color': '#999999', 'bg_color': '#fff9c4'},  # イエロー

        # パターン9～16: 暖色枠線 + 背景色セット
        {'number': 9, 'text_color': '#c62828', 'border_color': '#c62828', 'bg_color': '#ffebee'},  # レッド
        {'number': 10, 'text_color': '#e65100', 'border_color': '#e65100', 'bg_color': '#fff3e0'},  # オレンジ
        {'number': 11, 'text_color': '#c2185b', 'border_color': '#c2185b', 'bg_color': '#fce4ec'},  # ピンク
        {'number': 12, 'text_color': '#f57c00', 'border_color': '#f57c00', 'bg_color': '#fffde7'},  # イエロー
        {'number': 13, 'text_color': '#d32f2f', 'border_color': '#d32f2f', 'bg_color': '#ffcdd2'},  # レッド濃
        {'number': 14, 'text_color': '#ef6c00', 'border_color': '#ef6c00', 'bg_color': '#ffe0b2'},  # オレンジ濃
        {'number': 15, 'text_color': '#ad1457', 'border_color': '#ad1457', 'bg_color': '#f8bbd0'},  # ピンク濃
        {'number': 16, 'text_color': '#f9a825', 'border_color': '#f9a825', 'bg_color': '#fff9c4'},  # イエロー濃

        # パターン17～24: 背景色のみ（枠線なし）
        {'number': 17, 'text_color': '#000000', 'border_color': 'none', 'bg_color': '#ffebee'},
        {'number': 18, 'text_color': '#000000', 'border_color': 'none', 'bg_color': '#fff3e0'},
        {'number': 19, 'text_color': '#000000', 'border_color': 'none', 'bg_color': '#fce4ec'},
        {'number': 20, 'text_color': '#000000', 'border_color': 'none', 'bg_color': '#fffde7'},
        {'number': 21, 'text_color': '#000000', 'border_color': 'none', 'bg_color': '#ffcdd2'},
        {'number': 22, 'text_color': '#000000', 'border_color': 'none', 'bg_color': '#ffe0b2'},
        {'number': 23, 'text_color': '#000000', 'border_color': 'none', 'bg_color': '#f8bbd0'},
        {'number': 24, 'text_color': '#000000', 'border_color': 'none', 'bg_color': '#fff9c4'},

        # パターン25: テーブルヘッダー背景色
        {'number': 25, 'text_color': '#000000', 'border_color': 'none', 'bg_color': '#f8f9fa'},
    ]

    # 寒色パターン（青、緑、紫、ティール系）
    cool_color_patterns = [
        # パターン0: 通常時のボタン
        {'number': 0, 'text_color': '#6c757d', 'border_color': '#6c757d', 'bg_color': '#ffffff'},

        # パターン1～8: グレー枠線固定 + 寒色背景8色
        {'number': 1, 'text_color': '#000000', 'border_color': '#999999', 'bg_color': '#e3f2fd'},  # 淡いブルー
        {'number': 2, 'text_color': '#000000', 'border_color': '#999999', 'bg_color': '#e8f5e9'},  # 淡いグリーン
        {'number': 3, 'text_color': '#000000', 'border_color': '#999999', 'bg_color': '#f3e5f5'},  # 淡いパープル
        {'number': 4, 'text_color': '#000000', 'border_color': '#999999', 'bg_color': '#e0f2f1'},  # 淡いティール
        {'number': 5, 'text_color': '#000000', 'border_color': '#999999', 'bg_color': '#bbdefb'},  # ブルー
        {'number': 6, 'text_color': '#000000', 'border_color': '#999999', 'bg_color': '#c8e6c9'},  # グリーン
        {'number': 7, 'text_color': '#000000', 'border_color': '#999999', 'bg_color': '#e1bee7'},  # パープル
        {'number': 8, 'text_color': '#000000', 'border_color': '#999999', 'bg_color': '#b2dfdb'},  # ティール

        # パターン9～16: 寒色枠線 + 背景色セット
        {'number': 9, 'text_color': '#1565c0', 'border_color': '#1565c0', 'bg_color': '#e3f2fd'},  # ブルー
        {'number': 10, 'text_color': '#2e7d32', 'border_color': '#2e7d32', 'bg_color': '#e8f5e9'},  # グリーン
        {'number': 11, 'text_color': '#6a1b9a', 'border_color': '#6a1b9a', 'bg_color': '#f3e5f5'},  # パープル
        {'number': 12, 'text_color': '#00695c', 'border_color': '#00695c', 'bg_color': '#e0f2f1'},  # ティール
        {'number': 13, 'text_color': '#1976d2', 'border_color': '#1976d2', 'bg_color': '#bbdefb'},  # ブルー濃
        {'number': 14, 'text_color': '#388e3c', 'border_color': '#388e3c', 'bg_color': '#c8e6c9'},  # グリーン濃
        {'number': 15, 'text_color': '#7b1fa2', 'border_color': '#7b1fa2', 'bg_color': '#e1bee7'},  # パープル濃
        {'number': 16, 'text_color': '#00897b', 'border_color': '#00897b', 'bg_color': '#b2dfdb'},  # ティール濃

        # パターン17～24: 背景色のみ（枠線なし）
        {'number': 17, 'text_color': '#000000', 'border_color': 'none', 'bg_color': '#e3f2fd'},
        {'number': 18, 'text_color': '#000000', 'border_color': 'none', 'bg_color': '#e8f5e9'},
        {'number': 19, 'text_color': '#000000', 'border_color': 'none', 'bg_color': '#f3e5f5'},
        {'number': 20, 'text_color': '#000000', 'border_color': 'none', 'bg_color': '#e0f2f1'},
        {'number': 21, 'text_color': '#000000', 'border_color': 'none', 'bg_color': '#bbdefb'},
        {'number': 22, 'text_color': '#000000', 'border_color': 'none', 'bg_color': '#c8e6c9'},
        {'number': 23, 'text_color': '#000000', 'border_color': 'none', 'bg_color': '#e1bee7'},
        {'number': 24, 'text_color': '#000000', 'border_color': 'none', 'bg_color': '#b2dfdb'},

        # パターン25: テーブルヘッダー背景色
        {'number': 25, 'text_color': '#000000', 'border_color': 'none', 'bg_color': '#f8f9fa'},
    ]

    # 単色パターン - モノクロ
    monochrome_patterns = [
        {'number': 0, 'text_color': '#6c757d', 'border_color': '#6c757d', 'bg_color': '#ffffff'},
        # パターン1-5: 文字色#000000、枠線濃い色、背景薄い色（5段階）
        {'number': 1, 'text_color': '#000000', 'border_color': '#495057', 'bg_color': '#f8f9fa'},
        {'number': 2, 'text_color': '#000000', 'border_color': '#495057', 'bg_color': '#f5f5f5'},
        {'number': 3, 'text_color': '#000000', 'border_color': '#495057', 'bg_color': '#e9ecef'},
        {'number': 4, 'text_color': '#000000', 'border_color': '#495057', 'bg_color': '#dee2e6'},
        {'number': 5, 'text_color': '#000000', 'border_color': '#495057', 'bg_color': '#ced4da'},
        # パターン6-10: 文字色と枠線が濃い色（同じ）、背景は1-5と同じ
        {'number': 6, 'text_color': '#495057', 'border_color': '#495057', 'bg_color': '#f8f9fa'},
        {'number': 7, 'text_color': '#495057', 'border_color': '#495057', 'bg_color': '#f5f5f5'},
        {'number': 8, 'text_color': '#495057', 'border_color': '#495057', 'bg_color': '#e9ecef'},
        {'number': 9, 'text_color': '#ffffff', 'border_color': '#495057', 'bg_color': '#dee2e6'},
        {'number': 10, 'text_color': '#ffffff', 'border_color': '#495057', 'bg_color': '#ced4da'},
        # パターン11-15: 背景のみ（1-5の背景と同じ）、枠線なし
        {'number': 11, 'text_color': '#000000', 'border_color': 'none', 'bg_color': '#f8f9fa'},
        {'number': 12, 'text_color': '#000000', 'border_color': 'none', 'bg_color': '#f5f5f5'},
        {'number': 13, 'text_color': '#000000', 'border_color': 'none', 'bg_color': '#e9ecef'},
        {'number': 14, 'text_color': '#000000', 'border_color': 'none', 'bg_color': '#dee2e6'},
        {'number': 15, 'text_color': '#000000', 'border_color': 'none', 'bg_color': '#ced4da'},
        # パターン16: テーブルヘッダー
        {'number': 16, 'text_color': '#000000', 'border_color': 'none', 'bg_color': '#f8f9fa'},
    ]

    # 単色パターン - ブルー
    blue_patterns = [
        {'number': 0, 'text_color': '#6c757d', 'border_color': '#6c757d', 'bg_color': '#ffffff'},
        # パターン1-5
        {'number': 1, 'text_color': '#000000', 'border_color': '#084298', 'bg_color': '#e7f3ff'},
        {'number': 2, 'text_color': '#000000', 'border_color': '#084298', 'bg_color': '#cfe2ff'},
        {'number': 3, 'text_color': '#000000', 'border_color': '#084298', 'bg_color': '#9ec5fe'},
        {'number': 4, 'text_color': '#000000', 'border_color': '#084298', 'bg_color': '#6ea8fe'},
        {'number': 5, 'text_color': '#000000', 'border_color': '#084298', 'bg_color': '#3d8bfd'},
        # パターン6-10
        {'number': 6, 'text_color': '#084298', 'border_color': '#084298', 'bg_color': '#e7f3ff'},
        {'number': 7, 'text_color': '#084298', 'border_color': '#084298', 'bg_color': '#cfe2ff'},
        {'number': 8, 'text_color': '#084298', 'border_color': '#084298', 'bg_color': '#9ec5fe'},
        {'number': 9, 'text_color': '#ffffff', 'border_color': '#0a58ca', 'bg_color': '#6ea8fe'},
        {'number': 10, 'text_color': '#ffffff', 'border_color': '#0a58ca', 'bg_color': '#3d8bfd'},
        # パターン11-15
        {'number': 11, 'text_color': '#000000', 'border_color': 'none', 'bg_color': '#e7f3ff'},
        {'number': 12, 'text_color': '#000000', 'border_color': 'none', 'bg_color': '#cfe2ff'},
        {'number': 13, 'text_color': '#000000', 'border_color': 'none', 'bg_color': '#9ec5fe'},
        {'number': 14, 'text_color': '#000000', 'border_color': 'none', 'bg_color': '#6ea8fe'},
        {'number': 15, 'text_color': '#000000', 'border_color': 'none', 'bg_color': '#3d8bfd'},
        # パターン16
        {'number': 16, 'text_color': '#000000', 'border_color': 'none', 'bg_color': '#e7f3ff'},
    ]

    # 単色パターン - グリーン
    green_patterns = [
        {'number': 0, 'text_color': '#6c757d', 'border_color': '#6c757d', 'bg_color': '#ffffff'},
        # パターン1-5
        {'number': 1, 'text_color': '#000000', 'border_color': '#2e7d32', 'bg_color': '#e8f5e9'},
        {'number': 2, 'text_color': '#000000', 'border_color': '#2e7d32', 'bg_color': '#c8e6c9'},
        {'number': 3, 'text_color': '#000000', 'border_color': '#2e7d32', 'bg_color': '#a5d6a7'},
        {'number': 4, 'text_color': '#000000', 'border_color': '#2e7d32', 'bg_color': '#81c784'},
        {'number': 5, 'text_color': '#000000', 'border_color': '#2e7d32', 'bg_color': '#66bb6a'},
        # パターン6-10
        {'number': 6, 'text_color': '#2e7d32', 'border_color': '#2e7d32', 'bg_color': '#e8f5e9'},
        {'number': 7, 'text_color': '#2e7d32', 'border_color': '#2e7d32', 'bg_color': '#c8e6c9'},
        {'number': 8, 'text_color': '#2e7d32', 'border_color': '#2e7d32', 'bg_color': '#a5d6a7'},
        {'number': 9, 'text_color': '#ffffff', 'border_color': '#2e7d32', 'bg_color': '#81c784'},
        {'number': 10, 'text_color': '#ffffff', 'border_color': '#2e7d32', 'bg_color': '#66bb6a'},
        # パターン11-15
        {'number': 11, 'text_color': '#000000', 'border_color': 'none', 'bg_color': '#e8f5e9'},
        {'number': 12, 'text_color': '#000000', 'border_color': 'none', 'bg_color': '#c8e6c9'},
        {'number': 13, 'text_color': '#000000', 'border_color': 'none', 'bg_color': '#a5d6a7'},
        {'number': 14, 'text_color': '#000000', 'border_color': 'none', 'bg_color': '#81c784'},
        {'number': 15, 'text_color': '#000000', 'border_color': 'none', 'bg_color': '#66bb6a'},
        # パターン16
        {'number': 16, 'text_color': '#000000', 'border_color': 'none', 'bg_color': '#e8f5e9'},
    ]

    # 単色パターン - レッド
    red_patterns = [
        {'number': 0, 'text_color': '#6c757d', 'border_color': '#6c757d', 'bg_color': '#ffffff'},
        # パターン1-5
        {'number': 1, 'text_color': '#000000', 'border_color': '#c62828', 'bg_color': '#ffebee'},
        {'number': 2, 'text_color': '#000000', 'border_color': '#c62828', 'bg_color': '#ffcdd2'},
        {'number': 3, 'text_color': '#000000', 'border_color': '#c62828', 'bg_color': '#ef9a9a'},
        {'number': 4, 'text_color': '#000000', 'border_color': '#c62828', 'bg_color': '#e57373'},
        {'number': 5, 'text_color': '#000000', 'border_color': '#c62828', 'bg_color': '#ef5350'},
        # パターン6-10
        {'number': 6, 'text_color': '#c62828', 'border_color': '#c62828', 'bg_color': '#ffebee'},
        {'number': 7, 'text_color': '#c62828', 'border_color': '#c62828', 'bg_color': '#ffcdd2'},
        {'number': 8, 'text_color': '#c62828', 'border_color': '#c62828', 'bg_color': '#ef9a9a'},
        {'number': 9, 'text_color': '#ffffff', 'border_color': '#c62828', 'bg_color': '#e57373'},
        {'number': 10, 'text_color': '#ffffff', 'border_color': '#c62828', 'bg_color': '#ef5350'},
        # パターン11-15
        {'number': 11, 'text_color': '#000000', 'border_color': 'none', 'bg_color': '#ffebee'},
        {'number': 12, 'text_color': '#000000', 'border_color': 'none', 'bg_color': '#ffcdd2'},
        {'number': 13, 'text_color': '#000000', 'border_color': 'none', 'bg_color': '#ef9a9a'},
        {'number': 14, 'text_color': '#000000', 'border_color': 'none', 'bg_color': '#e57373'},
        {'number': 15, 'text_color': '#000000', 'border_color': 'none', 'bg_color': '#ef5350'},
        # パターン16
        {'number': 16, 'text_color': '#000000', 'border_color': 'none', 'bg_color': '#ffebee'},
    ]

    # 単色パターン - パープル
    purple_patterns = [
        {'number': 0, 'text_color': '#6c757d', 'border_color': '#6c757d', 'bg_color': '#ffffff'},
        # パターン1-5
        {'number': 1, 'text_color': '#000000', 'border_color': '#6a1b9a', 'bg_color': '#f3e5f5'},
        {'number': 2, 'text_color': '#000000', 'border_color': '#6a1b9a', 'bg_color': '#e1bee7'},
        {'number': 3, 'text_color': '#000000', 'border_color': '#6a1b9a', 'bg_color': '#ce93d8'},
        {'number': 4, 'text_color': '#000000', 'border_color': '#6a1b9a', 'bg_color': '#ba68c8'},
        {'number': 5, 'text_color': '#000000', 'border_color': '#6a1b9a', 'bg_color': '#ab47bc'},
        # パターン6-10
        {'number': 6, 'text_color': '#6a1b9a', 'border_color': '#6a1b9a', 'bg_color': '#f3e5f5'},
        {'number': 7, 'text_color': '#6a1b9a', 'border_color': '#6a1b9a', 'bg_color': '#e1bee7'},
        {'number': 8, 'text_color': '#6a1b9a', 'border_color': '#6a1b9a', 'bg_color': '#ce93d8'},
        {'number': 9, 'text_color': '#ffffff', 'border_color': '#6a1b9a', 'bg_color': '#ba68c8'},
        {'number': 10, 'text_color': '#ffffff', 'border_color': '#6a1b9a', 'bg_color': '#ab47bc'},
        # パターン11-15
        {'number': 11, 'text_color': '#000000', 'border_color': 'none', 'bg_color': '#f3e5f5'},
        {'number': 12, 'text_color': '#000000', 'border_color': 'none', 'bg_color': '#e1bee7'},
        {'number': 13, 'text_color': '#000000', 'border_color': 'none', 'bg_color': '#ce93d8'},
        {'number': 14, 'text_color': '#000000', 'border_color': 'none', 'bg_color': '#ba68c8'},
        {'number': 15, 'text_color': '#000000', 'border_color': 'none', 'bg_color': '#ab47bc'},
        # パターン16
        {'number': 16, 'text_color': '#000000', 'border_color': 'none', 'bg_color': '#f3e5f5'},
    ]

    # 単色パターン - オレンジ
    orange_patterns = [
        {'number': 0, 'text_color': '#6c757d', 'border_color': '#6c757d', 'bg_color': '#ffffff'},
        # パターン1-5
        {'number': 1, 'text_color': '#000000', 'border_color': '#984c0c', 'bg_color': '#ffe5d0'},
        {'number': 2, 'text_color': '#000000', 'border_color': '#984c0c', 'bg_color': '#ffc9a0'},
        {'number': 3, 'text_color': '#000000', 'border_color': '#984c0c', 'bg_color': '#ffad70'},
        {'number': 4, 'text_color': '#000000', 'border_color': '#984c0c', 'bg_color': '#ff9040'},
        {'number': 5, 'text_color': '#ffffff', 'border_color': '#984c0c', 'bg_color': '#fd7e14'},
        # パターン6-10
        {'number': 6, 'text_color': '#984c0c', 'border_color': '#984c0c', 'bg_color': '#ffe5d0'},
        {'number': 7, 'text_color': '#984c0c', 'border_color': '#984c0c', 'bg_color': '#ffc9a0'},
        {'number': 8, 'text_color': '#984c0c', 'border_color': '#984c0c', 'bg_color': '#ffad70'},
        {'number': 9, 'text_color': '#ffffff', 'border_color': '#cc6510', 'bg_color': '#ff9040'},
        {'number': 10, 'text_color': '#ffffff', 'border_color': '#cc6510', 'bg_color': '#fd7e14'},
        # パターン11-15
        {'number': 11, 'text_color': '#000000', 'border_color': 'none', 'bg_color': '#ffe5d0'},
        {'number': 12, 'text_color': '#000000', 'border_color': 'none', 'bg_color': '#ffc9a0'},
        {'number': 13, 'text_color': '#000000', 'border_color': 'none', 'bg_color': '#ffad70'},
        {'number': 14, 'text_color': '#000000', 'border_color': 'none', 'bg_color': '#ff9040'},
        {'number': 15, 'text_color': '#ffffff', 'border_color': 'none', 'bg_color': '#fd7e14'},
        # パターン16
        {'number': 16, 'text_color': '#000000', 'border_color': 'none', 'bg_color': '#ffe5d0'},
    ]

    # 単色パターン - ティール
    teal_patterns = [
        {'number': 0, 'text_color': '#6c757d', 'border_color': '#6c757d', 'bg_color': '#ffffff'},
        # パターン1-5
        {'number': 1, 'text_color': '#000000', 'border_color': '#087990', 'bg_color': '#cff4fc'},
        {'number': 2, 'text_color': '#000000', 'border_color': '#087990', 'bg_color': '#9eeaf9'},
        {'number': 3, 'text_color': '#000000', 'border_color': '#087990', 'bg_color': '#6edff6'},
        {'number': 4, 'text_color': '#000000', 'border_color': '#087990', 'bg_color': '#3dd5f3'},
        {'number': 5, 'text_color': '#ffffff', 'border_color': '#087990', 'bg_color': '#0dcaf0'},
        # パターン6-10
        {'number': 6, 'text_color': '#087990', 'border_color': '#087990', 'bg_color': '#cff4fc'},
        {'number': 7, 'text_color': '#087990', 'border_color': '#087990', 'bg_color': '#9eeaf9'},
        {'number': 8, 'text_color': '#087990', 'border_color': '#087990', 'bg_color': '#6edff6'},
        {'number': 9, 'text_color': '#ffffff', 'border_color': '#0aa2c0', 'bg_color': '#3dd5f3'},
        {'number': 10, 'text_color': '#ffffff', 'border_color': '#0aa2c0', 'bg_color': '#0dcaf0'},
        # パターン11-15
        {'number': 11, 'text_color': '#000000', 'border_color': 'none', 'bg_color': '#cff4fc'},
        {'number': 12, 'text_color': '#000000', 'border_color': 'none', 'bg_color': '#9eeaf9'},
        {'number': 13, 'text_color': '#000000', 'border_color': 'none', 'bg_color': '#6edff6'},
        {'number': 14, 'text_color': '#000000', 'border_color': 'none', 'bg_color': '#3dd5f3'},
        {'number': 15, 'text_color': '#ffffff', 'border_color': 'none', 'bg_color': '#0dcaf0'},
        # パターン16
        {'number': 16, 'text_color': '#000000', 'border_color': 'none', 'bg_color': '#cff4fc'},
    ]

    # アニメーションパターン
    animation_patterns = [
        {
            'number': 0,
            'name': 'フロート（浮き上がり）',
            'normal': {'text_color': '#6c757d', 'border_color': '#6c757d', 'bg_color': '#ffffff'},
            'animation': 'translateY(-1px) + box-shadow',
            'transition': '0.2s ease',
            'description': 'カーソルを合わせると1px上に浮き上がり、軽い影が追加されます（色変化なし）'
        }
    ]

    context = {
        'base_colors': base_colors,
        'client_list_colors': client_list_colors,
        'assessment_colors': assessment_colors,
        'client_detail_colors': client_detail_colors,
        'color_patterns': color_patterns,
        'warm_color_patterns': warm_color_patterns,
        'cool_color_patterns': cool_color_patterns,
        'monochrome_patterns': monochrome_patterns,
        'blue_patterns': blue_patterns,
        'green_patterns': green_patterns,
        'red_patterns': red_patterns,
        'purple_patterns': purple_patterns,
        'orange_patterns': orange_patterns,
        'teal_patterns': teal_patterns,
        'animation_patterns': animation_patterns,
        'pattern_ranges': {
            'range_1_5': list(range(1, 6)),
            'range_6_10': list(range(6, 11)),
            'range_11_15': list(range(11, 16)),
        }
    }

    return render(request, 'clients/color_reference.html', context)


# ====================================
# ユーザー管理
# ====================================

@login_required
def user_list(request):
    """ユーザー一覧"""
    from django.contrib.auth.models import User
    from .models import UserProfile

    # 検索クエリ
    search_query = request.GET.get('search', '')
    department_filter = request.GET.get('department', '')
    role_filter = request.GET.get('role', '')
    status_filter = request.GET.get('status', '')

    # ユーザー一覧を取得（UserProfileと結合）
    users = User.objects.select_related('profile').all()

    # 検索フィルタ
    if search_query:
        # Q オブジェクトを構築
        query = (
            Q(username__icontains=search_query) |
            Q(profile__last_name__icontains=search_query) |
            Q(profile__first_name__icontains=search_query) |
            Q(profile__last_name_kana__icontains=search_query) |
            Q(profile__first_name_kana__icontains=search_query)
        )

        users = users.filter(query)

    # 所属部署フィルタ
    if department_filter:
        users = users.filter(profile__department=department_filter)

    # 役職フィルタ
    if role_filter:
        users = users.filter(profile__role=role_filter)

    # ステータスフィルタ
    if status_filter == 'active':
        users = users.filter(profile__is_active=True)
    elif status_filter == 'inactive':
        users = users.filter(profile__is_active=False)

    # ページネーション
    paginator = Paginator(users, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'page_obj': page_obj,
        'search_query': search_query,
        'department_filter': department_filter,
        'role_filter': role_filter,
        'status_filter': status_filter,
        'department_choices': [('', 'すべて')] + list(UserProfile.DEPARTMENT_CHOICES),
        'role_choices': UserProfile.ROLE_CHOICES,
        'status_choices': [
            ('', 'すべて'),
            ('active', '有効'),
            ('inactive', '無効'),
        ],
    }

    return render(request, 'users/user_list.html', context)


@login_required
def user_create(request):
    """ユーザー登録"""
    from django.contrib.auth.models import User
    from .models import UserProfile
    from django import forms

    class UserCreateForm(forms.Form):
        username = forms.CharField(label='ログインID', max_length=150)
        password = forms.CharField(label='パスワード', widget=forms.PasswordInput)
        password_confirm = forms.CharField(label='パスワード（確認）', widget=forms.PasswordInput)

        last_name = forms.CharField(label='姓', max_length=50)
        first_name = forms.CharField(label='名', max_length=50)
        last_name_kana = forms.CharField(label='セイ', max_length=50, required=False)
        first_name_kana = forms.CharField(label='メイ', max_length=50, required=False)

        organization = forms.ChoiceField(label='事業所', choices=UserProfile.ORGANIZATION_CHOICES, initial='annotsuroman')
        job_type = forms.CharField(required=False)
        department = forms.ChoiceField(label='部署', choices=UserProfile.DEPARTMENT_CHOICES)

        # 居宅介護支援事業所
        from .models import HomeCareSupportOffice
        home_care_office = forms.ModelChoiceField(
            label='所属居宅介護支援事業所',
            queryset=HomeCareSupportOffice.objects.filter(is_active=True),
            required=False,
            empty_label='選択してください'
        )

        def clean(self):
            cleaned_data = super().clean()
            password = cleaned_data.get('password')
            password_confirm = cleaned_data.get('password_confirm')

            if password and password_confirm and password != password_confirm:
                raise forms.ValidationError('パスワードが一致しません。')

            return cleaned_data

        def clean_username(self):
            username = self.cleaned_data.get('username')
            if User.objects.filter(username=username).exists():
                raise forms.ValidationError('このログインIDは既に使用されています。')
            return username

    if request.method == 'POST':
        form = UserCreateForm(request.POST)
        if form.is_valid():
            # Userを作成
            user = User.objects.create_user(
                username=form.cleaned_data['username'],
                password=form.cleaned_data['password'],
            )

            # UserProfileを作成または更新
            profile, created = UserProfile.objects.get_or_create(user=user)
            profile.last_name = form.cleaned_data['last_name']
            profile.first_name = form.cleaned_data['first_name']
            profile.last_name_kana = form.cleaned_data.get('last_name_kana', '')
            profile.first_name_kana = form.cleaned_data.get('first_name_kana', '')
            profile.organization = form.cleaned_data['organization']
            profile.job_type = form.cleaned_data.get('job_type', '')
            profile.department = form.cleaned_data['department']
            profile.home_care_office = form.cleaned_data.get('home_care_office')
            profile.role = 'staff'  # デフォルトで一般スタッフ
            profile.is_active = True  # デフォルトで有効
            profile.save()

            messages.success(request, f'ユーザー「{profile.get_full_name()}」を登録しました。')
            return redirect('user_list')
    else:
        form = UserCreateForm()

    context = {
        'form': form,
    }

    return render(request, 'users/user_form.html', context)


@login_required
def user_edit(request, pk):
    """ユーザー編集"""
    from django.contrib.auth.models import User
    from .models import UserProfile
    from django import forms

    user = get_object_or_404(User, pk=pk)
    profile, created = UserProfile.objects.get_or_create(user=user)

    class UserEditForm(forms.Form):
        username = forms.CharField(label='ログインID', max_length=150)

        last_name = forms.CharField(label='姓', max_length=50)
        first_name = forms.CharField(label='名', max_length=50)
        last_name_kana = forms.CharField(label='セイ', max_length=50, required=False)
        first_name_kana = forms.CharField(label='メイ', max_length=50, required=False)

        organization = forms.ChoiceField(label='事業所', choices=UserProfile.ORGANIZATION_CHOICES)
        job_type = forms.CharField(required=False)
        department = forms.ChoiceField(label='部署', choices=UserProfile.DEPARTMENT_CHOICES)

        # 居宅介護支援事業所
        from .models import HomeCareSupportOffice
        home_care_office = forms.ModelChoiceField(
            label='所属居宅介護支援事業所',
            queryset=HomeCareSupportOffice.objects.filter(is_active=True),
            required=False,
            empty_label='選択してください'
        )

        def clean_username(self):
            username = self.cleaned_data.get('username')
            if User.objects.filter(username=username).exclude(pk=user.pk).exists():
                raise forms.ValidationError('このログインIDは既に使用されています。')
            return username

    if request.method == 'POST':
        form = UserEditForm(request.POST)
        if form.is_valid():
            # Userを更新
            user.username = form.cleaned_data['username']
            user.save()

            # UserProfileを更新
            profile.last_name = form.cleaned_data['last_name']
            profile.first_name = form.cleaned_data['first_name']
            profile.last_name_kana = form.cleaned_data.get('last_name_kana', '')
            profile.first_name_kana = form.cleaned_data.get('first_name_kana', '')
            profile.organization = form.cleaned_data['organization']
            profile.job_type = form.cleaned_data.get('job_type', '')
            profile.department = form.cleaned_data['department']
            profile.home_care_office = form.cleaned_data.get('home_care_office')
            profile.save()

            messages.success(request, f'ユーザー「{profile.get_full_name()}」を更新しました。')
            return redirect('user_list')
    else:
        form = UserEditForm(initial={
            'username': user.username,
            'last_name': profile.last_name,
            'first_name': profile.first_name,
            'last_name_kana': profile.last_name_kana,
            'first_name_kana': profile.first_name_kana,
            'organization': profile.organization,
            'job_type': profile.job_type,
            'department': profile.department,
            'home_care_office': profile.home_care_office,
        })

    context = {
        'form': form,
        'user': user,
        'profile': profile,
        'is_edit': True,
    }

    return render(request, 'users/user_form.html', context)


@login_required
def user_delete(request, pk):
    """ユーザー削除"""
    from django.contrib.auth.models import User

    user = get_object_or_404(User, pk=pk)
    profile = getattr(user, 'profile', None)

    if request.method == 'POST':
        user_name = profile.get_full_name() if profile else user.username
        user.delete()
        messages.success(request, f'ユーザー「{user_name}」を削除しました。')
        return redirect('user_list')

    context = {
        'user': user,
        'profile': profile,
    }

    return render(request, 'users/user_confirm_delete.html', context)


# ===== フィードバックシステム =====

@login_required
def feedback_submit(request):
    """フィードバック投稿"""
    if request.method == 'POST':
        form = FeedbackForm(request.POST)
        if form.is_valid():
            feedback = form.save(commit=False)
            feedback.user = request.user
            feedback.page_url = request.POST.get('page_url', '')
            feedback.save()
            messages.success(request, 'フィードバックを送信しました。ご協力ありがとうございます。')
            return JsonResponse({'success': True, 'message': '送信しました'})
        else:
            return JsonResponse({'success': False, 'errors': form.errors}, status=400)
    else:
        form = FeedbackForm()

    context = {
        'form': form,
    }
    return render(request, 'feedback/feedback_form.html', context)


@login_required
def feedback_list_admin(request):
    """管理者専用: フィードバック一覧"""
    # 管理者権限チェック
    if not request.user.is_staff and not request.user.is_superuser:
        messages.error(request, 'このページにアクセスする権限がありません。')
        return redirect('client_list')

    # フィルター処理
    feedbacks = Feedback.objects.select_related('user').all()

    # カテゴリフィルター
    category_filter = request.GET.get('category', '')
    if category_filter:
        feedbacks = feedbacks.filter(category=category_filter)

    # ステータスフィルター
    status_filter = request.GET.get('status', '')
    if status_filter:
        feedbacks = feedbacks.filter(status=status_filter)

    # 優先度フィルター
    priority_filter = request.GET.get('priority', '')
    if priority_filter:
        feedbacks = feedbacks.filter(priority=priority_filter)

    # 検索
    search_query = request.GET.get('search', '')
    if search_query:
        feedbacks = feedbacks.filter(
            Q(title__icontains=search_query) |
            Q(content__icontains=search_query) |
            Q(user__username__icontains=search_query)
        )

    # ページネーション
    paginator = Paginator(feedbacks, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'page_obj': page_obj,
        'category_choices': Feedback.CATEGORY_CHOICES,
        'status_choices': Feedback.STATUS_CHOICES,
        'priority_choices': Feedback.PRIORITY_CHOICES,
        'category_filter': category_filter,
        'status_filter': status_filter,
        'priority_filter': priority_filter,
        'search_query': search_query,
    }

    return render(request, 'feedback/feedback_list_admin.html', context)


@login_required
def feedback_detail_admin(request, pk):
    """管理者専用: フィードバック詳細"""
    # 管理者権限チェック
    if not request.user.is_staff and not request.user.is_superuser:
        messages.error(request, 'このページにアクセスする権限がありません。')
        return redirect('client_list')

    feedback = get_object_or_404(Feedback, pk=pk)

    # 返信一覧を取得（古い順）
    replies = feedback.replies.all()

    # 返信フォーム
    reply_form = FeedbackReplyForm()

    context = {
        'feedback': feedback,
        'replies': replies,
        'reply_form': reply_form,
    }

    return render(request, 'feedback/feedback_detail_admin.html', context)


@login_required
def feedback_status_update(request, pk):
    """フィードバックステータス更新（AJAX）"""
    # 管理者権限チェック
    if not request.user.is_staff and not request.user.is_superuser:
        return JsonResponse({'success': False, 'message': '権限がありません'}, status=403)

    if request.method == 'POST':
        try:
            feedback = get_object_or_404(Feedback, pk=pk)

            # JSONデータを読み取る
            data = json.loads(request.body)
            new_status = data.get('status')

            if new_status in dict(Feedback.STATUS_CHOICES):
                feedback.status = new_status
                feedback.save()
                return JsonResponse({
                    'success': True,
                    'message': 'ステータスを更新しました',
                    'status': feedback.get_status_display()
                })
            else:
                return JsonResponse({'success': False, 'message': '無効なステータスです'}, status=400)
        except json.JSONDecodeError:
            return JsonResponse({'success': False, 'message': 'JSONのパースに失敗しました'}, status=400)
        except Exception as e:
            return JsonResponse({'success': False, 'message': f'エラーが発生しました: {str(e)}'}, status=500)

    return JsonResponse({'success': False, 'message': '無効なリクエストです'}, status=400)


@login_required
def feedback_reply_submit(request, pk):
    """フィードバックへの返信投稿"""
    # 管理者権限チェック
    if not request.user.is_staff and not request.user.is_superuser:
        messages.error(request, '返信する権限がありません。')
        return redirect('feedback_detail_admin', pk=pk)

    feedback = get_object_or_404(Feedback, pk=pk)

    if request.method == 'POST':
        form = FeedbackReplyForm(request.POST)
        if form.is_valid():
            reply = form.save(commit=False)
            reply.feedback = feedback
            reply.user = request.user
            reply.save()
            messages.success(request, '返信を投稿しました。')
            return redirect('feedback_detail_admin', pk=pk)
        else:
            messages.error(request, '返信の投稿に失敗しました。入力内容を確認してください。')

    return redirect('feedback_detail_admin', pk=pk)


@login_required
def feedback_edit_admin(request, pk):
    """フィードバック編集（管理者専用）"""
    # 管理者権限チェック
    if not request.user.is_staff and not request.user.is_superuser:
        messages.error(request, 'フィードバックを編集する権限がありません。')
        return redirect('client_list')

    feedback = get_object_or_404(Feedback, pk=pk)

    if request.method == 'POST':
        form = FeedbackEditForm(request.POST, instance=feedback)
        if form.is_valid():
            form.save()  # updated_atは自動更新される
            messages.success(request, 'フィードバックを更新しました。')
            return redirect('feedback_detail_admin', pk=pk)
    else:
        form = FeedbackEditForm(instance=feedback)

    context = {
        'feedback': feedback,
        'form': form,
    }
    return render(request, 'feedback/feedback_edit_admin.html', context)


@login_required
def feedback_reply_edit(request, pk):
    """返信編集（管理者専用）"""
    # 管理者権限チェック
    if not request.user.is_staff and not request.user.is_superuser:
        messages.error(request, '返信を編集する権限がありません。')
        return redirect('client_list')

    reply = get_object_or_404(FeedbackReply, pk=pk)
    feedback_pk = reply.feedback.pk

    if request.method == 'POST':
        form = FeedbackReplyForm(request.POST, instance=reply)
        if form.is_valid():
            form.save()
            messages.success(request, '返信を更新しました。')
            return redirect('feedback_detail_admin', pk=feedback_pk)
        else:
            messages.error(request, '返信の更新に失敗しました。')
    else:
        form = FeedbackReplyForm(instance=reply)

    context = {
        'reply': reply,
        'feedback': reply.feedback,
        'form': form,
    }
    return render(request, 'feedback/feedback_reply_edit.html', context)


@login_required
def feedback_reply_delete(request, pk):
    """返信削除（管理者専用）"""
    # 管理者権限チェック
    if not request.user.is_staff and not request.user.is_superuser:
        messages.error(request, '返信を削除する権限がありません。')
        return redirect('client_list')

    reply = get_object_or_404(FeedbackReply, pk=pk)
    feedback_pk = reply.feedback.pk

    if request.method == 'POST':
        reply.delete()
        messages.success(request, '返信を削除しました。')
        return redirect('feedback_detail_admin', pk=feedback_pk)

    return redirect('feedback_detail_admin', pk=feedback_pk)


@login_required
def feedback_delete_admin(request, pk):
    """フィードバック削除（管理者専用）"""
    # 管理者権限チェック
    if not request.user.is_staff and not request.user.is_superuser:
        messages.error(request, 'フィードバックを削除する権限がありません。')
        return redirect('client_list')

    feedback = get_object_or_404(Feedback, pk=pk)

    if request.method == 'POST':
        feedback.delete()
        messages.success(request, 'フィードバックを削除しました。')
        return redirect('feedback_list_admin')

    return redirect('feedback_list_admin')


@login_required
def my_feedback_list(request):
    """ユーザー向け: 自分のフィードバック一覧"""
    feedbacks = Feedback.objects.filter(user=request.user).order_by('-created_at')

    context = {
        'feedbacks': feedbacks,
    }
    return render(request, 'feedback/my_feedback_list.html', context)


@login_required
def my_feedback_detail(request, pk):
    """ユーザー向け: フィードバック詳細（パスワード認証あり）"""
    feedback = get_object_or_404(Feedback, pk=pk)

    # 自分のフィードバック以外はアクセス拒否（管理者を除く）
    if not (request.user.is_staff or request.user.is_superuser):
        if feedback.user != request.user:
            messages.error(request, 'このフィードバックにアクセスする権限がありません。')
            return redirect('my_feedback_list')

    # 管理者の場合はパスワード不要
    if request.user.is_staff or request.user.is_superuser:
        # 返信一覧を取得
        replies = feedback.replies.all()

        context = {
            'feedback': feedback,
            'replies': replies,
            'is_authenticated': True,
        }
        return render(request, 'feedback/my_feedback_detail.html', context)

    # 一般ユーザー: 毎回パスワード認証が必要
    if request.method == 'POST':
        # パスワード認証
        entered_password = request.POST.get('password', '')
        if entered_password == feedback.password:
            # 認証成功 - 詳細を表示
            replies = feedback.replies.all()

            context = {
                'feedback': feedback,
                'replies': replies,
                'is_authenticated': True,
            }
            messages.success(request, 'パスワード認証に成功しました。')
            return render(request, 'feedback/my_feedback_detail.html', context)
        else:
            # 認証失敗 - パスワード入力画面を再表示
            messages.error(request, 'パスワードが正しくありません。')
            context = {
                'feedback': feedback,
                'is_authenticated': False,
            }
            return render(request, 'feedback/my_feedback_detail.html', context)

    # GETリクエスト: パスワード入力画面を表示
    context = {
        'feedback': feedback,
        'is_authenticated': False,
    }
    return render(request, 'feedback/my_feedback_detail.html', context)


# ========================================
# 居宅介護支援事業所管理
# ========================================

@login_required
def home_care_office_list(request):
    """居宅介護支援事業所一覧"""
    # 管理者権限チェック
    if not request.user.is_staff and not request.user.is_superuser:
        messages.error(request, 'このページにアクセスする権限がありません。')
        return redirect('client_list')

    offices = HomeCareSupportOffice.objects.all().order_by('name')
    context = {
        'offices': offices,
    }
    return render(request, 'clients/home_care_office_list.html', context)


@login_required
def home_care_office_create(request):
    """居宅介護支援事業所作成"""
    is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'

    if not request.user.is_staff and not request.user.is_superuser:
        if is_ajax:
            return JsonResponse({'success': False, 'error': 'アクセス権限がありません。'})
        messages.error(request, 'このページにアクセスする権限がありません。')
        return redirect('client_list')

    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        furigana = request.POST.get('furigana', '').strip()
        office_number = request.POST.get('office_number', '').strip()
        postal_code = request.POST.get('postal_code', '').strip()
        address = request.POST.get('address', '').strip()
        phone = request.POST.get('phone', '').strip()
        fax = request.POST.get('fax', '').strip()
        manager_name = request.POST.get('manager_name', '').strip()
        is_active = request.POST.get('is_active') == 'true'

        if not name or not office_number:
            if is_ajax:
                return JsonResponse({'success': False, 'error': '事業所名と事業所番号は必須です。'})
            messages.error(request, '事業所名と事業所番号は必須です。')
            return redirect('home_care_office_create')

        if HomeCareSupportOffice.objects.filter(office_number=office_number).exists():
            if is_ajax:
                return JsonResponse({'success': False, 'error': 'この事業所番号は既に登録されています。'})
            messages.error(request, 'この事業所番号は既に登録されています。')
            return redirect('home_care_office_create')

        office = HomeCareSupportOffice.objects.create(
            name=name,
            furigana=furigana,
            office_number=office_number,
            postal_code=postal_code,
            address=address,
            phone=phone,
            fax=fax,
            manager_name=manager_name,
            is_active=is_active,
        )
        if is_ajax:
            return JsonResponse({'success': True})
        messages.success(request, f'{office.name} を登録しました。')
        return redirect('home_care_office_list')

    return redirect('home_care_office_list')


@login_required
def home_care_office_edit(request, pk):
    """居宅介護支援事業所編集"""
    is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'

    if not request.user.is_staff and not request.user.is_superuser:
        if is_ajax:
            return JsonResponse({'success': False, 'error': 'アクセス権限がありません。'})
        messages.error(request, 'このページにアクセスする権限がありません。')
        return redirect('client_list')

    office = get_object_or_404(HomeCareSupportOffice, pk=pk)

    if request.method == 'POST':
        office.name = request.POST.get('name', '').strip()
        office.furigana = request.POST.get('furigana', '').strip()
        office.office_number = request.POST.get('office_number', '').strip()
        office.postal_code = request.POST.get('postal_code', '').strip()
        office.address = request.POST.get('address', '').strip()
        office.phone = request.POST.get('phone', '').strip()
        office.fax = request.POST.get('fax', '').strip()
        office.manager_name = request.POST.get('manager_name', '').strip()
        office.is_active = request.POST.get('is_active') == 'true'

        if not office.name or not office.office_number:
            if is_ajax:
                return JsonResponse({'success': False, 'error': '事業所名と事業所番号は必須です。'})
            messages.error(request, '事業所名と事業所番号は必須です。')
            return redirect('home_care_office_edit', pk=pk)

        if HomeCareSupportOffice.objects.filter(office_number=office.office_number).exclude(pk=pk).exists():
            if is_ajax:
                return JsonResponse({'success': False, 'error': 'この事業所番号は既に登録されています。'})
            messages.error(request, 'この事業所番号は既に登録されています。')
            return redirect('home_care_office_edit', pk=pk)

        office.save()
        if is_ajax:
            return JsonResponse({'success': True})
        messages.success(request, f'{office.name} を更新しました。')
        return redirect('home_care_office_list')

    return redirect('home_care_office_list')


@login_required
def home_care_office_delete(request, pk):
    """居宅介護支援事業所削除"""
    # 管理者権限チェック
    if not request.user.is_staff and not request.user.is_superuser:
        messages.error(request, 'このページにアクセスする権限がありません。')
        return redirect('client_list')

    office = get_object_or_404(HomeCareSupportOffice, pk=pk)

    if request.method == 'POST':
        name = office.name
        office.delete()
        messages.success(request, f'{name} を削除しました。')
        return redirect('home_care_office_list')

    return redirect('home_care_office_list')


# ========================================
# 介護保険認定情報 更新API
# ========================================

@login_required
def client_cert_info_update(request, client_id):
    """介護保険被保険者証の認定情報をDBに保存するAJAX endpoint"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'POST only'}, status=405)

    client = get_object_or_404(Client, pk=client_id)

    insurance_number   = request.POST.get('insurance_number', '').strip()
    care_level         = request.POST.get('care_level', '').strip()
    care_burden        = request.POST.get('care_burden', '').strip()
    cert_date_str      = request.POST.get('certification_date', '').strip()
    cert_start_str     = request.POST.get('cert_start', '').strip()
    cert_end_str       = request.POST.get('cert_end', '').strip()

    from datetime import date as date_cls
    def parse_date(s):
        if not s:
            return None
        try:
            return datetime.strptime(s, '%Y-%m-%d').date()
        except ValueError:
            return None

    def to_wareki_str(d):
        if not d:
            return ''
        y, m, day = d.year, d.month, d.day
        if d >= date_cls(2019, 5, 1):   return f"令和{y-2018}年{m}月{day}日"
        if d >= date_cls(1989, 1, 8):   return f"平成{y-1988}年{m}月{day}日"
        if d >= date_cls(1926, 12, 25): return f"昭和{y-1925}年{m}月{day}日"
        if d >= date_cls(1912, 7, 30):  return f"大正{y-1911}年{m}月{day}日"
        return f"明治{y-1867}年{m}月{day}日"

    if insurance_number:
        client.insurance_number = insurance_number
    if care_level:
        client.care_level = care_level
    client.care_burden            = care_burden
    client.certification_date     = parse_date(cert_date_str)
    client.certification_period_start = parse_date(cert_start_str)
    client.certification_period_end   = parse_date(cert_end_str)
    client.save()

    cert_date_obj  = client.certification_date
    cert_start_obj = client.certification_period_start
    cert_end_obj   = client.certification_period_end

    return JsonResponse({
        'success': True,
        'insurance_number':    client.insurance_number,
        'care_level_value':    client.care_level,
        'care_level_display':  client.get_care_level_display(),
        'care_burden':         client.care_burden,
        'certification_date':  cert_date_obj.strftime('%Y-%m-%d') if cert_date_obj else '',
        'cert_date_wareki':    to_wareki_str(cert_date_obj),
        'cert_start':          cert_start_obj.strftime('%Y-%m-%d') if cert_start_obj else '',
        'cert_end':            cert_end_obj.strftime('%Y-%m-%d')   if cert_end_obj   else '',
        'cert_start_wareki':   to_wareki_str(cert_start_obj),
        'cert_end_wareki':     to_wareki_str(cert_end_obj),
    })


# ========================================
# 医療保険情報 更新API
# ========================================

@login_required
def client_medical_info_update(request, client_id):
    """医療保険情報をDBに保存するAJAX endpoint"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'POST only'}, status=405)

    client = get_object_or_404(Client, pk=client_id)

    medical_type   = request.POST.get('medical_insurance_type', '').strip()
    symbol         = request.POST.get('medical_insurance_symbol', '').strip()
    number         = request.POST.get('medical_insurance_number', '').strip()
    branch         = request.POST.get('medical_insurance_branch', '').strip()
    burden         = request.POST.get('medical_burden', '').strip()
    insurer_name   = request.POST.get('medical_insurer_name', '').strip()
    insurer_number = request.POST.get('medical_insurer_number', '').strip()

    client.medical_insurance_type      = medical_type
    client.medical_insurance_symbol    = symbol
    client.medical_insurance_number    = number
    client.medical_insurance_branch    = branch
    client.medical_burden              = burden
    client.medical_insurer_name_issuer = insurer_name
    client.medical_insurer_number      = insurer_number
    client.save()

    return JsonResponse({
        'success': True,
        'medical_insurance_type':         client.medical_insurance_type,
        'medical_insurance_type_display': client.get_medical_insurance_type_display(),
        'medical_insurance_symbol':       client.medical_insurance_symbol,
        'medical_insurance_number':       client.medical_insurance_number,
        'medical_insurance_branch':       client.medical_insurance_branch,
        'medical_burden':                 client.medical_burden,
        'medical_insurer_name':           client.medical_insurer_name_issuer,
        'medical_insurer_number':         client.medical_insurer_number,
    })


# ========================================
# 更新認定申請書 / 区分変更申請書
# ========================================

def _document_create_ltc_base(request, client_id, doc_type, doc_name, template_name):
    """更新認定申請書・区分変更申請書の共通ロジック"""
    from django.conf import settings as dj_settings
    from openpyxl import load_workbook
    from openpyxl.cell.cell import MergedCell
    import tempfile

    client = get_object_or_404(Client, pk=client_id)

    # ユーザープロフィールと事業所情報を取得
    profile = getattr(request.user, 'profile', None)
    office = getattr(profile, 'home_care_office', None)

    # 担当者氏名（ログインユーザーの氏名）
    user_full_name = profile.get_full_name() if profile else request.user.username

    # 和暦変換（表示用）
    def _to_wareki_display(d):
        if not d:
            return ''
        from datetime import date as date_cls
        y, m, day = d.year, d.month, d.day
        if d >= date_cls(2019, 5, 1):
            return f"令和{y-2018}年{m}月{day}日"
        elif d >= date_cls(1989, 1, 8):
            return f"平成{y-1988}年{m}月{day}日"
        elif d >= date_cls(1926, 12, 25):
            return f"昭和{y-1925}年{m}月{day}日"
        elif d >= date_cls(1912, 7, 30):
            return f"大正{y-1911}年{m}月{day}日"
        else:
            return f"明治{y-1867}年{m}月{day}日"

    # 初期値（自動反映）
    initial = {
        # 被保険者
        'insurance_number': client.insurance_number or '',
        'client_furigana': client.furigana or '',
        'client_name': client.name or '',
        'client_gender': '男' if client.gender == 'male' else ('女' if client.gender == 'female' else ''),
        'birth_date': client.birth_date.strftime('%Y-%m-%d') if client.birth_date else '',
        'postal_code': client.postal_code or '',
        'client_address': client.address or '',
        'client_phone': client.phone or '',
        # 前回の認定
        'care_level': client.get_care_level_display() if client.care_level else '',
        'cert_start': client.certification_period_start.strftime('%Y-%m-%d') if client.certification_period_start else '',
        'cert_end': client.certification_period_end.strftime('%Y-%m-%d') if client.certification_period_end else '',
        'cert_start_wareki': _to_wareki_display(client.certification_period_start),
        'cert_end_wareki': _to_wareki_display(client.certification_period_end),
        'certification_date': client.certification_date.strftime('%Y-%m-%d') if client.certification_date else '',
        'cert_date_wareki': _to_wareki_display(client.certification_date),
        'care_burden': client.care_burden or '',
        # 医療保険
        'medical_insurance_type': client.get_medical_insurance_type_display() if client.medical_insurance_type else '',
        'medical_insurer_name': client.medical_insurer_name_issuer or '',
        'medical_insurer_number': client.medical_insurer_number or '',
        'medical_insurance_symbol': client.medical_insurance_symbol or '',
        'medical_insurance_number': client.medical_insurance_number or '',
        # 申請書提出者（事業所）
        'office_furigana': (office.furigana if office else '') or '',
        'office_name': (office.name if office else '') or '',
        'office_number': (office.office_number if office else '') or '',
        'office_postal_code': (office.postal_code if office else '') or '',
        'office_address': (office.address if office else '') or '',
        'office_phone': (office.phone if office else '') or '',
        'staff_name': user_full_name,
        'relation': '介護支援専門員',
        # 手入力項目
        'application_date': datetime.today().strftime('%Y-%m-%d'),
        'specific_disease': '',
        'survey_location': 'サービス付き高齢者向け住宅　安濃津ろまん',
        'survey_location_other': '',
        'survey_location_postal': '',
        'survey_location_address': '',
        'survey_location_phone': '',
        'transferred_in': '',
        'transfer_from_municipality': '',
        'transfer_applied': '',
        'transfer_applied_date': '',
        'survey_contact_type': '担当ケアマネ',
        'survey_contact_name': '',
        'survey_contact_furigana': '',
        'survey_contact_relation': '',
        'survey_contact_phone': '',
        'survey_preferred_time': '',
        'survey_notes': '',
        'user_full_name': user_full_name,
        'user_full_name_kana': profile.get_full_name_kana() if profile else '',
        'user_phone': profile.phone if profile else '',
        'doctor_hospital': '',
        'doctor_name': '',
        'doctor_phone': '',
        'doctor_outside_city': '',
        'doctor_postal_code': '',
        'doctor_address': '',
        'doctor_notes': '',
        'change_reason': '',
    }

    # 履歴からの再編集：保存済みデータで initial を上書き
    history_id = request.GET.get('history_id')
    if history_id and request.method == 'GET':
        try:
            hist = DocumentCreationHistory.objects.get(pk=history_id, client=client)
            saved = hist.form_data or {}
            for k in initial:
                if k in saved:
                    initial[k] = saved[k]

            # care_level: 生値（例: "care4"）を表示値に変換
            care_level_map = {
                'independent': '自立',
                'support1': '要支援1', 'support2': '要支援2',
                'care1': '要介護1', 'care2': '要介護2', 'care3': '要介護3',
                'care4': '要介護4', 'care5': '要介護5',
            }
            cl = initial.get('care_level', '')
            if cl in care_level_map:
                initial['care_level'] = care_level_map[cl]

            # cert_start_wareki / cert_end_wareki: 保存済みISO日付から再計算
            for date_key, wareki_key in [('cert_start', 'cert_start_wareki'), ('cert_end', 'cert_end_wareki'), ('certification_date', 'cert_date_wareki')]:
                iso = initial.get(date_key, '')
                if iso:
                    try:
                        from datetime import datetime as _dt
                        d = _dt.strptime(iso, '%Y-%m-%d').date()
                        initial[wareki_key] = _to_wareki_display(d)
                    except Exception:
                        pass

            # ユーザー情報は常に最新を使用（保存値ではなくログインユーザーから）
            initial['user_full_name']      = user_full_name
            initial['user_full_name_kana'] = profile.get_full_name_kana() if profile else ''
            initial['user_phone']          = profile.phone if profile else ''
            initial['office_phone']        = (office.phone if office else '') or ''

        except DocumentCreationHistory.DoesNotExist:
            history_id = None

    if request.method == 'POST':
        # フォームデータを収集
        form_data = {k: request.POST.get(k, '') for k in initial}
        # care_level: 生値を表示値に変換して保存
        _care_map = {
            'independent': '自立',
            'support1': '要支援1', 'support2': '要支援2',
            'care1': '要介護1', 'care2': '要介護2', 'care3': '要介護3',
            'care4': '要介護4', 'care5': '要介護5',
        }
        if form_data.get('care_level') in _care_map:
            form_data['care_level'] = _care_map[form_data['care_level']]
        # cert_start_wareki / cert_end_wareki を再計算して保存
        for date_key, wareki_key in [('cert_start', 'cert_start_wareki'), ('cert_end', 'cert_end_wareki'), ('certification_date', 'cert_date_wareki')]:
            iso = form_data.get(date_key, '')
            if iso:
                try:
                    from datetime import datetime as _dt2
                    d2 = _dt2.strptime(iso, '%Y-%m-%d').date()
                    form_data[wareki_key] = _to_wareki_display(d2)
                except Exception:
                    pass
        # ユーザー情報を最新値で上書き保存
        form_data['user_full_name']      = user_full_name
        form_data['user_full_name_kana'] = profile.get_full_name_kana() if profile else ''
        form_data['user_phone']          = profile.phone if profile else ''
        form_data['office_phone']        = (office.phone if office else '') or ''
        action = request.POST.get('action', 'excel')
        history_id = request.GET.get('history_id')

        # 保存処理（保存して戻る / 保存してExcel出力 共通）
        if history_id:
            try:
                hist = DocumentCreationHistory.objects.get(pk=history_id, client=client)
                hist.form_data = form_data
                hist.save()
            except DocumentCreationHistory.DoesNotExist:
                history_id = None
        if not history_id:
            hist = DocumentCreationHistory.objects.create(
                client=client,
                document_type=doc_type,
                document_name=doc_name,
                form_data=form_data,
                status='draft',
                created_by=request.user,
            )

        if action == 'save':
            messages.success(request, f'{doc_name}を保存しました。')
            url = reverse('client_detail', kwargs={'pk': client_id}) + '#documents'
            return HttpResponseRedirect(url)

        try:
            from urllib.parse import quote
            if doc_type == 'ltc_change':
                content = _generate_ltc_change_excel_bytes(client, form_data)
            else:
                content = _generate_ltc_renewal_excel_bytes(client, form_data)
            dl_name = _make_dl_filename(client, doc_name)
            response = HttpResponse(
                content,
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f"attachment; filename=\"download.xlsx\"; filename*=UTF-8''{quote(dl_name, safe='')}"
            return response

        except Exception as e:
            messages.error(request, f'Excel生成中にエラーが発生しました: {str(e)}')
            return redirect('client_detail', pk=client_id)

    # 年齢計算（第2号被保険者判定: 40〜64歳）
    is_second_insured = False
    if client.birth_date:
        from datetime import date as _date_cls
        today = _date_cls.today()
        age = today.year - client.birth_date.year - (
            (today.month, today.day) < (client.birth_date.month, client.birth_date.day)
        )
        is_second_insured = 40 <= age <= 64

    specific_diseases = [
        'がん（末期）',
        '関節リウマチ',
        '筋萎縮性側索硬化症（ALS）',
        '後縦靱帯骨化症',
        '骨折を伴う骨粗鬆症',
        '初老期における認知症',
        '進行性核上性麻痺・大脳皮質基底核変性症・パーキンソン病関連疾患',
        '脊髄小脳変性症',
        '脊柱管狭窄症',
        '早老症',
        '多系統萎縮症',
        '糖尿病性神経障害・糖尿病性腎症・糖尿病性網膜症',
        '脳血管疾患',
        '閉塞性動脈硬化症',
        '慢性閉塞性肺疾患（COPD）',
        '両側の膝関節または股関節に著しい変形を伴う変形性関節症',
    ]

    # アセスメントから医師・医療機関情報を取得（医師情報が入っている最新を優先）
    from assessments.models import Assessment
    def _has_doctor(a):
        hs = a.health_status or {}
        return any([
            hs.get('main_doctor_hospital'),
            hs.get('visiting_doctor_hospital'),
            *(hs.get(f'family_doctor_hospital_{i}') for i in range(1, 5)),
        ])
    latest_assessment = next(
        (a for a in client.assessments.order_by('-assessment_date', '-created_at') if _has_doctor(a)),
        client.assessments.order_by('-assessment_date', '-created_at').first()
    )
    doctor_options = []
    if latest_assessment:
        hs = latest_assessment.health_status or {}
        if hs.get('main_doctor_hospital'):
            doctor_options.append({
                'type': '主治医',
                'hospital': hs.get('main_doctor_hospital', ''),
                'name': hs.get('main_doctor_name', ''),
            })
        if hs.get('visiting_doctor_hospital'):
            doctor_options.append({
                'type': '往診医',
                'hospital': hs.get('visiting_doctor_hospital', ''),
                'name': hs.get('visiting_doctor_name', ''),
            })
        for i in range(1, 5):
            hospital = hs.get(f'family_doctor_hospital_{i}', '')
            if hospital:
                doctor_options.append({
                    'type': 'かかりつけ医',
                    'hospital': hospital,
                    'name': hs.get(f'family_doctor_name_{i}', ''),
                })

    context = {
        'client': client,
        'initial': initial,
        'is_second_insured': is_second_insured,
        'specific_diseases': specific_diseases,
        'doctor_options': doctor_options,
        'history_id': history_id or '',
    }
    return render(request, template_name, context)


@login_required
@user_passes_test(staff_required)
def document_create_ltc_renewal(request, client_id):
    """更新認定申請書 作成画面"""
    return _document_create_ltc_base(
        request, client_id,
        doc_type='ltc_renewal',
        doc_name='更新認定申請書',
        template_name='clients/document_create_ltc_renewal.html',
    )


@login_required
@user_passes_test(staff_required)
def document_create_ltc_change(request, client_id):
    """区分変更申請書 作成画面"""
    return _document_create_ltc_base(
        request, client_id,
        doc_type='ltc_change',
        doc_name='区分変更申請書',
        template_name='clients/document_create_ltc_change.html',
    )


# ========================================
# 地域包括支援センター管理
# ========================================

@login_required
def support_center_list(request):
    """地域包括支援センター一覧"""
    # 管理者権限チェック
    if not request.user.is_staff and not request.user.is_superuser:
        messages.error(request, 'このページにアクセスする権限がありません。')
        return redirect('client_list')

    from django.db.models import Case, When, IntegerField

    # 地域の優先順位を定義
    area_order = Case(
        When(area='tsu', then=1),
        When(area='suzuka', then=2),
        When(area='kameyama', then=3),
        When(area='matsusaka', then=4),
        When(area='ise', then=5),
        When(area='yokkaichi', then=6),
        When(area='kuwana', then=7),
        When(area='iga', then=8),
        When(area='nabari', then=9),
        When(area='inabe', then=10),
        When(area='kumano', then=11),
        When(area='owase', then=12),
        When(area='toba', then=13),
        When(area='other', then=14),
        output_field=IntegerField(),
    )

    centers = RegionalSupportCenter.objects.all().order_by(area_order, 'name')
    context = {
        'centers': centers,
    }
    return render(request, 'clients/support_center_list.html', context)


@login_required
def support_center_create(request):
    """地域包括支援センター作成"""
    # 管理者権限チェック
    if not request.user.is_staff and not request.user.is_superuser:
        messages.error(request, 'このページにアクセスする権限がありません。')
        return redirect('client_list')

    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        office_number = request.POST.get('office_number', '').strip()
        postal_code = request.POST.get('postal_code', '').strip()
        address = request.POST.get('address', '').strip()
        phone = request.POST.get('phone', '').strip()
        area = request.POST.get('area', 'other')

        if not name or not office_number:
            messages.error(request, '事業所名と事業所番号は必須です。')
            return redirect('support_center_create')

        # 事業所番号の重複チェック
        if RegionalSupportCenter.objects.filter(office_number=office_number).exists():
            messages.error(request, 'この事業所番号は既に登録されています。')
            return redirect('support_center_create')

        center = RegionalSupportCenter.objects.create(
            name=name,
            office_number=office_number,
            postal_code=postal_code,
            address=address,
            phone=phone,
            area=area,
        )
        messages.success(request, f'{center.name} を登録しました。')
        return redirect('support_center_list')

    context = {
        'action': 'create',
        'area_choices': RegionalSupportCenter.AREA_CHOICES,
    }
    return render(request, 'clients/support_center_form.html', context)


@login_required
def support_center_edit(request, pk):
    """地域包括支援センター編集"""
    # 管理者権限チェック
    if not request.user.is_staff and not request.user.is_superuser:
        messages.error(request, 'このページにアクセスする権限がありません。')
        return redirect('client_list')

    center = get_object_or_404(RegionalSupportCenter, pk=pk)

    if request.method == 'POST':
        center.name = request.POST.get('name', '').strip()
        center.office_number = request.POST.get('office_number', '').strip()
        center.postal_code = request.POST.get('postal_code', '').strip()
        center.address = request.POST.get('address', '').strip()
        center.phone = request.POST.get('phone', '').strip()
        center.area = request.POST.get('area', 'other')

        if not center.name or not center.office_number:
            messages.error(request, '事業所名と事業所番号は必須です。')
            return redirect('support_center_edit', pk=pk)

        # 事業所番号の重複チェック（自身以外）
        if RegionalSupportCenter.objects.filter(office_number=center.office_number).exclude(pk=pk).exists():
            messages.error(request, 'この事業所番号は既に登録されています。')
            return redirect('support_center_edit', pk=pk)

        center.save()
        messages.success(request, f'{center.name} を更新しました。')
        return redirect('support_center_list')

    context = {
        'center': center,
        'action': 'edit',
        'area_choices': RegionalSupportCenter.AREA_CHOICES,
    }
    return render(request, 'clients/support_center_form.html', context)


@login_required
def support_center_delete(request, pk):
    """地域包括支援センター削除"""
    # 管理者権限チェック
    if not request.user.is_staff and not request.user.is_superuser:
        messages.error(request, 'このページにアクセスする権限がありません。')
        return redirect('client_list')

    center = get_object_or_404(RegionalSupportCenter, pk=pk)

    if request.method == 'POST':
        name = center.name
        center.delete()
        messages.success(request, f'{name} を削除しました。')
        return redirect('support_center_list')

    return redirect('support_center_list')


@login_required
def support_center_api(request):
    """地域包括支援センターのAPI（書類作成画面用）"""
    from django.db.models import Case, When, IntegerField

    # 地域の優先順位を定義
    area_order = Case(
        When(area='tsu', then=1),
        When(area='suzuka', then=2),
        When(area='kameyama', then=3),
        When(area='matsusaka', then=4),
        When(area='ise', then=5),
        When(area='yokkaichi', then=6),
        When(area='kuwana', then=7),
        When(area='iga', then=8),
        When(area='nabari', then=9),
        When(area='inabe', then=10),
        When(area='kumano', then=11),
        When(area='owase', then=12),
        When(area='toba', then=13),
        When(area='other', then=14),
        output_field=IntegerField(),
    )

    centers = RegionalSupportCenter.objects.filter(is_active=True).order_by(area_order, 'name')
    data = []
    for center in centers:
        data.append({
            'id': center.id,
            'name': center.name,
            'office_number': center.office_number,
            'postal_code': center.postal_code,
            'address': center.address,
            'phone': center.phone,
            'area': center.area,
        })
    return JsonResponse(data, safe=False)
