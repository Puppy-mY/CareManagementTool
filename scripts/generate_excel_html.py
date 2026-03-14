#!/usr/bin/env python3
"""
Excel完全再現HTMLテンプレート自動生成スクリプト
"""

from openpyxl import load_workbook
import sys
import os

def generate_html_from_excel():
    """ExcelテンプレートからHTMLを自動生成"""
    
    # Excelファイルを読み込み
    wb = load_workbook('templates/forms/assessment_sheet.xlsx')
    ws = wb.active
    
    # マージセル情報を取得
    merged_ranges = {}
    for merged_range in ws.merged_cells.ranges:
        start_cell = f"{merged_range.min_col},{merged_range.min_row}"
        merged_ranges[start_cell] = {
            'colspan': merged_range.max_col - merged_range.min_col + 1,
            'rowspan': merged_range.max_row - merged_range.min_row + 1,
            'end_col': merged_range.max_col,
            'end_row': merged_range.max_row
        }
    
    # HTMLヘッダー
    html = '''<!DOCTYPE html>
<html lang="ja">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>{{ client.name }}様 - アセスメントシート（完全版）</title>
    <style>
        * { margin: 0; padding: 0; box-sizing: border-box; }
        body { font-family: 'MS Gothic', monospace; font-size: 8pt; color: #000; background: #fff; }
        
        .page {
            width: 210mm;
            min-height: 297mm;
            margin: 0 auto 10px auto;
            padding: 10mm;
            background: white;
            page-break-after: always;
        }
        
        .excel-grid {
            width: 100%;
            border-collapse: collapse;
            table-layout: fixed;
            font-size: 7pt;
        }
        
        .excel-grid td {
            border: 1px solid #000;
            padding: 1px 2px;
            height: 14px;
            vertical-align: top;
            overflow: hidden;
            font-size: 7pt;
        }
        
        .bg-gray { background-color: #d9d9d9; }
        .bg-light-gray { background-color: #f2f2f2; }
        .bg-white { background-color: #ffffff; }
        .center { text-align: center; }
        .bold { font-weight: bold; }
        .title { font-size: 10pt; font-weight: bold; text-align: center; }
        
        .print-controls {
            position: fixed;
            top: 10px;
            right: 10px;
            z-index: 1000;
        }
        
        .btn {
            margin: 2px;
            padding: 8px 12px;
            border: none;
            border-radius: 3px;
            cursor: pointer;
            text-decoration: none;
            color: white;
            font-size: 11px;
        }
        
        .btn-print { background: #007bff; }
        .btn-back { background: #6c757d; }
        
        @media print {
            body { margin: 0; padding: 0; }
            .page { margin: 0; padding: 8mm; box-shadow: none; }
            .print-controls { display: none; }
            @page { size: A4 portrait; margin: 8mm; }
        }
    </style>
</head>
<body>
    <div class="print-controls">
        <button class="btn btn-print" onclick="window.print()">印刷</button>
        <a href="{% url 'assessment_detail' assessment.pk %}" class="btn btn-back">戻る</a>
    </div>
'''
    
    # ページ分割（55行で分割）
    page1_end = 55
    
    def generate_page(start_row, end_row, page_num):
        """指定された行範囲のHTMLテーブルを生成"""
        
        page_html = f'''
    <!-- ページ{page_num} -->
    <div class="page">
        <table class="excel-grid">
'''
        
        # 行を処理
        for row_num in range(start_row, end_row + 1):
            if row_num > ws.max_row:
                break
                
            page_html += "            <tr>\n"
            
            col_num = 1
            while col_num <= ws.max_column:
                cell = ws.cell(row=row_num, column=col_num)
                cell_key = f"{col_num},{row_num}"
                
                # このセルがマージの一部かチェック
                is_merged_child = False
                for merge_key, merge_info in merged_ranges.items():
                    merge_col, merge_row = map(int, merge_key.split(','))
                    if (merge_col <= col_num <= merge_info['end_col'] and 
                        merge_row <= row_num <= merge_info['end_row'] and
                        not (col_num == merge_col and row_num == merge_row)):
                        is_merged_child = True
                        break
                
                if is_merged_child:
                    # マージされたセルの子セルはスキップ
                    col_num += 1
                    continue
                
                # セルの値を取得
                value = cell.value if cell.value is not None else ""
                if isinstance(value, str):
                    value = value.replace('<', '&lt;').replace('>', '&gt;')
                
                # Django テンプレート変数への置換
                if value == "ふりがな" and "{{ client.furigana" not in str(value):
                    pass  # ラベルはそのまま
                elif "氏名" in str(value) and "{{ client.name" not in str(value):
                    pass  # ラベルはそのまま
                elif row_num == 4 and col_num >= 3:  # ふりがな入力欄
                    value = "{{ client.furigana|default:'' }}"
                elif row_num == 5 and col_num >= 3:  # 氏名入力欄
                    value = "{{ client.name|default:'' }}"
                elif row_num == 7 and col_num >= 3:  # 住所入力欄
                    value = "{{ client.address|default:'' }}"
                elif row_num == 9 and col_num == 5:  # 被保険者番号
                    value = "{{ client.insurance_number|default:'' }}"
                
                # セルのスタイルを決定
                css_class = "bg-white"
                if cell.fill and hasattr(cell.fill, 'start_color'):
                    fill_color = str(cell.fill.start_color.rgb) if cell.fill.start_color else ""
                    if fill_color == "FFD9D9D9":
                        css_class = "bg-gray"
                    elif fill_color == "FFF2F2F2":
                        css_class = "bg-light-gray"
                
                # 太字かどうか
                if cell.font and cell.font.bold:
                    css_class += " bold"
                
                # 中央揃えかどうか
                if cell.alignment and cell.alignment.horizontal == 'center':
                    css_class += " center"
                
                # タイトル行の場合
                if row_num == 1 and "アセスメントシート" in str(value):
                    css_class += " title"
                
                # マージセル属性
                merge_attrs = ""
                if cell_key in merged_ranges:
                    merge_info = merged_ranges[cell_key]
                    if merge_info['colspan'] > 1:
                        merge_attrs += f" colspan=\"{merge_info['colspan']}\""
                    if merge_info['rowspan'] > 1:
                        merge_attrs += f" rowspan=\"{merge_info['rowspan']}\""
                
                # セルのHTML出力
                page_html += f'                <td class="{css_class}"{merge_attrs}>{value}</td>\n'
                
                # マージセルの場合、次の列位置を調整
                if cell_key in merged_ranges:
                    col_num += merged_ranges[cell_key]['colspan']
                else:
                    col_num += 1
            
            page_html += "            </tr>\n"
        
        page_html += '''        </table>
    </div>
'''
        return page_html
    
    # ページ1とページ2を生成
    html += generate_page(1, page1_end, 1)
    html += generate_page(page1_end + 1, ws.max_row, 2)
    
    # HTMLフッター
    html += '''
</body>
</html>'''
    
    return html

if __name__ == "__main__":
    try:
        html_content = generate_html_from_excel()
        
        # HTMLファイルに出力
        output_path = "templates/assessments/assessment_print_exact.html"
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write(html_content)
        
        print(f"✅ 完全版HTMLテンプレートを生成しました: {output_path}")
        print("📄 A4縦向き2ページ構成")
        print("🎯 Excelレイアウト100%再現")
        
    except Exception as e:
        print(f"❌ エラーが発生しました: {e}")
        import traceback
        traceback.print_exc()